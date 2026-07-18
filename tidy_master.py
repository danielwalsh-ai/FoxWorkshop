"""Tidy the wagon-master DAILY sheet after filling days.

Normalises every visible date column to a clean "golden" template column
(borders + number format), blanks stray £0 on wagon rows (the master convention
is blank, not 0, for a wagon that didn't earn), and widens the date columns so
figures never show as ###.  All at raw-XML level so the sheet's SharePoint
external links and chart sheets survive.

    from tidy_master import tidy
    tidy(master_path)            # in place
    tidy(master_path, out_path)  # to a new file
"""
import re
import zipfile
import datetime as dt
from lxml import etree
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from wagon_master_fill import VEHICLE_ROWS

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
def _q(t): return '{' + NS + '}' + t
FIRST_VISIBLE = 38          # date columns before this are the hidden monthly block
CHECK_ROWS = (124, 128, 129, 176, 177, 178)   # rows that were commonly mis-formatted


def tidy(src, out=None, width=13):
    out = out or src
    ws = openpyxl.load_workbook(src)['DAILY']
    wsv = openpyxl.load_workbook(src, data_only=True)['DAILY']
    datecols = {wsv.cell(2, c).value.date(): c for c in range(3, wsv.max_column + 2)
                if isinstance(wsv.cell(2, c).value, dt.datetime)}
    vis = sorted(c for c in datecols.values() if c >= FIRST_VISIBLE)
    if not vis:
        return out

    def _has_borders(col, row):
        b = ws.cell(row, col).border
        return all([b.left.style, b.right.style, b.top.style, b.bottom.style])
    # golden = first visible date column fully bordered on the tell-tale rows
    golden = next((c for c in vis if all(_has_borders(c, r) for r in CHECK_ROWS)), vis[0])

    z = zipfile.ZipFile(src)
    wbxml = z.read('xl/workbook.xml').decode()
    rels = z.read('xl/_rels/workbook.xml.rels').decode()
    rid = re.search(r'<sheet[^>]*name="DAILY"[^>]*r:id="(rId\d+)"', wbxml).group(1)
    tgt = 'xl/' + re.search(rf'<Relationship[^>]*Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1)
    tree = etree.fromstring(z.read(tgt))
    sd = tree.find(_q('sheetData'))
    rowmap = {int(r.get('r')): r for r in sd.findall(_q('row'))}

    def _cell(rowel, ref):
        for c in rowel.findall(_q('c')):
            if c.get('r') == ref:
                return c
        return None

    def _ensure(rowel, colidx, rownum):
        ref = f'{get_column_letter(colidx)}{rownum}'
        c = _cell(rowel, ref)
        if c is None:
            c = etree.Element(_q('c')); c.set('r', ref)
            for ex in rowel.findall(_q('c')):
                if column_index_from_string(re.match(r'([A-Z]+)', ex.get('r')).group(1)) > colidx:
                    ex.addprevious(c); break
            else:
                rowel.append(c)
        return c

    gL = get_column_letter(golden)
    for rownum, rowel in rowmap.items():
        if rownum > 205:
            continue
        gs = _cell(rowel, f'{gL}{rownum}')
        gstyle = gs.get('s') if gs is not None else None
        for c in vis:
            if c == golden:
                continue
            cc = _ensure(rowel, c, rownum)
            if gstyle is not None:
                cc.set('s', gstyle)                       # match template borders/format
            if rownum in VEHICLE_ROWS:                    # blank stray £0 on wagon rows
                v = cc.find(_q('v'))
                if v is not None and cc.get('t') in (None, 'n') \
                        and v.text not in (None, '') and float(v.text) == 0:
                    cc.remove(v)

    # widen: cover all visible date columns to a generous, uniform width
    cols = tree.find(_q('cols'))
    if cols is not None:
        keep = [e for e in cols.findall(_q('col')) if int(e.get('min')) < FIRST_VISIBLE]
        for e in list(cols):
            cols.remove(e)
        for e in keep:
            cols.append(e)
        wcol = etree.SubElement(cols, _q('col'))
        wcol.set('min', str(FIRST_VISIBLE)); wcol.set('max', '500')
        wcol.set('width', str(width)); wcol.set('customWidth', '1')

    data = etree.tostring(tree, xml_declaration=True, encoding='UTF-8', standalone=True)
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zo:
        for it in z.infolist():
            zo.writestr(it, data if it.filename == tgt else z.read(it.filename))
    z.close()
    return out


if __name__ == '__main__':
    import sys
    tidy(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
    print("tidied", sys.argv[1])
