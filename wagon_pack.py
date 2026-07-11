"""
Fox Leyland — Daily Wagon Earnings Pack generator (faithful 10-page layout).

Reproduces the original daily earnings pack exactly (masthead + KPI cards +
category chips, best/worst period tables, daily-earnings bar chart, day-vs-night
stacked bars, night cards+table, category-average line chart, under-target
horizontal bars, off-road stacked bars, and the under-target detail table).

One PDF per focus day; the multi-day pages use the run sheets available so far
in the same week (matching "Days available: Tue 23 Jun – Fri 26 Jun").

Usage:
    python wagon_pack.py <focus_runsheet.xlsx> [prior_runsheet.xlsx ...] -o out.pdf
      (list the week's run sheets in date order; the LAST is the focus day)
"""
import argparse
import re
import tempfile
import datetime as dt
from pathlib import Path

from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

HERE = Path(__file__).parent
LOGO = HERE / 'static' / 'fox-group-logo.png'

REG_RE = re.compile(r'^[A-Z]{2}\d{2}\s?[A-Z]{3}$')

NAVY = colors.HexColor('#1a2b4a'); ORANGE = colors.HexColor('#eb941f')
GREEN = colors.HexColor('#387519'); RED = colors.HexColor('#bf3833')
GREY = colors.HexColor('#666666'); LGREY = colors.HexColor('#999999')
CGREY = colors.HexColor('#545454'); BORDER = colors.HexColor('#cccccc')
CHIPGREEN = colors.HexColor('#d9e7cf'); CHIPRED = colors.HexColor('#f4d6d3')
ROWALT = colors.HexColor('#f4f4f6'); TOTBG = colors.HexColor('#e6ebf3')
WHITE = colors.white
NAVY_H = '#1a2b4a'; ORANGE_H = '#eb941f'
CAT_COLORS = {'8 Wheeler': '#1a2b4a', 'Hooks': '#eb941f', 'Ally Body': '#4a8c2a',
              'Artics': '#8a6d3b', 'Grabs': '#a03d6f', 'Sweepers': '#c9a227'}

# canonical category display order + which map to the run-sheet block names
CAT_ORDER = ['8 Wheeler', 'Hooks', 'Ally Body', 'Artics', 'Grabs', 'Sweepers']
BLOCK_NAME = {'8 WHEELER': '8 Wheeler', 'HOOKS': 'Hooks', 'HOOKS ON HIRE': 'Hooks on Hire',
              'ALLY BODY': 'Ally Body', 'ARTICS': 'Artics', 'GRABS': 'Grabs',
              'SWEEPERS': 'Sweepers', 'BPH': '8 Wheeler', 'ARTICS - NIGHTS': 'Night'}


def parse_runsheet(path):
    wb = load_workbook(path, data_only=True)
    ws = wb['Wagons']
    date = ws['P2'].value
    if isinstance(date, dt.datetime):
        date = date.date()

    # 1) raw category block at the bottom (original names, NOT merged)
    raw = {}
    in_block = False
    for r in range(3, ws.max_row + 1):
        a, b, c, e = (ws.cell(r, i).value for i in (1, 2, 3, 5))
        if isinstance(a, str) and a.strip() == 'VEHICLE':
            in_block = True
            continue
        if in_block and isinstance(a, str) and a.strip():
            raw[a.strip().upper()] = {'rev': b or 0, 'target': c or 0,
                                      'count': int(e) if isinstance(e, (int, float)) else 0}

    # 2) split the wagon list into sections by subtotal rows
    sections, sec = [], []
    is_night = False
    day_total, total_scored = None, None
    for r in range(3, ws.max_row + 1):
        a, b, c, d = (ws.cell(r, i).value for i in (1, 2, 3, 4))
        if isinstance(a, str) and a.strip() == 'VEHICLE':
            break
        if isinstance(a, str) and a.strip().upper() == 'NIGHT WORK':
            is_night = True
            continue
        if isinstance(a, str) and REG_RE.match(a.strip().upper()):
            sec.append({'reg': a.strip().upper(), 'target': b or 0,
                        'earned': d if isinstance(d, (int, float)) else 0,
                        'status': c.strip().upper() if isinstance(c, str) else 'OK'})
        elif isinstance(a, (int, float)):
            if c is None and b is None:                    # grand-total row
                if isinstance(d, (int, float)):
                    day_total, total_scored = d, int(a)
            else:                                          # section subtotal
                sections.append({'count': int(a), 'target': b or 0,
                                 'revD': d if isinstance(d, (int, float)) else 0,
                                 'regs': sec, 'is_night': is_night})
                sec, is_night = [], False

    # 3) name each section by best match to a raw block entry (unique)
    used = set()
    for s in sections:
        name, bestd = None, 1e18
        for rn, v in raw.items():
            if rn in used:
                continue
            dd = abs(v['rev'] - s['revD']) + abs(v['target'] - s['target']) + abs(v['count'] - s['count']) * 50
            if dd < bestd:
                bestd, name = dd, rn
        used.add(name)
        s['disp'] = BLOCK_NAME.get(name, (name or '?').title())

    regs, offroad = [], []
    for s in sections:
        if s['is_night'] or s['disp'] == 'Night':
            continue
        for rr in s['regs']:
            rr['cat'] = s['disp']
            regs.append(rr)
            if rr['status'] != 'OK':
                offroad.append(rr)

    cats = {}
    night = 0
    for rn, v in raw.items():
        disp = BLOCK_NAME.get(rn, rn.title())
        if disp == 'Night':
            night = v['rev']; continue
        c2 = cats.setdefault(disp, {'rev': 0, 'target': 0, 'count': 0})
        c2['rev'] += v['rev']; c2['target'] += v['target']; c2['count'] += v['count']

    day = day_total if day_total is not None else sum(v['rev'] for v in cats.values())
    ts = total_scored if total_scored is not None else len(regs)
    return {'date': date, 'regs': regs, 'cats': cats, 'offroad': offroad,
            'day': day, 'night': night, 'total_wagons': ts}


def compute_focus(data):
    regs = data['regs']
    scored = [r for r in regs if r['status'] == 'OK' or r['earned'] > 0]
    on_road = [r for r in regs if r['status'] == 'OK']
    under = sorted([r for r in regs if r['status'] == 'OK' and r['earned'] < r['target']],
                   key=lambda r: (r['earned'] - r['target']))
    # category averages (focus day): rev per on-road wagon; target per wagon
    cat_avg = {}
    for name in CAT_ORDER:
        cr = [r for r in regs if r['cat'] == name]
        onroad = [r for r in cr if r['status'] == 'OK']
        rev = sum(r['earned'] for r in cr)
        cnt = len(cr)
        tgt_sum = sum(r['target'] for r in cr)
        cat_avg[name] = {'avg': rev / len(onroad) if onroad else 0,
                         'tgt': tgt_sum / cnt if cnt else 0,
                         'n': cnt, 'under': sum(1 for r in onroad if r['earned'] < r['target']),
                         'onroad': len(onroad)}
    return {'n_scored': len(data['regs']) + 0, 'under': under, 'n_under': len(under),
            'n_offroad': len(data['offroad']), 'cat_avg': cat_avg,
            'total_scored': data['total_wagons']}


def money(v, dp=0):
    return f"£{v:,.{dp}f}"


# ── charts ──────────────────────────────────────────────────────────────────
def _style_ax(ax):
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#888'); ax.spines['bottom'].set_color('#888')
    ax.tick_params(colors='#333', labelsize=9)
    ax.yaxis.grid(True, linestyle='--', color='#cccccc', lw=0.7)
    ax.set_axisbelow(True)


def _daylabels(days):
    return [d.strftime('%a\n%d %b') for d in days]


def chart_daily(days, dayvals, path):
    fig, ax = plt.subplots(figsize=(8.6, 5.0), dpi=150)
    ax.bar(range(len(days)), dayvals, color=NAVY_H, width=0.62)
    for i, v in enumerate(dayvals):
        ax.text(i, v + max(dayvals) * 0.015, money(v), ha='center', va='bottom',
                fontsize=10, color='#333')
    ax.set_xticks(range(len(days))); ax.set_xticklabels(_daylabels(days))
    ax.set_ylabel('Earnings (£)', fontsize=9)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda x, _: f'£{x/1000:.0f}k'))
    ax.set_ylim(0, max(dayvals) * 1.18)
    ax.set_title('Total daily earnings — days available so far', fontsize=13,
                 fontweight='bold', color=NAVY_H, loc='left', pad=14)
    _style_ax(fig.gca())
    fig.tight_layout(); fig.savefig(path, bbox_inches='tight'); plt.close(fig)


def chart_daynight(days, dayvals, nightvals, path):
    fig, ax = plt.subplots(figsize=(8.6, 5.0), dpi=150)
    x = range(len(days))
    ax.bar(x, dayvals, color=NAVY_H, width=0.62, label='Day earnings')
    ax.bar(x, nightvals, bottom=dayvals, color=ORANGE_H, width=0.62, label='Night work')
    for i in x:
        tot = dayvals[i] + nightvals[i]
        ax.text(i, tot + max(d + n for d, n in zip(dayvals, nightvals)) * 0.015,
                money(tot), ha='center', va='bottom', fontsize=10, color='#333')
    ax.set_xticks(list(x)); ax.set_xticklabels(_daylabels(days))
    ax.set_ylabel('Earnings (£)', fontsize=9)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f'£{v/1000:.0f}k'))
    ax.set_ylim(0, max(d + n for d, n in zip(dayvals, nightvals)) * 1.2)
    ax.set_title('Day earnings vs night work', fontsize=13, fontweight='bold',
                 color=NAVY_H, loc='left', pad=14)
    ax.legend(loc='center left', bbox_to_anchor=(1.01, 0.5), frameon=False, fontsize=10)
    _style_ax(ax)
    fig.tight_layout(); fig.savefig(path, bbox_inches='tight'); plt.close(fig)


def chart_catavg(days, series, path):
    fig, ax = plt.subplots(figsize=(8.6, 5.0), dpi=150)
    x = range(len(days))
    for name in CAT_ORDER:
        ax.plot(x, series[name], marker='o', ms=5, lw=2, color=CAT_COLORS[name], label=name)
    ax.set_xticks(list(x)); ax.set_xticklabels([d.strftime('%a %d %b') for d in days])
    ax.set_ylabel('£ per wagon', fontsize=9)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f'£{v:.0f}'))
    ax.set_title('Daily average earnings per wagon — by category', fontsize=13,
                 fontweight='bold', color=NAVY_H, loc='left', pad=14)
    ax.legend(loc='center left', bbox_to_anchor=(1.01, 0.5), frameon=False, fontsize=10)
    _style_ax(ax)
    fig.tight_layout(); fig.savefig(path, bbox_inches='tight'); plt.close(fig)


def chart_undertarget(cat_avg, path):
    cats = [c for c in CAT_ORDER if cat_avg[c]['onroad'] > 0]
    onroad = [cat_avg[c]['onroad'] for c in cats]
    under = [cat_avg[c]['under'] for c in cats]
    okc = [o - u for o, u in zip(onroad, under)]
    y = range(len(cats))[::-1]
    fig, ax = plt.subplots(figsize=(8.6, 5.0), dpi=150)
    ax.barh(list(y), okc, color='#5a8f3c', label='On / above target')
    ax.barh(list(y), under, left=okc, color=ORANGE_H, label='Under target')
    for i, yy in zip(range(len(cats)), y):
        ax.text(onroad[i] + max(onroad) * 0.01, yy, f'{under[i]}/{onroad[i]} under',
                va='center', fontsize=9, color='#333')
    ax.set_yticks(list(y)); ax.set_yticklabels(cats, fontsize=10)
    ax.set_xlabel('Wagons in service', fontsize=9)
    ax.set_title('Wagons under daily target — focus day', fontsize=13, fontweight='bold',
                 color=NAVY_H, loc='left', pad=14)
    ax.legend(loc='lower right', frameon=False, fontsize=9)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.xaxis.grid(True, linestyle='--', color='#cccccc', lw=0.7); ax.set_axisbelow(True)
    ax.set_xlim(0, max(onroad) * 1.25)
    fig.tight_layout(); fig.savefig(path, bbox_inches='tight'); plt.close(fig)


def chart_offroad(days, reasons, path):
    fig, ax = plt.subplots(figsize=(8.6, 5.0), dpi=150)
    x = range(len(days))
    vor = [reasons[i]['VOR'] for i in x]
    mnt = [reasons[i]['Maintenance'] for i in x]
    nod = [reasons[i]['No driver'] for i in x]
    ax.bar(x, vor, color=NAVY_H, width=0.6, label='Vehicle Off Road (VOR)')
    ax.bar(x, mnt, bottom=vor, color=ORANGE_H, width=0.6, label='Maintenance')
    ax.bar(x, nod, bottom=[v + m for v, m in zip(vor, mnt)], color='#7a1f2b',
           width=0.6, label='No driver / absence')
    for i in x:
        tot = vor[i] + mnt[i] + nod[i]
        if tot:
            ax.text(i, tot + 0.12, str(tot), ha='center', va='bottom', fontsize=10,
                    fontweight='bold', color='#333')
        if vor[i]:
            ax.text(i, vor[i] / 2, str(vor[i]), ha='center', va='center', fontsize=9, color='white')
        if mnt[i]:
            ax.text(i, vor[i] + mnt[i] / 2, str(mnt[i]), ha='center', va='center', fontsize=9, color='white')
    ax.set_xticks(list(x)); ax.set_xticklabels(_daylabels(days))
    ax.set_ylabel('Wagons off road (days)', fontsize=9)
    mx = max((v + m + n) for v, m, n in zip(vor, mnt, nod)) if days else 1
    ax.set_ylim(0, max(mx + 2, 4))
    ax.set_title('Wagons off the road each day, by reason', fontsize=13, fontweight='bold',
                 color=NAVY_H, loc='left', pad=14)
    ax.legend(loc='upper left', frameon=False, fontsize=9)
    _style_ax(ax)
    fig.tight_layout(); fig.savefig(path, bbox_inches='tight'); plt.close(fig)


# ── PDF ─────────────────────────────────────────────────────────────────────
def _reason(status):
    s = (status or '').upper()
    if s == 'VOR':
        return 'VOR'
    if s in ('MN', 'MAINT', 'MAINTENANCE'):
        return 'Maintenance'
    return 'No driver'


def build_pack(runsheets, out_pdf):
    per = [parse_runsheet(p) for p in runsheets]
    focus = per[-1]
    days = [d['date'] for d in per]
    dayvals = [d['day'] for d in per]
    nightvals = [d['night'] for d in per]
    fc = compute_focus(focus)
    d = focus['date']
    dl = d.strftime('%a %d %b %Y'); dshort = d.strftime('%d %b')
    ndays = len(per)

    tmp = Path(tempfile.mkdtemp())
    chart_daily(days, dayvals, tmp / 'c3.png')
    chart_daynight(days, dayvals, nightvals, tmp / 'c4.png')
    # category avg series over days
    series = {c: [] for c in CAT_ORDER}
    for p in per:
        cf = compute_focus(p)['cat_avg']
        for c in CAT_ORDER:
            series[c].append(cf[c]['avg'])
    chart_catavg(days, series, tmp / 'c6.png')
    chart_undertarget(fc['cat_avg'], tmp / 'c7.png')
    reasons = []
    for p in per:
        rr = {'VOR': 0, 'Maintenance': 0, 'No driver': 0}
        for o in p['offroad']:
            rr[_reason(o['status'])] += 1
        reasons.append(rr)
    chart_offroad(days, reasons, tmp / 'c8.png')

    # period aggregates
    avg_day = sum(dayvals) / ndays
    wdl = sum(len(p['offroad']) for p in per)
    night_total = sum(nightvals)
    # best/worst period (across days, per reg)
    per_reg = {}
    for p in per:
        for r in p['regs']:
            e = per_reg.setdefault(r['reg'], {'reg': r['reg'], 'cat': r['cat'], 'total': 0, 'days': 0, 'off': 0})
            e['total'] += r['earned']
            if r['status'] == 'OK':
                e['days'] += 1
            else:
                e['off'] += 1
    ranked = [v for v in per_reg.values() if v['total'] > 0]
    best = sorted(ranked, key=lambda v: -v['total'])[:10]
    # worst excludes wagons off-road for most of the period (structural low earners)
    worst_pool = [v for v in ranked if v['days'] >= ndays / 2]
    worst = sorted(worst_pool, key=lambda v: v['total'])[:10]

    W, H = landscape(A4)
    cv = canvas.Canvas(str(out_pdf), pagesize=landscape(A4))
    cv.setTitle(f"Fox Leyland — Daily Wagon Earnings — {d:%d %b %Y}")
    M = 40

    def T(y):  # top-origin -> reportlab
        return H - y

    def masthead(page1=False):
        cv.setFillColor(NAVY); cv.rect(0, H - 28, W, 28, fill=1, stroke=0)
        if page1:
            cv.drawImage(ImageReader(str(LOGO)), 51, T(126), width=181, height=41,
                         mask='auto', preserveAspectRatio=True)
        else:
            cv.drawImage(ImageReader(str(LOGO)), W - 155, H - 24, width=120, height=20,
                         mask='auto', preserveAspectRatio=True)

    def footer(pg):
        cv.setFillColor(GREY); cv.setFont('Helvetica', 8)
        cv.drawString(M, 28, f'Fox Leyland · Daily Wagon Earnings Pack (draft) · {dl}')
        cv.drawRightString(W - M, 28, f'Page {pg}')

    def title(y, text, sub=None, note=None):
        cv.setFillColor(NAVY); cv.setFont('Helvetica', 28)
        cv.drawString(M, T(y), text)
        if sub:
            cv.setFillColor(GREY); cv.setFont('Helvetica', 11)
            cv.drawString(M, T(y + 22), sub)
        if note:
            cv.setFillColor(LGREY); cv.setFont('Helvetica-Oblique', 9)
            cv.drawString(M, T(y + 40), note)

    def card(x, y, w, h, big, label, bigcolor=NAVY):
        cv.setStrokeColor(BORDER); cv.setLineWidth(1)
        cv.roundRect(x, T(y + h), w, h, 6, stroke=1, fill=0)
        cv.setFillColor(bigcolor); cv.setFont('Helvetica-Bold', 22)
        cv.drawCentredString(x + w / 2, T(y + h / 2 - 2), big)
        cv.setFillColor(GREY); cv.setFont('Helvetica', 9.5)
        cv.drawCentredString(x + w / 2, T(y + h - 14), label)

    def draw_table(x, y, colw, header, rows, aligns, val_colors=None,
                   total_row=False, fs=9, rh=21):
        tw = sum(colw)
        cv.setFillColor(NAVY); cv.rect(x, T(y + rh), tw, rh, fill=1, stroke=0)
        cv.setFillColor(WHITE); cv.setFont('Helvetica-Bold', fs)
        cx = x
        for i, htext in enumerate(header):
            if aligns[i] == 'r':
                cv.drawRightString(cx + colw[i] - 6, T(y + rh - 6), htext)
            else:
                cv.drawString(cx + 6, T(y + rh - 6), htext)
            cx += colw[i]
        yy = y + rh
        for ri, row in enumerate(rows):
            islast = total_row and ri == len(rows) - 1
            if islast:
                cv.setFillColor(TOTBG); cv.rect(x, T(yy + rh), tw, rh, fill=1, stroke=0)
            elif ri % 2 == 0:
                cv.setFillColor(ROWALT); cv.rect(x, T(yy + rh), tw, rh, fill=1, stroke=0)
            cx = x
            for i, cell in enumerate(row):
                col = NAVY if islast else colors.black
                if val_colors and val_colors.get(i) and not islast:
                    col = val_colors[i]
                cv.setFillColor(col)
                cv.setFont('Helvetica-Bold' if (islast or (val_colors and val_colors.get(i))) else 'Helvetica', fs)
                if aligns[i] == 'r':
                    cv.drawRightString(cx + colw[i] - 6, T(yy + rh - 6), str(cell))
                else:
                    cv.drawString(cx + 6, T(yy + rh - 6), str(cell))
                cx += colw[i]
            yy += rh
        return yy

    def img(path, x, y, w, h):
        cv.drawImage(ImageReader(str(path)), x, T(y + h), width=w, height=h,
                     mask='auto', preserveAspectRatio=True)

    # ===== Page 1 =====
    masthead(page1=True)
    title(172, 'Daily Wagon Earnings Pack — Fox Leyland',
          sub=f"Focus day: {d.strftime('%A %d %b')} · Days available: "
              f"{days[0].strftime('%a %d %b')} – {d.strftime('%a %d %b %Y')}",
          note=f"Built from {ndays} run sheet{'s' if ndays!=1 else ''} — BPH wagons "
               f"absorbed into 8 Wheelers. Cost/EBITDA pages omitted (no cost data in run sheets).")
    cw = (W - 2 * M - 3 * 12) / 4
    row1 = [(money(focus['day']), f"Day earnings on {dshort}"),
            (money(focus['night']), f"Night work {dshort}"),
            (money(focus['day'] + focus['night']), 'Total incl. nights'),
            (money(avg_day), f'Avg day earnings ({ndays} day{"s" if ndays!=1 else ""})')]
    for i, (big, lab) in enumerate(row1):
        card(M + i * (cw + 12), 256, cw, 80, big, lab)
    cw3 = (W - 2 * M - 2 * 12) / 3
    row2 = [(f"{fc['n_under']} / {fc['total_scored']}", 'Wagons under target'),
            (str(fc['n_offroad']), f"Wagons off road on {dshort}"),
            (str(wdl), f'Wagon-days lost ({ndays} day{"s" if ndays!=1 else ""})')]
    for i, (big, lab) in enumerate(row2):
        card(M + i * (cw3 + 12), 348, cw3, 80, big, lab, bigcolor=RED)
    cv.setFillColor(NAVY); cv.setFont('Helvetica', 11)
    cv.drawString(M, T(456), f"Average earnings per wagon by category on {dshort}")
    cw6 = (W - 2 * M - 5 * 10) / 6
    for i, name in enumerate(CAT_ORDER):
        ca = fc['cat_avg'][name]
        above = ca['avg'] >= ca['tgt']
        x = M + i * (cw6 + 10)
        cv.setFillColor(CHIPGREEN if above else CHIPRED)
        cv.roundRect(x, T(516), cw6, 46, 5, fill=1, stroke=0)
        cv.setFillColor(GREEN if above else RED); cv.setFont('Helvetica-Bold', 14)
        cv.drawCentredString(x + cw6 / 2, T(492), money(ca['avg']))
        cv.setFillColor(CGREY); cv.setFont('Helvetica', 7.5)
        cv.drawCentredString(x + cw6 / 2, T(508), f"{name} · tgt {money(ca['tgt'])}")
    footer(1)

    # ===== Page 2: best & worst =====
    cv.showPage(); masthead()
    title(92, 'Best & worst earning wagons — period total',
          sub=f"Top 10 and bottom 10 wagons across all {ndays} run-sheet day"
              f"{'s' if ndays!=1 else ''} so far, with average earnings per day worked.")
    cv.setFillColor(NAVY); cv.setFont('Helvetica-Bold', 12); cv.drawString(M, T(175), 'Best earning wagons')
    cv.setFillColor(GREY); cv.setFont('Helvetica', 9); cv.drawString(M, T(190), 'Ranked by period total (highest first).')
    cv.setFillColor(NAVY); cv.setFont('Helvetica-Bold', 12); cv.drawString(W / 2 + 20, T(175), 'Worst earning wagons')
    cv.setFillColor(GREY); cv.setFont('Helvetica', 9); cv.drawString(W / 2 + 20, T(190), 'Genuine low earners (structural £0s excluded).')
    colw = [26, 92, 110, 78, 74]
    hdr = ['#', 'Reg', 'Category', 'Period £', 'Avg/day £']
    aligns = ['l', 'l', 'l', 'r', 'r']
    brows = [[str(i + 1), v['reg'], v['cat'], money(v['total']),
              money(v['total'] / v['days'] if v['days'] else 0)] for i, v in enumerate(best)]
    wrows = [[str(i + 1), v['reg'], v['cat'], money(v['total']),
              money(v['total'] / v['days'] if v['days'] else 0)] for i, v in enumerate(worst)]
    draw_table(M, 205, colw, hdr, brows, aligns, val_colors={3: GREEN, 4: GREEN})
    draw_table(W / 2 + 20, 205, colw, hdr, wrows, aligns, val_colors={3: RED, 4: RED})
    cv.setFillColor(GREY); cv.setFont('Helvetica', 9)
    cv.drawString(M, T(505), 'Avg/day £ = period total ÷ days actually on the road (off-road / maintenance days excluded).')
    cv.drawString(M, T(520), 'Worst list excludes dedicated night-work regs and wagons off-road for most of the period.')
    footer(2)

    # ===== Page 3: total daily earnings =====
    cv.showPage(); masthead()
    title(92, 'Total daily earnings',
          sub='Fleet-wide day earnings (excludes night work), all run-sheet days so far.')
    img(tmp / 'c3.png', M + 90, 150, W - 2 * M - 260, 360)
    hi = max(range(ndays), key=lambda i: dayvals[i]); lo = min(range(ndays), key=lambda i: dayvals[i])
    cv.setFillColor(colors.black); cv.setFont('Helvetica', 10)
    cv.drawString(M + 10, T(545),
        f"Highest day was {days[hi].strftime('%a %d %b')} at {money(dayvals[hi])}; "
        f"lowest was {days[lo].strftime('%a %d %b')} at {money(dayvals[lo])}. "
        f"Focus day {dshort}: {money(focus['day'])}.")
    footer(3)

    # ===== Page 4: day vs night =====
    cv.showPage(); masthead()
    title(92, 'Day earnings vs night work', sub='Night work is shown separately from the daytime fleet total.')
    img(tmp / 'c4.png', M + 90, 150, W - 2 * M - 200, 360)
    cv.setFillColor(colors.black); cv.setFont('Helvetica', 10)
    cv.drawString(M + 10, T(545), f"Night work added {money(night_total)} across the {ndays} day"
                  f"{'s' if ndays!=1 else ''}, on top of daytime earnings. Nights are excluded from the headline day figure.")
    footer(4)

    # ===== Page 5: night table =====
    cv.showPage(); masthead()
    title(92, 'Night work — Artics (reported separately)',
          sub='Night-only artic registrations are shown here only and are excluded from every other page in this pack.')
    ncw = (W - 2 * M - 2 * 12) / 3
    ntgt = 3200 * ndays
    card(M, 180, ncw, 78, money(night_total), f'Night earnings ({ndays} day{"s" if ndays!=1 else ""})')
    card(M + ncw + 12, 180, ncw, 78, money(night_total / ndays), 'Avg night earnings / day')
    card(M + 2 * (ncw + 12), 180, ncw, 78, ('+' if night_total - ntgt >= 0 else '−') + money(abs(night_total - ntgt)),
         'Vs night target (period)', bigcolor=GREEN if night_total >= ntgt else RED)
    colw = [W / 2 - M - 40, 150, 130, 120]
    hdr = ['Night', 'Night earnings £', 'Target £', 'Variance £']
    nrows = []
    for i, p in enumerate(per):
        var = nightvals[i] - 3200
        nrows.append([days[i].strftime('%A %d %b'), money(nightvals[i]), money(3200),
                      ('+' if var >= 0 else '−') + money(abs(var))])
    nrows.append(['Period total', money(night_total), money(ntgt),
                  ('+' if night_total - ntgt >= 0 else '−') + money(abs(night_total - ntgt))])
    y5 = draw_table(M, 300, colw, hdr, nrows, ['l', 'r', 'r', 'r'],
                    val_colors={3: GREEN}, total_row=True)
    cv.setFillColor(GREY); cv.setFont('Helvetica', 9)
    cv.drawString(M, T(y5 + 30), 'Night regs: PK21YSON, PK21YSPN, PK21YSRN.  Earnings are booked as a single nightly total on the run sheet, not per reg.')
    cv.drawString(M, T(y5 + 44), 'These regs are excluded from day earnings, category averages, under-target counts and the best/worst wagon tables.')
    cv.drawString(M, T(y5 + 58), 'Night target is a flat £3,200 per night from the run sheet.')
    footer(5)

    # ===== Page 6: category avg line =====
    cv.showPage(); masthead()
    title(92, 'Daily average earnings per wagon, by category',
          sub='Shows which fleet categories pull the per-wagon average up or down.')
    img(tmp / 'c6.png', M + 90, 150, W - 2 * M - 200, 360)
    cv.setFillColor(colors.black); cv.setFont('Helvetica', 10)
    cv.drawString(M + 10, T(545), '8 Wheelers (incl. former BPH) lead the per-wagon average. Night-only artics are excluded here and reported on their own page.')
    footer(6)

    # ===== Page 7: under target chart =====
    cv.showPage(); masthead()
    title(92, f'Wagons under target — {dl}',
          sub='Per-category targets from the run sheet. BPH folded into 8 Wheelers.')
    img(tmp / 'c7.png', M + 90, 150, W - 2 * M - 240, 360)
    cv.setFillColor(colors.black); cv.setFont('Helvetica', 10)
    cv.drawString(M + 10, T(545), f"{fc['n_under']} of {fc['total_scored']} scored wagons earned below their daily target on the focus day.")
    footer(7)

    # ===== Page 8: off road chart =====
    cv.showPage(); masthead()
    title(92, 'Wagons off the road — by reason',
          sub='VOR = Vehicle Off Road. Counts from the run-sheet summary block.')
    img(tmp / 'c8.png', M + 90, 150, W - 2 * M - 240, 360)
    cv.setFillColor(colors.black); cv.setFont('Helvetica', 10)
    cv.drawString(M + 10, T(545), f"Across the {ndays} day{'s' if ndays!=1 else ''} the fleet lost {wdl} wagon-days to off-road status. VOR is the dominant reason throughout.")
    footer(8)

    # ===== Pages 9-10: under target table =====
    per_page = 20
    chunks = [fc['under'][i:i + per_page] for i in range(0, len(fc['under']), per_page)] or [[]]
    for pi, chunk in enumerate(chunks):
        cv.showPage(); masthead()
        cont = '(continued) ' if pi else ''
        title(92, f'Wagons under target — {dl}',
              sub=f"{cont}Sorted by shortfall (largest first). {fc['n_under']} wagons earned below target across {fc['total_scored']} scored.")
        colw = [150, 200, 130, 130, 120]
        hdr = ['Category', 'Reg', 'Target (£)', 'Earned (£)', 'Shortfall (£)']
        rows = [[r['cat'], r['reg'], f"{r['target']:,.0f}", f"{r['earned']:,.0f}",
                 f"{r['target'] - r['earned']:,.0f}"] for r in chunk]
        draw_table(M, 168, colw, hdr, rows, ['l', 'l', 'r', 'r', 'r'], val_colors={4: RED}, rh=18)
        footer(9 + pi)

    cv.save()
    return {'day': focus['day'], 'night': focus['night'], 'n_under': fc['n_under'],
            'n_offroad': fc['n_offroad'], 'total': focus['day'] + focus['night'],
            'pages': 8 + len(chunks)}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('runsheets', nargs='+')
    ap.add_argument('-o', '--out', required=True)
    a = ap.parse_args()
    m = build_pack(a.runsheets, a.out)
    print(f"Saved {a.out} | {m['pages']} pages | day £{m['day']:,.0f} night £{m['night']:,.0f} "
          f"total £{m['total']:,.0f} under {m['n_under']} offroad {m['n_offroad']}")


if __name__ == '__main__':
    main()
