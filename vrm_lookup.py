"""
Vehicle lookup built from the master file (vehicle_master.xlsx).

The master has one tab per division, each with columns: REG | MAKE | AREA.
We turn it into two dicts the report builder needs:
  reg_to_area[REG] -> a Cover-sheet AREA label (one of COVER_AREAS)
  reg_make[REG]    -> make (SCANIA / VOLVO / ...) [bonus, not required]
  reg_plate[REG]   -> '25plate' | '24plate' | 'pre24'  (from the plate year)

The master's AREA spellings vary (e.g. '8 WHEELER', 'ARTIC', 'HOOK'), so we
normalise them to the exact labels used on the Cover sheet.
"""
import re
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).parent
MASTER = HERE / "vehicle_master.xlsx"
ORIGINAL = HERE / "vrm_original.xlsx"

# Original VRM_list.xlsx Type -> Cover area (used only as a fallback for regs
# the master doesn't contain).
ORIGINAL_TYPE_MAP = {
    '8W': '8 WHEELERS', '8 W DAY CABS': '8 WHEELERS', '8W SLEEPERS': '8 WHEELERS',
    'SLEEPERS': '8 WHEELERS', 'MIDLAND': '8 WHEELERS', 'ALLY BODY': '8 ALI BODY',
    'ALLY BODY - ASPHALT': '8 ALI BODY', 'ARTICS': 'ARTICS', "EV'S": '8 EV',
    'GRABS': 'GRABS', 'HOOKS': 'HOOKS', 'SWEEPER': 'SWEEPER', 'TIPWORX': 'TIPPER',
    'WHITE VOLVOS': 'ARTICS',
}

# Exact AREA labels that exist as rows on the Cover sheet (valid destinations)
COVER_AREAS = {
    '8 ALI BODY', '8 EV', '8 WHEELERS', 'ARTICS', 'CONCRETE MIXER', 'WORKSHOP',
    'GRABS', 'HOOKS', 'SWEEPER', 'TRAILERS', 'UNIDENTIFIED', 'BEAVER TAIL', 'CAR',
    'FUEL TANKER', 'PICK UP', 'SHUNTER', 'TIPPER', 'VAN', 'PLANT', 'JAY/FABSHOP',
    'TYRES', 'TARMAC/ASPHALT',
}

# Map the master file's AREA spellings -> Cover-sheet label.
# Keys are UPPER-CASED and space-trimmed.
AREA_NORMALISE = {
    '8 WHEELER': '8 WHEELERS', '8 WHEELERS': '8 WHEELERS',
    '8 SLEEPER': '8 WHEELERS', '8 W SLEEPER': '8 WHEELERS', 'SLEEPER': '8 WHEELERS',
    'ARTIC': 'ARTICS', 'ARTICS': 'ARTICS',
    '8 ALI BODY': '8 ALI BODY', 'ALI BODY': '8 ALI BODY',
    '8 EV': '8 EV',
    'BEAVER TAIL': 'BEAVER TAIL', '18T BEAVERTAIL': 'BEAVER TAIL', 'BEAVERTAIL': 'BEAVER TAIL',
    'HOOK': 'HOOKS', 'HOOKS': 'HOOKS',
    'VAN': 'VAN', 'VANS': 'VAN', 'FITTER VAN': 'VAN',
    'PICK UP': 'PICK UP', 'PICKUP': 'PICK UP',
    'CONCRETE MIXER': 'CONCRETE MIXER',
    'SWEEPER': 'SWEEPER',
    'GRABS': 'GRABS', 'GRAB': 'GRABS',
    'CAR': 'CAR', 'CARS': 'CAR',
    'SHUNTER': 'SHUNTER',
    'FUEL TANKER': 'FUEL TANKER',
    'TRAILER': 'TRAILERS', 'TRAILERS': 'TRAILERS',
    'TIPPER': 'TIPPER',
    # Plant & plant machinery -> PLANT
    'PLANT': 'PLANT', 'PLANT BODY': 'PLANT', 'PLANER': 'PLANT', 'SHOVEL': 'PLANT',
    'TELEHANDLER': 'PLANT', 'SCREENER': 'PLANT', 'CRUSHER': 'PLANT',
    # --- best-guess mappings (flagged for Daniel to confirm) ---
    'CHAIN LIFT': 'HOOKS',        # chain-lift ~ hook loader?
    'BOWSER': 'FUEL TANKER',      # bowser = mobile tank?
    'SIDELOADER': 'PLANT',        # sideloader?
}


def plate_cat(reg):
    reg = str(reg).upper().replace(' ', '')
    m = re.match(r'[A-Z]{2}(\d{2})[A-Z]{3}', reg)
    if not m:
        return None
    num = int(m.group(1))
    year = 1950 + num if num >= 51 else 2000 + num
    return '25plate' if year >= 2025 else ('24plate' if year == 2024 else 'pre24')


def _load_master(path=MASTER):
    wb = load_workbook(path, read_only=True, data_only=True)
    reg_to_area, reg_make, reg_plate = {}, {}, {}
    unknown_areas = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            reg = str(row[0]).strip().upper().replace(' ', '')
            if not reg:
                continue
            make = str(row[1]).strip().upper() if len(row) > 1 and row[1] else None
            area_raw = str(row[2]).strip().upper() if len(row) > 2 and row[2] else None
            if make:
                reg_make[reg] = make
            if area_raw:
                mapped = AREA_NORMALISE.get(area_raw)
                if mapped:
                    reg_to_area[reg] = mapped
                else:
                    reg_to_area[reg] = 'UNIDENTIFIED'
                    unknown_areas.setdefault(area_raw, []).append(f"{sheet}:{reg}")
            cat = plate_cat(reg)
            if cat:
                reg_plate[reg] = cat
    return reg_to_area, reg_make, reg_plate, unknown_areas


def _load_original(path=ORIGINAL):
    """Original VRM_list.xlsx (single sheet: VRM | Type | Base)."""
    reg_to_area = {}
    if not Path(path).exists():
        return reg_to_area
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        reg = str(row[0]).strip().upper().replace(' ', '')
        t = str(row[1]).strip().upper() if len(row) > 1 and row[1] else None
        area = ORIGINAL_TYPE_MAP.get(t)
        if area:
            reg_to_area[reg] = area
    return reg_to_area


def load_lookup(merge_original=True, report_unknowns=False):
    """Master is primary; original fills regs the master doesn't have."""
    reg_to_area, reg_make, reg_plate, unknown_areas = _load_master()
    filled = 0
    if merge_original:
        for reg, area in _load_original().items():
            if reg not in reg_to_area:
                reg_to_area[reg] = area
                if plate_cat(reg):
                    reg_plate.setdefault(reg, plate_cat(reg))
                filled += 1
    if report_unknowns:
        return reg_to_area, reg_make, reg_plate, unknown_areas, filled
    return reg_to_area, reg_make, reg_plate


if __name__ == "__main__":
    from collections import Counter
    reg_to_area, reg_make, reg_plate, unknown, filled = load_lookup(report_unknowns=True)
    print(f"Vehicles loaded: {len(reg_to_area)}  (incl. {filled} filled from original)")
    print("\nArea distribution:")
    for area, n in Counter(reg_to_area.values()).most_common():
        print(f"  {area:<16} {n}")
    print("\nPlate distribution:")
    for cat, n in Counter(reg_plate.values()).most_common():
        print(f"  {cat:<10} {n}")
    if unknown:
        print("\n!! AREA labels not recognised (routed to UNIDENTIFIED):")
        for area, regs in unknown.items():
            print(f"  '{area}': {len(regs)} -> {regs[:5]}")
    else:
        print("\nAll AREA labels recognised.")
