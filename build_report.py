"""
Fox Group Daily Report — builder (adapted from fox_daily_report_builder.py).

Differences from the original:
  * date args computed automatically (date_args.py)
  * vehicle areas from the merged master list (vrm_lookup.py via classify.py)
  * per-row budgets READ from the workbook's hidden column 35 (not hardcoded)
  * Windows-safe date formatting
  * blank template auto-generated from a base workbook when needed

Usage:
    python build_report.py <csv> <report_date YYYY-MM-DD> [base_xlsx]
"""
import sys
import shutil
from pathlib import Path
import datetime as dt

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle

from classify import (
    COVER_TOP_ROW, COVER_BOTTOM_ROW, INFO_ROWS, SCANIA_ROW, VOLVO_ROW, CAPITAL_ROW,
    PRE24_ROW, R24_ROW, R25_ROW, BOT_TOTAL_ROW, TOP_SHEETS, sheet_to_row, PLATE_ROWS,
    reg_year,
)
from date_args import compute as compute_dates, working_day_index, last_working_day
from parts_category import CATEGORIES
import queries

HERE = Path(__file__).parent
FIXED_WD = 22  # working days per month (Daniel's convention)
BUDGET_COL = 35
MTD_COL = 33
REMAIN_COL = 37


def fmt_date_long(d: dt.date) -> str:
    return f"{d.day} {d:%B %Y}"          # e.g. "2 July 2026" (no %-d, Windows-safe)


def make_blank_template(base_path: Path, out_path: Path):
    """Create a reusable blank template from a real report: keep structure,
    labels, formulas and budgets; wipe all daily data + tab rows."""
    shutil.copy(base_path, out_path)
    wb = load_workbook(out_path)
    cover = wb['Cover']
    data_rows = list(range(3, 36, 2)) + INFO_ROWS + list(COVER_BOTTOM_ROW.values())
    for r in data_rows:
        for c in range(2, 33):                     # daily columns
            cover.cell(r, c, 0)
        v33 = cover.cell(r, MTD_COL).value
        if not (isinstance(v33, str) and v33.startswith('=')):
            cover.cell(r, MTD_COL, None)           # clear static MTD (keep formulas)
        cover.cell(r, REMAIN_COL, None)            # clear remaining (recomputed)
    for tab in wb.sheetnames:
        if tab == 'Cover':
            continue
        ws = wb[tab]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    wb.save(out_path)
    return out_path


def _write_parts_sheet(wb, parts, report_date_long):
    """Dedicated 'Parts (25 & newer)' tab: category breakdown for 2025/2026
    plates, life-to-date and this-month, all divisions (incl. Capital fit-out)."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    name = 'Parts (25 & newer)'
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    NAVY = 'FF24214A'; ORANGE = 'FFEB941F'; LIGHT = 'FFF2F3F7'
    hdr_font = Font(bold=True, color='FFFFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor=NAVY)
    title_font = Font(bold=True, color='FF24214A', size=13)
    note_font = Font(italic=True, color='FF666666', size=9)
    bold = Font(bold=True, color='FF24214A')
    money = '£#,##0.00'
    thin = Side(style='thin', color='FFCCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws['A1'] = 'Parts Category Breakdown — 2025 & Newer Trucks'
    ws['A1'].font = title_font
    ws['A2'] = (f'Life-to-date spend on 2025/2026-plate vehicles (all divisions, '
                f'incl. capital fit-out).  As at {report_date_long}.')
    ws['A2'].font = note_font

    def section(start_row, scope_key, tot_key, tr_key, heading):
        r = start_row
        ws.cell(r, 1, heading).font = bold
        r += 1
        headers = ['Category', '2025 plate', '2026 plate', 'Total']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(r, ci, h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
            cell.alignment = Alignment(horizontal='left' if ci == 1 else 'right')
        r += 1
        d25 = parts[scope_key][2025]; d26 = parts[scope_key][2026]
        rowi = 0
        for cat in CATEGORIES:
            v25 = round(d25.get(cat, 0.0), 2); v26 = round(d26.get(cat, 0.0), 2)
            if not v25 and not v26:
                continue
            fill = PatternFill('solid', fgColor=LIGHT) if rowi % 2 == 0 else None
            vals = [cat, v25, v26, round(v25 + v26, 2)]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = border
                if ci > 1:
                    cell.number_format = money
                if fill:
                    cell.fill = fill
            r += 1; rowi += 1
        # total row
        t25 = parts[tot_key][2025]; t26 = parts[tot_key][2026]
        totrow = ['TOTAL', t25, t26, round(t25 + t26, 2)]
        for ci, v in enumerate(totrow, 1):
            cell = ws.cell(r, ci, v)
            cell.font = bold; cell.border = border
            cell.fill = PatternFill('solid', fgColor='FFE0E3EE')
            if ci > 1:
                cell.number_format = money
        r += 1
        # truck counts
        ws.cell(r, 1, f"Trucks: {parts[tr_key][2025]} × 2025, "
                      f"{parts[tr_key][2026]} × 2026").font = note_font
        return r + 2

    nxt = section(4, 'ltd', 'total_ltd', 'trucks_ltd', 'Life-to-date (since the trucks joined the fleet)')
    section(nxt, 'mtd', 'total_mtd', 'trucks_mtd', f'This month so far ({report_date_long})')

    ws.column_dimensions['A'].width = 32
    for col in ('B', 'C', 'D'):
        ws.column_dimensions[col].width = 15


def _write_fisher_nms_sheets(wb, fn, report_date_long):
    """Two tabs — 'J Fisher Breakdown' and 'NMS Breakdown' — age band x parts
    category, month-to-date and today, in the Parts (25 & newer) visual style."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from queries import FN_BANDS
    NAVY = 'FF24214A'; LIGHT = 'FFF2F3F7'
    hdr_font = Font(bold=True, color='FFFFFFFF', size=10)
    hdr_fill = PatternFill('solid', fgColor=NAVY)
    title_font = Font(bold=True, color='FF24214A', size=13)
    note_font = Font(italic=True, color='FF666666', size=9)
    bold = Font(bold=True, color='FF24214A')
    money = '£#,##0.00'
    thin = Side(style='thin', color='FFCCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    titles = {'J FISHER': ('J Fisher Breakdown', 'J Fisher (Trucks + Plant) — Age x Category'),
              'NMS': ('NMS Breakdown', 'NMS (Civil + Plant) — Age x Category')}

    def matrix_section(ws, start_row, mat, heading):
        r = start_row
        ws.cell(r, 1, heading).font = bold
        r += 1
        headers = ['Category'] + FN_BANDS + ['Total']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(r, ci, h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
            cell.alignment = Alignment(horizontal='left' if ci == 1 else 'right')
        r += 1
        col_tot = {b: 0.0 for b in FN_BANDS}
        rowi = 0
        for cat in CATEGORIES:
            bands = mat.get(cat, {})
            if not any(bands.get(b, 0) for b in FN_BANDS):
                continue
            fill = PatternFill('solid', fgColor=LIGHT) if rowi % 2 == 0 else None
            rowvals = [round(bands.get(b, 0.0), 2) for b in FN_BANDS]
            for b, v in zip(FN_BANDS, rowvals):
                col_tot[b] += v
            vals = [cat] + rowvals + [round(sum(rowvals), 2)]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(r, ci, v)
                cell.border = border
                if ci > 1:
                    cell.number_format = money
                if fill:
                    cell.fill = fill
            r += 1; rowi += 1
        tot_vals = ['TOTAL'] + [round(col_tot[b], 2) for b in FN_BANDS] + \
                   [round(sum(col_tot.values()), 2)]
        for ci, v in enumerate(tot_vals, 1):
            cell = ws.cell(r, ci, v)
            cell.font = bold; cell.border = border
            cell.fill = PatternFill('solid', fgColor='FFE0E3EE')
            if ci > 1:
                cell.number_format = money
        return r + 3

    for g, (tab, title) in titles.items():
        if tab in wb.sheetnames:
            del wb[tab]
        ws = wb.create_sheet(tab)
        ws['A1'] = title
        ws['A1'].font = title_font
        ws['A2'] = ("Registered vehicles banded by plate year; plant kit and stock "
                    f"lines have no VRM and sit under 'Unregistered / Plant'.  As at {report_date_long}.")
        ws['A2'].font = note_font
        nxt = matrix_section(ws, 4, fn[g]['mtd'], f"This month so far  (£{fn[g]['mtd_total']:,.2f})")
        matrix_section(ws, nxt, fn[g]['today'], f"Today  (£{fn[g]['today_total']:,.2f})")
        ws.column_dimensions['A'].width = 30
        from openpyxl.utils import get_column_letter
        for cc in range(2, len(FN_BANDS) + 3):
            ws.column_dimensions[get_column_letter(cc)].width = 14


def build(report_date: dt.date, fixed_wd=FIXED_WD):
    """Build the workbook + PDF for report_date, sourcing every figure for the
    whole month from the database — so month-to-date is always correct and no
    fragile rolling 'state' file is needed."""
    info = compute_dates(report_date, fixed_wd=fixed_wd)
    today_col = info['today_col']
    days_elapsed = info['days_elapsed']
    days_remaining = info['days_remaining']
    wd = info['working_days_in_month']
    y, m = report_date.year, report_date.month

    out_xlsx = HERE / f"fox_transaction_report_{report_date:%d-%m-%Y}.xlsx"
    out_pdf = HERE / f"daily_kpi_report_{report_date:%d-%m-%Y}.pdf"
    report_date_long = fmt_date_long(report_date)
    print(f"Building {report_date_long}  col={today_col} elapsed={days_elapsed} "
          f"remaining={days_remaining} wd={wd}  (from database)")

    # structural template only (labels, formulas, budgets in col 35 — no figures)
    shutil.copy(HERE / "blank_template.xlsx", out_xlsx)
    wb = load_workbook(out_xlsx)
    cover = wb['Cover']

    for tab in wb.sheetnames:                       # clear all division tabs
        if tab == 'Cover':
            continue
        ws = wb[tab]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

    # Registration-year rows — 2021-2026 + older/private (below the main layout)
    YEAR_ROWS = [(2021, 72), (2022, 73), (2023, 74), (2024, 75), (2025, 76), (2026, 77)]
    OTHER_ROW = 78          # vehicles with a reg that isn't a 2021-2026 plate
    YEAR_TOTAL_ROW = 79
    year_row_of = {yr: r for yr, r in YEAR_ROWS}
    all_year_rows = [r for _, r in YEAR_ROWS] + [OTHER_ROW]

    data_rows = list(range(3, 36, 2)) + INFO_ROWS + list(COVER_BOTTOM_ROW.values()) + all_year_rows
    for r in data_rows:
        for c in range(2, today_col + 1):
            cover.cell(r, c, 0)

    def col_for(d):
        # Cover columns are calendar days: col 2 = 1st ... col 32 = 31st.
        return d.day + 1

    # whitespace-tolerant lookups (DB strips names; some labels e.g.
    # 'Asphalt Plant ' carry a trailing space in the workbook)
    ROW_BY_DIV = {k.strip(): v for k, v in sheet_to_row.items()}
    BOTTOM_BY_AREA = {k.strip(): v for k, v in COVER_BOTTOM_ROW.items()}
    TOP_STRIPPED = {s.strip() for s in TOP_SHEETS}

    # accumulate every transaction of the month into (row, column) cells
    acc = {}

    def add(row, col, cost):
        acc[(row, col)] = acc.get((row, col), 0.0) + cost

    for rdate, division, area, plate, supplier, cost, vehicle_reg in queries.month_rows(y, m):
        col = col_for(rdate)
        if col < 2 or col > today_col:
            continue
        cost = float(cost or 0)
        division = (division or '').strip()
        area = (area or '').strip()
        drow = ROW_BY_DIV.get(division)
        if drow:
            add(drow, col, cost)
        if division in TOP_STRIPPED:
            arow = BOTTOM_BY_AREA.get(area)
            if arow:
                add(arow, col, cost)
            prow = PLATE_ROWS.get(plate)
            if prow:
                add(prow, col, cost)
            reg = (vehicle_reg or '').strip()
            if reg:  # a vehicle line: 2021-2026 -> its year row, else older/private
                add(year_row_of.get(reg_year(reg), OTHER_ROW), col, cost)
        sup = (supplier or '').upper()
        if 'SCANIA' in sup:
            add(SCANIA_ROW, col, cost)
        elif 'VOLVO' in sup or 'THOMAS HARDIE' in sup:
            add(VOLVO_ROW, col, cost)

    for (r, c), v in acc.items():
        cover.cell(r, c, round(v, 2))

    def gcv(r, c):
        v = cover.cell(r, c).value
        if v is None or (isinstance(v, str) and v.startswith('=')):
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def get_mtd(rn):
        return round(sum(gcv(rn, c) for c in range(2, today_col + 1)), 2)

    for rn in list(COVER_TOP_ROW.values()) + [CAPITAL_ROW]:
        cover.cell(rn, MTD_COL, get_mtd(rn))
    for rn in list(COVER_TOP_ROW.values()) + [CAPITAL_ROW]:
        budget = gcv(rn, BUDGET_COL)
        if budget:
            cover.cell(rn, REMAIN_COL, round(budget - get_mtd(rn), 2))

    # ── column totals + MTD written as VALUES, not =SUM() formulas ──
    # (Excel formulas in openpyxl-generated files kept showing blank / "breaking";
    #  hard numbers always display correctly, everywhere.)
    div_rows_all = list(range(3, 36, 2))
    area_rows_all = list(COVER_BOTTOM_ROW.values())
    for col in range(2, today_col + 1):
        cover.cell(37, col, round(sum(gcv(r, col) for r in div_rows_all), 2))              # Total row
        cover.cell(BOT_TOTAL_ROW, col, round(sum(gcv(r, col) for r in area_rows_all), 2))  # Area Total row
    for rn in INFO_ROWS + area_rows_all + [37, BOT_TOTAL_ROW]:
        cover.cell(rn, MTD_COL, round(sum(gcv(rn, c) for c in range(2, today_col + 1)), 2))

    # ── Spend by Registration Year block (2021-2026, rows 71-78) ──
    cover.cell(71, 1, 'SPEND BY REGISTRATION YEAR')
    for yr, r in YEAR_ROWS:
        cover.cell(r, 1, str(yr))
        cover.cell(r, MTD_COL, get_mtd(r))
    cover.cell(OTHER_ROW, 1, 'Older / private plates')
    cover.cell(OTHER_ROW, MTD_COL, get_mtd(OTHER_ROW))
    cover.cell(YEAR_TOTAL_ROW, 1, 'Registration Year Total')
    for col in range(2, today_col + 1):
        cover.cell(YEAR_TOTAL_ROW, col, round(sum(gcv(r, col) for r in all_year_rows), 2))
    cover.cell(YEAR_TOTAL_ROW, MTD_COL, get_mtd(YEAR_TOTAL_ROW))

    # ── Hook Fleet — spend by registration (PF request; rows 81+) ──
    HOOK_HDR = 81
    fleet, per_reg, unmatched = queries.hook_split(report_date)
    extras = [r for r in sorted(per_reg) if r not in fleet]
    hook_regs = fleet + extras
    cover.cell(HOOK_HDR, 1, 'HOOK FLEET — SPEND BY REGISTRATION')
    hr = HOOK_HDR + 1
    def _daycell(d):
        return col_for(d)
    for reg in hook_regs:
        cover.cell(hr, 1, f"{reg[:4]} {reg[4:]}")
        days = per_reg.get(reg, {})
        for d, amt in days.items():
            cc = _daycell(d)
            prev = cover.cell(hr, cc).value
            prev = float(prev) if isinstance(prev, (int, float)) else 0.0
            cover.cell(hr, cc, round(prev + amt, 2))
        cover.cell(hr, MTD_COL, round(sum(days.values()), 2))
        hr += 1
    UNM_ROW = hr
    cover.cell(UNM_ROW, 1, 'Hook lines with no registration')
    for d, amt in unmatched.items():
        cover.cell(UNM_ROW, _daycell(d), round(amt, 2))
    cover.cell(UNM_ROW, MTD_COL, round(sum(unmatched.values()), 2))
    HOOK_TOTAL_ROW = UNM_ROW + 1
    cover.cell(HOOK_TOTAL_ROW, 1, 'Hooks Total')
    for col in range(2, today_col + 1):
        tot = 0.0
        for r in range(HOOK_HDR + 1, HOOK_TOTAL_ROW):
            v = cover.cell(r, col).value
            if isinstance(v, (int, float)):
                tot += float(v)
        cover.cell(HOOK_TOTAL_ROW, col, round(tot, 2))
    cover.cell(HOOK_TOTAL_ROW, MTD_COL,
               round(sum(float(cover.cell(r, MTD_COL).value or 0)
                         for r in range(HOOK_HDR + 1, HOOK_TOTAL_ROW)), 2))

    # write the day's transactions into the division tabs
    tabmap = {s.strip(): s for s in wb.sheetnames}
    nextrow = {}
    for row in queries.day_tab_rows(report_date):
        sn = tabmap.get((row['division'] or '').strip())
        if not sn:
            continue
        ws = wb[sn]
        nr = nextrow.get(sn, 2)
        po = row['po_created_date']
        ws.cell(nr, 1, row['supplier'] or '')
        ws.cell(nr, 2, row['supplier_source_depot'] or '')
        ws.cell(nr, 3, row['system_no'] or '')
        ws.cell(nr, 4, row['supplier_pn'] or '')
        ws.cell(nr, 5, row['part_name'] or '')
        ws.cell(nr, 6, float(row['cost'] or 0))
        ws.cell(nr, 7, row['po_no'] or '')
        ws.cell(nr, 8, row['attached_order_no'])
        ws.cell(nr, 9, row['attached_customer'])
        ws.cell(nr, 10, str(po) if po else '')
        ws.cell(nr, 11, row['supply_type'] or '')
        ws.cell(nr, 12, row['item_count'])
        ws.cell(nr, 13, row['goods_received'])
        ws.cell(nr, 14, row['target_depot'] or '')
        ws.cell(nr, 15, row['assigned_depot'] or '')
        ws.cell(nr, 16, row['supplier_ref'] or '')
        ws.cell(nr, 17, row['custom_ref'] or '')
        ws.cell(nr, 18, row['area'] or '')
        nextrow[sn] = nr + 1

    # ── Parts Category breakdown for 2025 & newer trucks (dedicated tab) ──
    parts = queries.parts_category_split(report_date)
    _write_parts_sheet(wb, parts, report_date_long)

    # ── J Fisher / NMS age-band x category breakdown tabs ──
    fn = queries.fisher_nms_split(report_date)
    _write_fisher_nms_sheets(wb, fn, report_date_long)

    wb.save(out_xlsx)

    # ── verify balance + PDF ──
    wb2 = load_workbook(out_xlsx)
    cover2 = wb2['Cover']

    def g2(r, c):
        v = cover2.cell(r, c).value
        if v is None or (isinstance(v, str) and v.startswith('=')):
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    top = sum(g2(r, today_col) for r in range(3, 36, 2))
    bot = sum(g2(r, today_col) for r in COVER_BOTTOM_ROW.values())
    diff = abs(top - bot)
    print(f"Balance: top £{top:,.2f}  bottom £{bot:,.2f}  "
          f"{'BALANCED' if diff < 0.01 else f'GAP £{diff:,.2f}'}")

    year_today, year_mtd = queries.reg_year_split(report_date)
    hook_pdf = (hook_regs, per_reg, unmatched)
    _build_pdf(out_pdf, cover2, g2, report_date, report_date_long,
               today_col, days_elapsed, days_remaining, wd, year_today, year_mtd, parts,
               fn=fn, hooks=hook_pdf)
    print(f"Saved: {out_xlsx.name}")
    print(f"Saved: {out_pdf.name}")
    return out_xlsx, out_pdf, diff, top, report_date_long


# ── PDF (ported from original, budgets read from sheet) ─────────────
def _build_pdf(out_pdf, cover2, g2, report_date, REPORT_DATE, TODAY_COL,
               DAYS_ELAPSED, DAYS_REMAINING, WD, year_today=None, year_mtd=None,
               parts=None, fn=None, hooks=None):
    NAVY = colors.HexColor('#24214a'); ORANGE = colors.HexColor('#eb941f')
    BLUE = colors.HexColor('#00579e'); WHITE = colors.white
    LIGHT = colors.HexColor('#F2F3F7'); GREY = colors.HexColor('#666666')
    MID = colors.HexColor('#CCCCCC'); GREEN = colors.HexColor('#27ae60')
    PURPLE = colors.HexColor('#8e44ad'); AMBER = colors.HexColor('#f39c12')
    RED = colors.HexColor('#c0392b'); AMBG = colors.HexColor('#FFF3CD')
    RDBG = colors.HexColor('#FFCCCC'); TEAL = colors.HexColor('#16a085')

    def fmt(v):  return f"£{v:,.2f}"
    def fmts(v): return f"£{v:,.0f}"

    daily_total = sum(g2(r, TODAY_COL) for r in COVER_TOP_ROW.values())
    daily_damage = g2(33, TODAY_COL); daily_glass = g2(35, TODAY_COL)
    daily_capital = g2(43, TODAY_COL); daily_tyres = g2(21, TODAY_COL)
    mtd_total = sum(g2(r, c) for r in COVER_TOP_ROW.values() for c in range(2, TODAY_COL + 1))
    mtd_damage = sum(g2(33, c) for c in range(2, TODAY_COL + 1))
    mtd_tyres = sum(g2(21, c) for c in range(2, TODAY_COL + 1))
    mtd_capital = sum(g2(43, c) for c in range(2, TODAY_COL + 1))
    daily_avg = mtd_total / DAYS_ELAPSED if DAYS_ELAPSED > 0 else 0

    day_nz = {}
    for ci in range(2, TODAY_COL + 1):
        lbl = cover2.cell(2, ci).value or str(ci)
        tot = sum(g2(r, ci) for r in COVER_TOP_ROW.values())
        if tot > 0:
            day_nz[lbl] = tot
    bd_lbl = max(day_nz, key=day_nz.get) if day_nz else '-'
    bd_val = day_nz.get(bd_lbl, 0)

    # budgets from the workbook (hidden col 35)
    BUDGET_ROWS = [(3, 'Fox Wagons'), (5, 'Leyland Wagons'), (11, 'J FISHER (Trucks)'),
                   (13, 'NMS CIVIL'), (21, 'Tyres'), (29, 'J Fisher Plant'), (31, 'NMS Plant')]
    budget_data = [(name, g2(row, BUDGET_COL), sum(g2(row, c) for c in range(2, TODAY_COL + 1)))
                   for row, name in BUDGET_ROWS]
    div_rows = [(name, g2(row, TODAY_COL)) for name, row in COVER_TOP_ROW.items() if g2(row, TODAY_COL) > 0]
    day_rows = []
    for ci in range(2, TODAY_COL + 1):
        lb = cover2.cell(2, ci).value or ''
        dt2 = sum(g2(r, ci) for r in COVER_TOP_ROW.values())
        mo = REPORT_DATE.split()[-2][:3]
        if dt2 > 0:
            day_rows.append((f"{lb} {mo}", dt2, g2(33, ci), g2(21, ci), g2(43, ci)))

    W, H = A4; M = 12 * mm; HH = 22 * mm; FH = 10 * mm; CW = W - 2 * M
    YT = H - HH - 5 * mm; YB = FH + 4 * mm
    CARD_H = 22 * mm; RH = 6.5 * mm; RH_B = 7.5 * mm; LH = 6 * mm; G = 5 * mm
    n_half = (len(div_rows) + 1) // 2
    DIV_H = (n_half + 1) * RH; BGT_H = (len(budget_data) + 1) * RH_B; NOTE_H = 7 * mm
    P1_FIXED = LH + G + CARD_H + G + LH + G + DIV_H + G + LH + G + NOTE_H + G / 2 + BGT_H
    PAD1 = max((YT - YB - P1_FIXED) / 3, 2 * mm)
    MTD_H = CARD_H + 3 * mm + CARD_H; DAY_H = (len(day_rows) + 2) * RH
    PAD2 = max((YT - YB - LH - G - MTD_H - G - LH - G - DAY_H) / 2, 2 * mm)

    cv = canvas.Canvas(str(out_pdf), pagesize=A4)

    def chrome(pg, total):
        cv.setFillColor(NAVY); cv.rect(0, H - HH, W, HH, fill=1, stroke=0)
        cv.setFillColor(ORANGE); cv.rect(0, H - HH, W, 1.5 * mm, fill=1, stroke=0)
        cv.setFillColor(WHITE); cv.setFont('Helvetica-Bold', 16)
        cv.drawString(M, H - 10 * mm, 'Fox Group Ltd — Daily Workshop Spend Report')
        cv.setFont('Helvetica', 9)
        cv.drawString(M, H - 17.5 * mm, f'{REPORT_DATE}   |   Prepared by Knowles Farm Ltd')
        cv.drawRightString(W - M, H - 17.5 * mm, f'Page {pg} of {total}')
        cv.setFillColor(NAVY); cv.rect(0, 0, W, FH, fill=1, stroke=0)
        cv.setFillColor(WHITE); cv.setFont('Helvetica', 7.5)
        cv.drawString(M, 4 * mm, 'Confidential — Fox Group Ltd   |   Prepared by Knowles Farm Ltd')
        cv.drawRightString(W - M, 4 * mm, f'Generated {REPORT_DATE}')

    def sec_lbl(y, text):
        cv.setFillColor(NAVY); cv.setFont('Helvetica-Bold', 11); cv.drawString(M, y, text)
        cv.setFillColor(ORANGE); cv.rect(M, y - 2 * mm, CW, 0.7 * mm, fill=1, stroke=0)

    def draw_cards(y, ch, items, n):
        gap = 3 * mm; kw = (CW - (n - 1) * gap) / n
        for i, (title, val, col, sub) in enumerate(items):
            kx = M + i * (kw + gap); ky = y - ch
            cv.setFillColor(col); cv.roundRect(kx, ky, kw, ch, 2 * mm, fill=1, stroke=0)
            cv.setFillColor(WHITE); cv.setFont('Helvetica-Bold', 7.5)
            cv.drawString(kx + 3 * mm, ky + ch - 5.5 * mm, title)
            cv.setFont('Helvetica-Bold', 16); cv.drawCentredString(kx + kw / 2, ky + ch - 13 * mm, val)
            cv.setFillColor(colors.HexColor('#FFFFFF50', hasAlpha=True))
            cv.rect(kx + 3 * mm, ky + 6.5 * mm, kw - 6 * mm, 0.3 * mm, fill=1, stroke=0)
            cv.setFillColor(WHITE); cv.setFont('Helvetica', 6.5)
            cv.drawCentredString(kx + kw / 2, ky + 3.5 * mm, sub)

    def base_tbl():
        return [('BACKGROUND', (0, 0), (-1, 0), NAVY), ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [LIGHT, WHITE]),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'), ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('LINEBELOW', (0, 0), (-1, 0), 0.5, ORANGE), ('GRID', (0, 0), (-1, -1), 0.2, MID),
                ('TOPPADDING', (0, 0), (-1, -1), 1.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 1.5),
                ('LEFTPADDING', (0, 0), (-1, -1), 3), ('RIGHTPADDING', (0, 0), (-1, -1), 3)]

    # Page 1
    chrome(1, 7); y = YT
    sec_lbl(y, "TODAY'S SPEND"); y -= LH + G
    draw_cards(y, CARD_H, [("Today's Total", fmt(daily_total), BLUE, ''),
        ('Damage', fmt(daily_damage), RED, ''), ('Tyres', fmt(daily_tyres), NAVY, ''),
        ('Windscreen/Glass', fmt(daily_glass), TEAL, ''), ('Capital', fmt(daily_capital), ORANGE, '')], 5)
    y -= CARD_H + G + PAD1

    sec_lbl(y, 'Spend by Division'); y -= LH + G
    left_div = div_rows[:n_half]; right_div = div_rows[n_half:]
    while len(right_div) < len(left_div):
        right_div.append(('', ''))
    div_data = [['Division', 'Spend', 'Division', 'Spend']]
    for i in range(len(left_div)):
        lr = left_div[i]; rr = right_div[i]
        div_data.append([lr[0], fmt(lr[1]) if lr[1] else '', rr[0], fmt(rr[1]) if rr[1] else ''])
    dt_tbl = Table(div_data, colWidths=[71 * mm, 22 * mm, 71 * mm, 22 * mm], rowHeights=RH)
    dt_tbl.setStyle(TableStyle(base_tbl() + [('ALIGN', (2, 0), (2, -1), 'LEFT'),
                                             ('LINEAFTER', (1, 0), (1, -1), 0.5, MID)]))
    dt_tbl.wrapOn(cv, W, H); dt_tbl.drawOn(cv, M, y - DIV_H)
    y -= DIV_H + G + PAD1

    mo_name = REPORT_DATE.split()[-2]
    sec_lbl(y, f'Budget Tracker — {mo_name} {REPORT_DATE.split()[-1]}'); y -= LH + G
    cv.setFillColor(GREY); cv.setFont('Helvetica', 8.5)
    cv.drawString(M, y, f'{DAYS_ELAPSED} day{"s" if DAYS_ELAPSED > 1 else ""} elapsed — '
                        f'{DAYS_REMAINING} working days remaining of {WD}')
    y -= NOTE_H + G / 2
    bgt_cols = [52 * mm, 26 * mm, 28 * mm, 28 * mm, 52 * mm]
    bgt_data = [['Division', 'Budget', 'MTD Spend', 'Remaining', 'Daily Avg to stay on budget']]
    bgt_sty = base_tbl()
    for i, (name, budget, spent) in enumerate(budget_data):
        rem = budget - spent; dn = rem / DAYS_REMAINING if DAYS_REMAINING > 0 else 0
        bgt_data.append([name, fmts(budget), fmt(spent), fmt(rem), fmt(dn)])
        bg = LIGHT if i % 2 == 0 else WHITE
        bgt_sty += [('BACKGROUND', (0, i + 1), (2, i + 1), bg), ('BACKGROUND', (4, i + 1), (4, i + 1), bg),
                    ('TOPPADDING', (0, i + 1), (-1, i + 1), 2), ('BOTTOMPADDING', (0, i + 1), (-1, i + 1), 2)]
        if rem < 0:
            bgt_sty += [('BACKGROUND', (3, i + 1), (3, i + 1), RDBG), ('TEXTCOLOR', (3, i + 1), (3, i + 1), RED),
                        ('FONTNAME', (3, i + 1), (3, i + 1), 'Helvetica-Bold')]
        elif budget > 0 and spent / budget > 0.7:
            bgt_sty += [('BACKGROUND', (3, i + 1), (3, i + 1), AMBG), ('TEXTCOLOR', (3, i + 1), (3, i + 1), AMBER),
                        ('FONTNAME', (3, i + 1), (3, i + 1), 'Helvetica-Bold')]
        else:
            bgt_sty += [('TEXTCOLOR', (3, i + 1), (3, i + 1), GREEN), ('FONTNAME', (3, i + 1), (3, i + 1), 'Helvetica-Bold')]
    btbl = Table(bgt_data, colWidths=bgt_cols, rowHeights=RH_B)
    btbl.setStyle(TableStyle(bgt_sty)); btbl.wrapOn(cv, W, H); btbl.drawOn(cv, M, y - BGT_H)

    # Page 2 — Month-to-Date + Spend by Registration Year
    cv.showPage(); chrome(2, 7); y = YT
    sec_lbl(y, f'Month-to-Date — {mo_name} {REPORT_DATE.split()[-1]}'); y -= LH + G
    draw_cards(y, CARD_H, [('MTD Total', fmt(mtd_total), BLUE, f'{DAYS_ELAPSED} of {WD} working days'),
        ('MTD Damage', fmt(mtd_damage), RED, ''), ('MTD Tyres', fmt(mtd_tyres), NAVY, '')], 3)
    y -= CARD_H + 3 * mm
    draw_cards(y, CARD_H, [('MTD Capital', fmt(mtd_capital), ORANGE, ''),
        ('Daily Average', fmt(daily_avg), GREEN, f'Target {fmts(mtd_total / WD)} over {WD} days'),
        ('Biggest Day', fmt(bd_val), PURPLE, f'{bd_lbl}')], 3)
    y -= CARD_H + G + 6 * mm

    sec_lbl(y, 'Spend by Registration Year (vehicle spend)'); y -= LH + G
    yt = year_today or {}
    ym = year_mtd or {}
    yr_data = [['Registration Year', "Today's Spend", 'Month-to-Date']]
    tot_t = tot_m = 0.0
    for yr in (2021, 2022, 2023, 2024, 2025, 2026):
        t = yt.get(yr, 0.0); mt = ym.get(yr, 0.0)
        yr_data.append([str(yr), fmt(t), fmt(mt)]); tot_t += t; tot_m += mt
    ot = yt.get('other', 0.0); om = ym.get('other', 0.0)
    yr_data.append(['Older / private plates', fmt(ot), fmt(om)]); tot_t += ot; tot_m += om
    yr_data.append(['Total (all vehicle spend)', fmt(tot_t), fmt(tot_m)])
    n_yr = len(yr_data)
    yr_sty = base_tbl() + [('BACKGROUND', (0, n_yr - 1), (-1, n_yr - 1), colors.HexColor('#E0E3EE')),
                           ('FONTNAME', (0, n_yr - 1), (-1, n_yr - 1), 'Helvetica-Bold'),
                           ('LINEABOVE', (0, n_yr - 1), (-1, n_yr - 1), 0.5, NAVY)]
    yr_tbl = Table(yr_data, colWidths=[70 * mm, 58 * mm, 58 * mm], rowHeights=RH_B)
    yr_tbl.setStyle(TableStyle(yr_sty)); yr_tbl.wrapOn(cv, W, H)
    yr_tbl.drawOn(cv, M, y - n_yr * RH_B)

    # Page 3 — Parts Category Breakdown (2025 & newer trucks)
    cv.showPage(); chrome(3, 7); y = YT
    sec_lbl(y, 'Parts Category Breakdown — 2025 & Newer Trucks'); y -= LH + G
    p = parts or {'ltd': {2025: {}, 2026: {}}, 'mtd': {2025: {}, 2026: {}},
                  'total_ltd': {2025: 0, 2026: 0}, 'total_mtd': {2025: 0, 2026: 0},
                  'trucks_ltd': {2025: 0, 2026: 0}, 'trucks_mtd': {2025: 0, 2026: 0}}
    cv.setFillColor(GREY); cv.setFont('Helvetica', 8.5)
    cv.drawString(M, y, 'Life-to-date spend on 2025/2026-plate vehicles — all divisions, '
                        'including capital fit-out.')
    y -= 5 * mm
    cv.drawString(M, y, f"This month so far: 25-plate {fmt(p['total_mtd'][2025])}  "
                        f"| 26-plate {fmt(p['total_mtd'][2026])}")
    y -= NOTE_H + G

    d25 = p['ltd'][2025]; d26 = p['ltd'][2026]
    pc_data = [['Category', '2025 plate', '2026 plate', 'Total']]
    for cat in CATEGORIES:
        v25 = d25.get(cat, 0.0); v26 = d26.get(cat, 0.0)
        if not v25 and not v26:
            continue
        pc_data.append([cat, fmt(v25), fmt(v26), fmt(v25 + v26)])
    t25 = p['total_ltd'][2025]; t26 = p['total_ltd'][2026]
    pc_data.append(['TOTAL', fmt(t25), fmt(t26), fmt(t25 + t26)])
    n_pc = len(pc_data)
    pc_sty = base_tbl() + [('BACKGROUND', (0, n_pc - 1), (-1, n_pc - 1), colors.HexColor('#E0E3EE')),
                           ('FONTNAME', (0, n_pc - 1), (-1, n_pc - 1), 'Helvetica-Bold'),
                           ('LINEABOVE', (0, n_pc - 1), (-1, n_pc - 1), 0.5, NAVY),
                           ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5)]
    pc_tbl = Table(pc_data, colWidths=[66 * mm, 40 * mm, 40 * mm, 40 * mm], rowHeights=RH_B)
    pc_tbl.setStyle(TableStyle(pc_sty)); pc_tbl.wrapOn(cv, W, H)
    pc_tbl.drawOn(cv, M, y - n_pc * RH_B)
    y -= n_pc * RH_B + G + 4 * mm
    cv.setFillColor(GREY); cv.setFont('Helvetica-Oblique', 8)
    cv.drawString(M, y, f"Based on {p['trucks_ltd'][2025]} × 2025-plate and "
                        f"{p['trucks_ltd'][2026]} × 2026-plate trucks on fleet.")

    # Pages 4-5 — J Fisher / NMS Age x Category (PF request 16/07)
    from queries import FN_BANDS
    fn = fn or {}
    fn_titles = [('J FISHER', 'J Fisher (Trucks + Plant) — Age Band x Parts Category'),
                 ('NMS', 'NMS (Civil + Plant) — Age Band x Parts Category')]
    band_short = ['2021', '2022', '2023', '2024', '2025', '2026', 'Older', 'Unreg/Plant']
    for pi, (g, title) in enumerate(fn_titles):
        cv.showPage(); chrome(4 + pi, 7); y = YT
        sec_lbl(y, title); y -= LH + G
        gdat = fn.get(g, {'mtd': {}, 'today': {}, 'mtd_total': 0, 'today_total': 0})
        cv.setFillColor(GREY); cv.setFont('Helvetica', 8.5)
        cv.drawString(M, y, "Registered vehicles banded by plate year. Plant kit and stock lines "
                            "carry no VRM and sit under Unreg/Plant.")
        y -= 5 * mm
        cv.drawString(M, y, f"Month to date: {fmt(gdat['mtd_total'])}   |   Today: {fmt(gdat['today_total'])}")
        y -= NOTE_H + G
        mat = gdat['mtd']
        fnd = [['Category'] + band_short + ['Total']]
        col_tot = {b: 0.0 for b in FN_BANDS}
        for cat in CATEGORIES:
            bands = mat.get(cat, {})
            if not any(bands.get(b, 0) for b in FN_BANDS):
                continue
            vals = [round(bands.get(b, 0.0), 2) for b in FN_BANDS]
            for b, v in zip(FN_BANDS, vals):
                col_tot[b] += v
            fnd.append([cat] + [fmt(v) if v else '—' for v in vals] + [fmt(sum(vals))])
        n_fn = len(fnd)
        fnd.append(['TOTAL'] + [fmt(round(col_tot[b], 2)) for b in FN_BANDS] +
                   [fmt(round(sum(col_tot.values()), 2))])
        fn_sty = base_tbl() + [('FONTSIZE', (0, 0), (-1, -1), 6.8),
                               ('BACKGROUND', (0, n_fn), (-1, n_fn), colors.HexColor('#E0E3EE')),
                               ('FONTNAME', (0, n_fn), (-1, n_fn), 'Helvetica-Bold'),
                               ('LINEABOVE', (0, n_fn), (-1, n_fn), 0.5, NAVY)]
        fn_cols = [50 * mm] + [15 * mm] * 8 + [16 * mm]
        fn_tbl = Table(fnd, colWidths=fn_cols, rowHeights=RH_B)
        fn_tbl.setStyle(TableStyle(fn_sty)); fn_tbl.wrapOn(cv, W, H)
        fn_tbl.drawOn(cv, M, y - len(fnd) * RH_B)

    # Page 6 — Hook Fleet spend by registration
    cv.showPage(); chrome(6, 7); y = YT
    sec_lbl(y, 'Hook Fleet — Spend by Registration'); y -= LH + G
    hook_regs, per_reg, unmatched = hooks if hooks else ([], {}, {})
    cv.setFillColor(GREY); cv.setFont('Helvetica', 8.5)
    cv.drawString(M, y, 'Every hook wagon on fleet — spend today and month-to-date. '
                        'Reconciles with the HOOKS area row on the cover sheet.')
    y -= NOTE_H + G
    hk_data = [['Registration', "Today", 'Month-to-Date']]
    tot_t = tot_m = 0.0
    for reg in hook_regs:
        days = per_reg.get(reg, {})
        t = round(days.get(report_date, 0.0), 2)
        mtd_v = round(sum(days.values()), 2)
        tot_t += t; tot_m += mtd_v
        hk_data.append([f"{reg[:4]} {reg[4:]}", fmt(t) if t else '—', fmt(mtd_v) if mtd_v else '—'])
    if unmatched:
        ut = round(unmatched.get(report_date, 0.0), 2)
        um = round(sum(unmatched.values()), 2)
        tot_t += ut; tot_m += um
        hk_data.append(['No registration on line', fmt(ut) if ut else '—', fmt(um)])
    n_hk = len(hk_data)
    hk_data.append(['HOOKS TOTAL', fmt(round(tot_t, 2)), fmt(round(tot_m, 2))])
    hk_sty = base_tbl() + [('BACKGROUND', (0, n_hk), (-1, n_hk), colors.HexColor('#E0E3EE')),
                           ('FONTNAME', (0, n_hk), (-1, n_hk), 'Helvetica-Bold'),
                           ('LINEABOVE', (0, n_hk), (-1, n_hk), 0.5, NAVY)]
    hk_tbl = Table(hk_data, colWidths=[70 * mm, 58 * mm, 58 * mm], rowHeights=RH)
    hk_tbl.setStyle(TableStyle(hk_sty)); hk_tbl.wrapOn(cv, W, H)
    hk_tbl.drawOn(cv, M, y - len(hk_data) * RH)

    # Page 7 — Day-by-Day Summary
    cv.showPage(); chrome(7, 7); y = YT
    sec_lbl(y, 'Day-by-Day Summary'); y -= LH + G
    dcols = [28 * mm, 40 * mm, 32 * mm, 32 * mm, 32 * mm, 22 * mm]
    ddata = [['Date', 'Total Spend', 'Damage', 'Tyres', 'Capital', 'Daily Avg']]
    for dlbl, tot, dmg, tyr, cap in day_rows:
        ddata.append([dlbl, fmt(tot), fmt(dmg), fmt(tyr), fmt(cap), fmt(tot)])
    n_dd = len(ddata)
    ddata.append(['MTD', fmt(mtd_total), fmt(mtd_damage), fmt(mtd_tyres), fmt(mtd_capital), fmt(daily_avg)])
    dsty = base_tbl()
    dsty += [('BACKGROUND', (0, n_dd), (-1, n_dd), colors.HexColor('#E0E3EE')),
             ('FONTNAME', (0, n_dd), (-1, n_dd), 'Helvetica-Bold'),
             ('LINEABOVE', (0, n_dd), (-1, n_dd), 0.5, NAVY)]
    day_tbl = Table(ddata, colWidths=dcols, rowHeights=RH)
    day_tbl.setStyle(TableStyle(dsty))
    day_tbl.wrapOn(cv, W, H); day_tbl.drawOn(cv, M, y - len(ddata) * RH)
    cv.save()


def main():
    if len(sys.argv) < 2:
        print("Usage: python build_report.py <report_date YYYY-MM-DD>")
        sys.exit(1)
    build(dt.date.fromisoformat(sys.argv[1]))


if __name__ == "__main__":
    main()
