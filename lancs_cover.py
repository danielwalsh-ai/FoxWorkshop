"""
Cover sheet for Mel's Lancashire master — the 44 rows Paul highlighted, by month.

Paul shaded 44 rows on the DAILY tab in dark blue (0070C0). This builds a new
first tab holding those rows aggregated month by month from January 2025, with a
native Excel bar chart and trend line, and opens scrolled to the latest month.

The workbook carries 180 chart parts and 16 drawings, so it is never re-saved
through openpyxl — that would destroy them. The new sheet is built standalone and
its parts are transplanted into the original zip.
"""
import re
import shutil
import zipfile
import datetime as dt
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage

HIGHLIGHT = "FF0070C0"          # the dark blue Paul used
START = dt.date(2026, 1, 1)
DATE_ROW, FIRST_DATA_COL = 1, 4
SHEET_NAME = "COVER"
FOX_NAVY = "1A2646"      # Fox Group navy, sampled from the pack
CHARTS_ACROSS = 4        # grid, not one long column
CHART_W, CHART_H = 13.0, 7.0     # cm
COL_STEP, ROW_STEP = 7, 19       # columns / rows between chart anchors
FOX_ORANGE = "E97A1F"
LOGO = Path(__file__).parent / "static" / "fox-group-logo.png"
TREND_RED = "C00000"      # trend line

# Rows that are a rate rather than a running total have to be averaged over the
# month, not summed.
def agg_for(label):
    u = (label or "").upper()
    if "AVERAGE" in u or "AVG" in u or u.startswith("NO WAGONS"):
        return "mean"
    return "sum"


# Tidy display names for the section a row sits under. Derived from the block's
# "<X> AVERAGE" row, so it works on the Leyland sheet too without a second list.
SECTION_NAMES = {
    "EVS": "EV's", "EV'S": "EV's",
    "SLEEEPERS": "Sleepers", "SLEEPERS": "Sleepers",
    "8 W": "8 Wheelers", "8W": "8 Wheelers", "8 W DAY CABS": "8 Wheelers",
    "WHITE": "White Volvos", "WHITE VOLVOS": "White Volvos",
    "ARTICS": "Artics", "ARTIC": "Artics",
    "MIDLAND": "Midland", "MIDLANDS": "Midland",
    "TIPWORX": "Tipworx", "BRAMPTON/TIPWORX": "Tipworx",
    "ALLY BODY": "Alloy", "ALLY": "Alloy",
    "HOOKS": "Hooks", "HOOKS ON HIRE": "Hooks on hire",
    "SWEEPER": "Sweepers", "SWEEPERS": "Sweepers",
    "GRABS": "Grabs", "GRAB": "Grabs",
    "8W SLEEPERS": "8W Sleepers", "8 W SLEEPERS": "8W Sleepers",
}


def _norm(s):
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def section_map(path, sheet="DAILY", limit=200):
    """{row: section} for every row inside a category block.

    Keyed off the block *headers*, not the AVERAGE rows: each block runs header,
    wagons, AVERAGE, then TOTAL EARNINGS / WAGES / FUEL / MARGIN, so anchoring on
    AVERAGE puts a category's totals in the next section along. A block runs from
    its header to the row before the next one; anything after the last block is
    fleet-wide and gets no section.
    """
    ws = openpyxl.load_workbook(path)[sheet]
    heads = []
    for r in range(1, limit):
        v = ws.cell(r, 1).value
        if not isinstance(v, str) or not v.strip():
            continue
        key = v.strip().upper()
        if key.endswith(("AVERAGE", "TOTAL")):
            continue
        name = SECTION_NAMES.get(key) or SECTION_NAMES.get(_norm(key))
        if name:
            heads.append((r, name))
    # The last block has no following header, so run it to its own AVERAGE row
    # plus the four total rows that trail it.
    last_avg = max((r for r in range(1, limit)
                    if isinstance(ws.cell(r, 1).value, str)
                    and ws.cell(r, 1).value.strip().upper().endswith("AVERAGE")),
                   default=None)
    out = {}
    for i, (r, name) in enumerate(heads):
        if i + 1 < len(heads):
            end = heads[i + 1][0] - 1
        else:
            end = (last_avg + 4) if last_avg and last_avg > r else r
        for rr in range(r, end + 1):
            out[rr] = name
    return out


def highlighted_rows(path, sheet="DAILY", limit=200):
    """[(row, label)] for every row Paul shaded, in sheet order."""
    ws = openpyxl.load_workbook(path)[sheet]
    out = []
    for r in range(1, limit):
        c = ws.cell(r, 1)
        try:
            rgb = c.fill.fgColor.rgb
        except Exception:
            rgb = None
        if isinstance(rgb, str) and rgb.upper() == HIGHLIGHT:
            lab = str(c.value or "").strip()
            if lab:
                out.append((r, lab))
    return out


def date_layout(ws):
    """(date_row, first_data_col). Lancashire has dates on row 1 from column D,
    Leyland on row 2 from column C, so this is detected rather than assumed."""
    best = (DATE_ROW, FIRST_DATA_COL, -1)
    for r in (1, 2, 3):
        cs = [c for c in range(2, min(ws.max_column, 60) + 1)
              if isinstance(ws.cell(r, c).value, dt.datetime)]
        if len(cs) > best[2]:
            best = (r, cs[0] if cs else FIRST_DATA_COL, len(cs))
    return best[0], best[1]


def default_rows(path, sheet="DAILY", limit=200):
    """Leyland has no highlighting, so mirror Paul's Lancashire selection by
    meaning: each block's AVERAGE and TOTAL, then the fleet-wide summary rows
    that follow the last block."""
    ws = openpyxl.load_workbook(path)[sheet]
    secs = section_map(path, sheet, limit)
    last_block = max(secs) if secs else 0
    out = []
    for r in range(1, limit):
        v = ws.cell(r, 1).value
        if not isinstance(v, str) or not v.strip():
            continue
        lab, up = v.strip(), v.strip().upper()
        if r <= last_block:
            if up.endswith(("AVERAGE", "TOTAL")):
                out.append((r, lab))
        # The night-work block lists individual wagons as reg+N (PJ21OWEN), so
        # allow a 4-letter tail or they land in the summary as their own charts.
        elif not re.match(r"^[A-Z]{2}[0-9]{2}[A-Z]{3,4}$", up.replace(" ", "")):
            if not up.startswith("DATE "):
                out.append((r, lab))
    return out


# ── formula fallback ─────────────────────────────────────────────────
# Columns written by our own XML-level fill carry formulas Excel has never
# recalculated, so openpyxl's cached values read back as None and a freshly
# filled month would fall off the cover charts (Paul spotted August missing,
# 07/08/2026). This evaluates the handful of formula shapes the masters
# actually use — SUM / AVERAGE / AVERAGEIF / COUNTIF / IFERROR, cell refs and
# plain arithmetic — from the raw wagon values, but only when the cache is
# empty, so workbooks Excel has saved (Mel's) behave exactly as before.
_FUNC_RE = re.compile(r"(IFERROR|SUM|AVERAGEIF|AVERAGE|COUNTIF)\(([^()]*)\)", re.I)
_REF_RE = re.compile(r"\$?[A-Z]{1,3}\$?[0-9]{1,7}")
_ERR = "#ERR"


class CellResolver:
    """Cached value if Excel left one; otherwise evaluate the formula."""

    def __init__(self, ws_formulas, ws_values):
        self.wsf, self.wsv = ws_formulas, ws_values
        self.memo = {}

    def value(self, row, col):
        key = (row, col)
        if key in self.memo:
            return self.memo[key]
        self.memo[key] = None                     # cycle guard
        v = self.wsv.cell(row, col).value
        if v is None:
            f = self.wsf.cell(row, col).value
            if isinstance(f, str) and f.startswith("="):
                try:
                    v = self._eval(f[1:])
                except Exception:
                    v = None
        self.memo[key] = v
        return v

    def _range(self, ref):
        from openpyxl.utils import range_boundaries
        ref = ref.replace("$", "").strip()
        if "!" in ref:
            raise ValueError("cross-sheet reference")
        b = range_boundaries(ref if ":" in ref else f"{ref}:{ref}")
        return [self.value(r, c)
                for r in range(b[1], b[3] + 1) for c in range(b[0], b[2] + 1)]

    @staticmethod
    def _nums(vals):
        return [float(v) for v in vals if isinstance(v, (int, float))]

    def _call(self, name, args):
        name = name.upper()
        parts = [a.strip() for a in args.split(",")]
        if name == "IFERROR":
            if len(parts) != 2:
                raise ValueError("IFERROR arity")
            a, b = parts
            if a != _ERR:
                return a
            return b.strip('"').strip() or _ERR   # blank fallback = no value
        if name == "SUM":
            vals = []
            for p in parts:
                if p == _ERR:
                    return _ERR
                if _REF_RE.fullmatch(p.replace(":", "").replace("$", "")) or ":" in p:
                    vals += self._nums(self._range(p))
                else:
                    vals.append(float(p))
            return repr(sum(vals))
        if name == "AVERAGE":
            vals = []
            for p in parts:
                vals += self._nums(self._range(p)) if not _is_number(p) else [float(p)]
            return repr(sum(vals) / len(vals)) if vals else _ERR
        if name == "AVERAGEIF":
            if len(parts) != 2:
                raise ValueError("AVERAGEIF arity")
            crit = parts[1].strip().strip('"')
            vals = self._nums(self._range(parts[0]))
            if crit == "<>0":
                vals = [v for v in vals if v != 0]
            elif crit != "<>":
                raise ValueError(f"AVERAGEIF criteria {crit!r}")
            return repr(sum(vals) / len(vals)) if vals else _ERR
        if name == "COUNTIF":
            if len(parts) != 2:
                raise ValueError("COUNTIF arity")
            crit = parts[1].strip().strip('"')
            vals = self._range(parts[0])
            if _is_number(crit):
                return repr(sum(1 for v in vals
                                if isinstance(v, (int, float)) and float(v) == float(crit)))
            return repr(sum(1 for v in vals
                            if isinstance(v, str) and v.strip().upper() == crit.upper()))
        raise ValueError(name)

    def _eval(self, expr):
        expr = expr.strip()
        # a bare reference can hold text (a VOR'd wagon) — hand it back as is
        if _REF_RE.fullmatch(expr):
            b = expr.replace("$", "")
            m = re.match(r"([A-Z]{1,3})([0-9]+)$", b)
            from openpyxl.utils import column_index_from_string
            return self.value(int(m.group(2)), column_index_from_string(m.group(1)))
        for _ in range(30):                       # innermost call first
            m = _FUNC_RE.search(expr)
            if not m:
                break
            expr = expr[:m.start()] + str(self._call(m.group(1), m.group(2))) + expr[m.end():]
        else:
            raise ValueError("formula too deep")
        if expr.strip() == _ERR:
            return None
        def sub_ref(m):
            b = m.group(0).replace("$", "")
            mm = re.match(r"([A-Z]{1,3})([0-9]+)$", b)
            from openpyxl.utils import column_index_from_string
            v = self.value(int(mm.group(2)), column_index_from_string(mm.group(1)))
            if not isinstance(v, (int, float)):
                raise ValueError(f"{b} is not numeric")
            return repr(float(v))
        expr = _REF_RE.sub(sub_ref, expr)
        if _ERR in expr:
            return None
        if not re.fullmatch(r"[0-9.eE+\-*/() ]+", expr):
            raise ValueError(f"unsupported formula remainder {expr!r}")
        return eval(expr, {"__builtins__": {}})


def _is_number(s):
    try:
        float(s)
        return True
    except (TypeError, ValueError):
        return False


def monthly(path, rows, start=START, sheet="DAILY"):
    """{(row,label): {month_start: value}} plus the ordered month list."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    res = CellResolver(openpyxl.load_workbook(path)[sheet], ws)
    DATE_ROW, FIRST_DATA_COL = date_layout(ws)
    cols = {}
    for c in range(FIRST_DATA_COL, ws.max_column + 1):
        v = ws.cell(DATE_ROW, c).value
        if isinstance(v, dt.datetime) and v.date() >= start:
            cols.setdefault(dt.date(v.year, v.month, 1), []).append(c)
    months = sorted(cols)
    data = {}
    for r, lab in rows:
        how = agg_for(lab)
        series = {}
        for mth in months:
            vals = [res.value(r, c) for c in cols[mth]]
            vals = [float(v) for v in vals if isinstance(v, (int, float))]
            if not vals:
                series[mth] = None
            else:
                series[mth] = round(sum(vals), 2) if how == "sum" \
                    else round(sum(vals) / len(vals), 2)
        data[(r, lab)] = series
    return months, data


def build_cover_workbook(months, data, out_path, chart_rows=None, sections=None,
                         company="Fox Brothers (Lancashire)"):
    """A standalone workbook holding just the cover sheet + its native chart."""
    from openpyxl.chart import BarChart, Reference, Series
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.trendline import Trendline
    from openpyxl.chart.label import DataLabelList
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.text import (Paragraph, ParagraphProperties,
                                       CharacterProperties)
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    navy = "FF" + FOX_NAVY
    ws.sheet_view.showGridLines = False          # cleaner behind the charts

    # Fox mark top-left, then the title block, then a thin navy rule.
    if LOGO.exists():
        img = XLImage(str(LOGO))
        img.width, img.height = 232, 52          # keeps the 4.47:1 aspect
        img.anchor = "A1"
        ws.add_image(img)
    for r, h in ((1, 15), (2, 15), (3, 15), (4, 6)):
        ws.row_dimensions[r].height = h

    ws["A5"] = company
    ws["A5"].font = Font(name="Calibri", size=20, bold=True, color=navy)
    ws["A6"] = "Monthly summary"
    ws["A6"].font = Font(name="Calibri", size=13, color=navy)
    ws["A7"] = (f"Rows selected by Paul Fox, by month from {min(months):%B %Y}. "
                f"Totals summed; averages and wagon counts averaged.")
    ws["A7"].font = Font(name="Calibri", size=9, italic=True, color="FF7F7F7F")
    ws.row_dimensions[5].height = 27
    ws.row_dimensions[6].height = 19
    ws.row_dimensions[8].height = 5

    hdr = 9
    thin = Side(style="thin", color="FFD9D9D9")
    edge = Border(bottom=thin)
    hc = ws.cell(hdr, 1, "Metric")
    hc.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
    hc.fill = PatternFill("solid", fgColor=navy)
    hc.alignment = Alignment(vertical="center")
    for j, mth in enumerate(months):
        c = ws.cell(hdr, 2 + j, dt.datetime(mth.year, mth.month, 1))
        c.number_format = "mmm yy"
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=navy)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[hdr].height = 20
    rule = PatternFill("solid", fgColor="FF" + FOX_ORANGE)
    for j in range(max(len(months) + 1, 8)):
        ws.cell(hdr - 1, 1 + j).fill = rule

    band = PatternFill("solid", fgColor="FFF4F6F9")
    sections = sections or {}
    for i, ((r, lab), series) in enumerate(data.items()):
        rr = hdr + 1 + i
        # "TOTAL EARNINGS" appears once per category, so without the section in
        # front of it six charts carry the same title and can't be told apart.
        sec = sections.get(r)
        # Both sheets word these unevenly — SLEEEPERS, 8 W, ALLY BODY, SWEEPER —
        # so normalise rather than echo them.
        if sec and lab.strip().upper().endswith("AVERAGE"):
            shown = f"{sec} — Average"
        elif sec and lab.strip().upper().endswith("TOTAL"):
            shown = f"{sec} — Total"
        elif sec and _norm(sec) not in _norm(lab):
            shown = f"{sec} — {lab.title()}"
        else:
            shown = lab.title()
        nc = ws.cell(rr, 1, shown)
        nc.font = Font(name="Calibri", size=9.5, bold=True, color="FF262626")
        nc.border = edge
        for j, mth in enumerate(months):
            cell = ws.cell(rr, 2 + j, series.get(mth))
            cell.number_format = "#,##0;-#,##0;\"–\""
            cell.font = Font(name="Calibri", size=9.5, color="FF262626")
            cell.alignment = Alignment(horizontal="right")
            cell.border = edge
        if i % 2:                                # subtle banding, easier to read across
            for j in range(len(months) + 1):
                ws.cell(rr, 1 + j).fill = band
    last_row = hdr + len(data)

    ws.column_dimensions["A"].width = 32
    # Every column the charts sit on gets the same width, not just the ones the
    # table uses — otherwise the 4-across grid drifts once it runs past the table.
    grid_cols = max(len(months), CHARTS_ACROSS * COL_STEP + 1)
    for j in range(grid_cols):
        ws.column_dimensions[get_column_letter(2 + j)].width = 10.5
    # No frozen panes. A frozen column A clips any chart anchored there — you'd
    # have to widen the column to see it — and a frozen header row leaves month
    # headings floating over empty space once you scroll past the table.

    # One chart per highlighted row, stacked down the sheet. They can't share an
    # axis: the 44 rows span single-figure wagon counts to £1.9m a month, so on one
    # chart everything but the money is a flat line.
    wanted = list(chart_rows if chart_rows is not None
                  else range(hdr + 1, last_row + 1))
    # Laid out four across rather than one long column. Anchors start at column B
    # so every chart sits on the same uniform column width — column A is much wider
    # for the metric names and would throw the grid out.
    top = last_row + 3
    skipped = []
    placed = 0
    for rr in wanted:
        label = str(ws.cell(rr, 1).value)
        # Several rows only start part-way through — the category TOTAL EARNINGS
        # rows don't begin until 2026 — so each chart starts at its own first
        # populated month rather than carrying a run of empty columns.
        filled = [j for j in range(len(months))
                  if ws.cell(rr, 2 + j).value is not None]
        if not filled:
            skipped.append(label)
            continue
        c0, c1 = 2 + filled[0], 2 + filled[-1]
        ch = BarChart()
        ch.type = "col"
        ch.grouping = "clustered"
        ch.title = label
        ch.y_axis.title = None
        ch.x_axis.title = None
        ch.height, ch.width = CHART_H, CHART_W
        ch.legend = None
        ch.gapWidth = 55                       # chunkier bars, less white space
        ch.overlap = -10

        s = Series(Reference(ws, min_col=c0, max_col=c1, min_row=rr),
                   title_from_data=False, title=label)
        # every bar the same Fox navy, no outline
        s.graphicalProperties = GraphicalProperties(solidFill=FOX_NAVY)
        s.graphicalProperties.line.noFill = True
        # Excel shows the equation and R² by default when these are absent, so
        # they have to be turned off explicitly rather than just left unset.
        s.trendline = Trendline(trendlineType="linear", dispEq=False, dispRSqr=False)
        s.trendline.graphicalProperties = GraphicalProperties()
        s.trendline.graphicalProperties.line = LineProperties(solidFill=TREND_RED, w=22000)
        ch.series.append(s)
        ch.set_categories(Reference(ws, min_col=c0, max_col=c1, min_row=hdr))

        # Paul asked for the month's actual figure on each bar. Put it on the
        # series, not the chart, so the trendline doesn't get labelled too.
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.showSerName = False
        s.dLbls.showCatName = False
        s.dLbls.showLegendKey = False
        s.dLbls.showBubbleSize = False
        s.dLbls.numFmt = "#,##0"
        # Inside the bar top, in white — at outEnd the trendline runs straight
        # through the figures.
        s.dLbls.dLblPos = "inEnd"
        s.dLbls.txPr = RichText(
            p=[Paragraph(pPr=ParagraphProperties(
                defRPr=CharacterProperties(sz=750, b=True, solidFill="FFFFFF")),
                endParaRPr=None)])

        # quiet axes: thin grey gridlines, no clutter
        ch.y_axis.majorGridlines = ChartLines(
            spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=6000)))
        ch.x_axis.majorGridlines = None
        ch.y_axis.numFmt = "#,##0"
        for ax in (ch.x_axis, ch.y_axis):
            ax.majorTickMark = "none"
            ax.minorTickMark = "none"
            ax.spPr = GraphicalProperties(ln=LineProperties(solidFill="BFBFBF", w=6000))
            ax.txPr = RichText(
                p=[Paragraph(pPr=ParagraphProperties(
                    defRPr=CharacterProperties(sz=800, solidFill="595959")), endParaRPr=None)])
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        col = get_column_letter(2 + (placed % CHARTS_ACROSS) * COL_STEP)
        row = top + (placed // CHARTS_ACROSS) * ROW_STEP
        ws.add_chart(ch, f"{col}{row}")
        placed += 1
    if skipped:
        print(f"  no data, no chart: {', '.join(skipped)}")

    # Land on the latest month without freezing anything: just park the cursor
    # there so Excel scrolls it into view, and leave the charts unclipped.
    last_col = get_column_letter(1 + len(months))
    try:
        ws.sheet_view.selection[0].activeCell = f"{last_col}{hdr + 1}"
        ws.sheet_view.selection[0].sqref = f"{last_col}{hdr + 1}"
    except Exception:
        pass

    wb.save(out_path)
    return last_row, len(months)


AVG_SHEET = "MONTHLY AVERAGES"
DASH_SHEET = "DASHBOARD"


def averages_rows(path, sheet="DAILY", limit=250):
    """[(row, display name)] for the averages tab: each category's TOTAL row,
    the night-work total, and the fleet-wide TOTAL EARNINGS."""
    ws = openpyxl.load_workbook(path)[sheet]
    secs = section_map(path, sheet, limit)
    te = night = None
    out = []
    for r in range(1, limit):
        v = ws.cell(r, 1).value
        if not isinstance(v, str) or not v.strip():
            continue
        up = v.strip().upper()
        if up == "TOTAL EARNINGS" and te is None:
            te = r
        elif up == "NIGHT WORK" and te is None:
            night = r                              # keep the last one before TOTAL EARNINGS
        elif up.endswith(" TOTAL") and r in secs:
            out.append((r, secs[r]))
    if night:
        out.append((night, "Night work"))
    if te:
        out.append((te, "Whole fleet"))
    return out


def monthly_daily_averages(path, rows, start=START, sheet="DAILY"):
    """{(row,name): {month: mean of that row's non-zero daily figures}}.

    Zero days are left out — a category that didn't turn a wheel on a bank
    holiday shouldn't drag its average down."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    res = CellResolver(openpyxl.load_workbook(path)[sheet], ws)
    DATE_ROW, FIRST_DATA_COL = date_layout(ws)
    cols = {}
    for c in range(FIRST_DATA_COL, ws.max_column + 1):
        v = ws.cell(DATE_ROW, c).value
        if isinstance(v, dt.datetime) and v.date() >= start:
            cols.setdefault(dt.date(v.year, v.month, 1), []).append(c)
    months = sorted(cols)
    data = {}
    for r, name in rows:
        series = {}
        for mth in months:
            vals = [res.value(r, c) for c in cols[mth]]
            vals = [float(v) for v in vals if isinstance(v, (int, float)) and v]
            series[mth] = round(sum(vals) / len(vals), 2) if vals else None
        data[(r, name)] = series
    return months, data


def build_averages_workbook(months, data, out_path, company):
    """A standalone workbook holding just the MONTHLY AVERAGES table —
    daily average earnings per category per month, plain numbers, no charts."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = AVG_SHEET
    navy = "FF" + FOX_NAVY
    ws.sheet_view.showGridLines = False

    ws["A1"] = company
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color=navy)
    ws["A2"] = "Daily average earnings by month"
    ws["A2"].font = Font(name="Calibri", size=13, color=navy)
    ws["A3"] = ("Average per working day: each month's earnings divided by the days "
                "that category worked. Days with nothing recorded are left out.")
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="FF7F7F7F")
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 19
    ws.row_dimensions[4].height = 5

    hdr = 5
    rule = PatternFill("solid", fgColor="FF" + FOX_ORANGE)
    for j in range(max(len(months) + 1, 8)):
        ws.cell(hdr - 1, 1 + j).fill = rule
    thin = Side(style="thin", color="FFD9D9D9")
    edge = Border(bottom=thin)
    hc = ws.cell(hdr, 1, "Category")
    hc.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
    hc.fill = PatternFill("solid", fgColor=navy)
    hc.alignment = Alignment(vertical="center")
    for j, mth in enumerate(months):
        c = ws.cell(hdr, 2 + j, dt.datetime(mth.year, mth.month, 1))
        c.number_format = "mmm yy"
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=navy)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[hdr].height = 20

    band = PatternFill("solid", fgColor="FFF4F6F9")
    for i, ((r, name), series) in enumerate(data.items()):
        rr = hdr + 1 + i
        whole_fleet = name == "Whole fleet"
        nc = ws.cell(rr, 1, name)
        nc.font = Font(name="Calibri", size=9.5, bold=True, color="FF262626")
        nc.border = edge
        for j, mth in enumerate(months):
            cell = ws.cell(rr, 2 + j, series.get(mth))
            cell.number_format = "£#,##0;-£#,##0;\"–\""
            cell.font = Font(name="Calibri", size=9.5, bold=whole_fleet, color="FF262626")
            cell.alignment = Alignment(horizontal="right")
            cell.border = edge
        if whole_fleet:
            for j in range(len(months) + 1):
                ws.cell(rr, 1 + j).fill = PatternFill("solid", fgColor="FFEAEDF3")
        elif i % 2:
            for j in range(len(months) + 1):
                ws.cell(rr, 1 + j).fill = band

    ws.column_dimensions["A"].width = 24
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11
    wb.save(out_path)


def build_onto(book, out_dir, company="Fox Brothers (Lancashire)", averages=False,
               dashboard=None):
    """Rebuild the cover tab onto `book` and hand back the covered copy.

    Same filename as the original, so whoever owns the master carries on saving
    over their own file. With averages=True a MONTHLY AVERAGES tab (daily average
    per category per month — Paul asked, 07/08/2026) goes in as well. Pass a
    period_compare.compare() result as `dashboard` for a DASHBOARD tab. Returns
    None on failure — the daily report still goes out; a cover problem shouldn't
    hold up the numbers.
    """
    import lancs_inject
    try:
        # The masters round-trip through our own email, so they arrive already
        # carrying our tabs. Strip them first, exactly as the Leyland run always
        # did for the cover — injecting on top would orphan the old sheet and its
        # chart parts, which Excel reads as a damaged file.
        src = Path(book)
        Path(out_dir).mkdir(parents=True, exist_ok=True)
        bare = Path(out_dir) / ("_bare_" + src.name)
        if lancs_inject.strip(str(src), str(bare))["removed"]:
            book = bare
        bare2 = Path(out_dir) / ("_bare2_" + src.name)
        if lancs_inject.strip(str(book), str(bare2), sheet_name=AVG_SHEET)["removed"]:
            book = bare2
        bare3 = Path(out_dir) / ("_bare3_" + src.name)
        if lancs_inject.strip(str(book), str(bare3), sheet_name=DASH_SHEET)["removed"]:
            book = bare3
        rows = highlighted_rows(str(book)) or default_rows(str(book))
        months, data = monthly(str(book), rows)
        if not months:
            print("  ! cover: no months in range, skipping")
            return None
        tmp = Path(out_dir) / "_cover.xlsx"
        build_cover_workbook(months, data, str(tmp),
                             sections=section_map(str(book)), company=company)
        out = Path(out_dir) / src.name
        if out.resolve() == src.resolve():
            raise ValueError("out_dir would overwrite the source workbook")

        # The averages tab goes in FIRST: each inject marks its own sheet as the
        # selected one, and Excel opens on tabSelected — the cover must win.
        if averages:
            avg_rows = averages_rows(str(book))
            if avg_rows:
                m2, d2 = monthly_daily_averages(str(book), avg_rows)
                avg_tmp = Path(out_dir) / "_averages.xlsx"
                build_averages_workbook(m2, d2, str(avg_tmp), company=company)
                mid = Path(out_dir) / ("_mid_" + src.name)
                lancs_inject.inject(str(book), str(avg_tmp), str(mid),
                                    sheet_name=AVG_SHEET, first=False)
                book = mid
                avg_tmp.unlink(missing_ok=True)
                print(f"  averages tab: {len(avg_rows)} categories, {len(m2)} months")

        if dashboard:
            import period_compare
            dash_tmp = Path(out_dir) / "_dashboard.xlsx"
            period_compare.build_dashboard_workbook(dashboard, str(dash_tmp), company)
            mid2 = Path(out_dir) / ("_mid2_" + src.name)
            lancs_inject.inject(str(book), str(dash_tmp), str(mid2),
                                sheet_name=DASH_SHEET, first=False)
            book = mid2
            dash_tmp.unlink(missing_ok=True)
            print(f"  dashboard tab: {dashboard['this_label']} v "
                  f"{dashboard['prev_label']}, {dashboard['days']} trading days")

        lancs_inject.inject(str(book), str(tmp), str(out))
        tmp.unlink(missing_ok=True)
        for leftover in ("_bare_", "_bare2_", "_bare3_", "_mid_", "_mid2_"):
            (Path(out_dir) / (leftover + src.name)).unlink(missing_ok=True)
        print(f"  cover rebuilt: {len(rows)} rows, {len(months)} months "
              f"({months[0]:%b %y}..{months[-1]:%b %y})")
        return out
    except Exception as e:
        print(f"  ! cover sheet failed ({e}) — sending without it")
        return None


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("master")
    ap.add_argument("-o", "--out", required=True)
    ap.add_argument("--company", default="Fox Brothers (Lancashire)")
    a = ap.parse_args()
    rows = highlighted_rows(a.master) or default_rows(a.master)
    months, data = monthly(a.master, rows)
    secs = section_map(a.master)
    lr, nm = build_cover_workbook(months, data, a.out, sections=secs,
                                  company=a.company)
    print(f"{len(rows)} highlighted rows x {nm} months "
          f"({months[0]:%b %Y} .. {months[-1]:%b %Y}) -> {a.out}")
