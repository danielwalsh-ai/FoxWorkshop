"""
Combined Leyland + Lancashire daily wagon earnings, both operations on one page.

Individual figures for each, then a combined total, plus the running-month KPIs:
wagons utilised, best day of the month, month to date, and month to date measured
against full utilisation (every wagon on the list earning the £700 target).

Leaves a Gmail DRAFT. Nothing is ever sent. Default recipient is Daniel alone.

Usage:
    python combined_report.py --once                 # leave tonight's draft
    python combined_report.py --once --html out.html # build only, write to disk
    python combined_report.py --scheduled            # cron entry point
"""
import os
import re
import ssl
import time
import imaplib
import argparse
import datetime as dt
from pathlib import Path
from email.message import EmailMessage

import openpyxl

import wagon_auto
import lancs_auto
import lancs_inject
from wagon_master_fill import detect_layout, DATE_ROW as LEY_DATE_ROW
import lancs_data

HERE = Path(__file__).parent
ENV = lancs_auto.ENV          # same .env-or-real-environment loader as the others
TARGET = 700.0                     # the per-wagon daily target Paul measures against
TRADING_MIN = 0.25                 # below this share of the fleet out, not a trading day
DRAFT_TO = ["daniel.walsh@kfltd.uk"]

NAVY, ORANGE, GREEN, RED = "#24214a", "#eb941f", "#94c21f", "#b22222"
BLUE, GREY = "#00579e", "#f2f2f6"


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def money(v):
    if v is None:
        return "-"
    return f"£{v:,.0f}"


# ── reading the two masters ─────────────────────────────────────────────

def leyland_series(path):
    """[{date, earnings, on_list, utilised}] for every day the master carries."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["DAILY"]
    lay = detect_layout(ws)
    regs = [r for r in lay.vehicle_rows
            if isinstance(ws.cell(r, 1).value, str) and ws.cell(r, 1).value.strip()]
    out = []
    for c in range(1, ws.max_column + 1):
        d = ws.cell(LEY_DATE_ROW, c).value
        if not isinstance(d, dt.datetime):
            continue
        vals = [_num(ws.cell(r, c).value) for r in regs]
        # TOTAL EARNINGS is a formula, and the masters we build ourselves carry no
        # cached result for it, so the wagon rows are the only reliable source.
        earn = sum(v for v in vals if v)
        if not earn:
            continue
        out.append({
            "date": d.date(),
            "earnings": earn,
            "on_list": len(regs),
            "utilised": sum(1 for v in vals if v),
        })
    return sorted(out, key=lambda x: x["date"])


def lancs_series(path):
    m = lancs_data.LancsMaster(path)
    out = []
    for d, c in sorted(m.cols.items()):
        earn = _num(m.ws.cell(lancs_data.ROW["total_earnings"], c).value)
        if not earn:
            continue
        wag = m.wagons(d)
        out.append({
            "date": d,
            "earnings": earn,
            "on_list": len(wag),
            "utilised": sum(1 for w in wag if w["earned"]),
        })
    return out


# ── the maths ───────────────────────────────────────────────────────────

def combine(ley, lan):
    """Merge both series by date. A day present in only one still counts."""
    days = sorted({r["date"] for r in ley} | {r["date"] for r in lan})
    by_l = {r["date"]: r for r in ley}
    by_m = {r["date"]: r for r in lan}
    rows = []
    for d in days:
        a, b = by_l.get(d), by_m.get(d)
        rows.append({
            "date": d,
            "ley": a, "lan": b,
            "earnings": (a["earnings"] if a else 0) + (b["earnings"] if b else 0),
            "on_list": (a["on_list"] if a else 0) + (b["on_list"] if b else 0),
            "utilised": (a["utilised"] if a else 0) + (b["utilised"] if b else 0),
        })
    return rows


def summarise(rows):
    """Today, the month it falls in, and the running figures for that month."""
    if not rows:
        return None
    # The two operations report on different lags. Anchoring on the latest day
    # either one has would show a "combined" total missing a whole company, so
    # the headline day is the newest one both have reported.
    both = [r for r in rows if r["ley"] and r["lan"]]
    today = both[-1] if both else rows[-1]
    last = {k: max((r["date"] for r in rows if r[k]), default=None)
            for k in ("ley", "lan")}
    # Only days up to the headline, so the running totals match what is on show.
    inmonth = [r for r in rows
               if r["date"] <= today["date"]
               and (r["date"].year, r["date"].month) == (today["date"].year,
                                                         today["date"].month)]
    # A Saturday with 15 wagons out is not a day you would measure against 131
    # wagons at target, so it is kept out of the running figures and reported
    # separately. Real days run 80% or more of the list; weekends run 11%.
    month = [r for r in inmonth if r["utilised"] >= TRADING_MIN * r["on_list"]]
    skipped = [r for r in inmonth if r not in month]
    best = max(month, key=lambda r: r["earnings"])
    mtd = sum(r["earnings"] for r in month)
    # Full utilisation = every wagon on the list that day earning the target.
    potential = sum(r["on_list"] * TARGET for r in month)
    return {
        "today": today, "month": month, "best": best, "mtd": mtd,
        "mtd_ley": sum((r["ley"]["earnings"] if r["ley"] else 0) for r in month),
        "mtd_lan": sum((r["lan"]["earnings"] if r["lan"] else 0) for r in month),
        "potential": potential,
        "gap": potential - mtd,
        "pct": (mtd / potential * 100) if potential else None,
        "days": len(month),
        "last": last,
        "skipped": skipped,
    }


# ── the email ───────────────────────────────────────────────────────────

def tile(label, value, bg=NAVY, fg="#ffffff"):
    return (f'<td width="33%" style="background:{bg};padding:14px 12px;'
            f'border-radius:8px" align="center">'
            f'<div style="font:700 22px Arial,sans-serif;color:{fg}">{value}</div>'
            f'<div style="font:700 9px Arial,sans-serif;color:#ffffff;opacity:.75;'
            f'letter-spacing:1px;text-transform:uppercase;padding-top:4px">'
            f'{label}</div></td>')


def build_html(s):
    t = s["today"]
    d = t["date"]
    ley, lan = t["ley"], t["lan"]
    pct = f'{s["pct"]:.1f}%' if s["pct"] is not None else "-"
    util = (f'{t["utilised"]} of {t["on_list"]}' if t["on_list"] else "-")

    stale = [f"{n} runs to {s['last'][k]:%d %b}"
             for n, k in (("Leyland", "ley"), ("Lancashire", "lan"))
             if s["last"][k] and s["last"][k] != d]
    note = ("" if not stale else
            f'<tr><td style="padding:0 22px 10px"><div style="background:#fff4e0;'
            f'border-left:4px solid {ORANGE};padding:10px 12px;font:400 12px '
            f'Arial,sans-serif;color:#7a5b1e">Figures shown are for the latest day '
            f'both operations have reported. {"; ".join(stale)}.</div></td></tr>')

    sk = s["skipped"]
    skip_note = "" if not sk else (
        f'{len(sk)} non-trading day{"s" if len(sk) > 1 else ""} '
        f'({", ".join(f"{r["date"]:%d %b}" for r in sk)}) left out of the running '
        f'figures, {money(sum(r["earnings"] for r in sk))} between them. &nbsp;|&nbsp; ')

    rows = []
    for r in s["month"]:
        rows.append(
            f'<tr style="background:{GREY if len(rows) % 2 else "#ffffff"}">'
            f'<td style="padding:7px 9px">{r["date"]:%d %b}</td>'
            f'<td align="right" style="padding:7px 9px">'
            f'{money(r["ley"]["earnings"]) if r["ley"] else "-"}</td>'
            f'<td align="right" style="padding:7px 9px">'
            f'{money(r["lan"]["earnings"]) if r["lan"] else "-"}</td>'
            f'<td align="right" style="padding:7px 9px;font-weight:700">'
            f'{money(r["earnings"])}</td>'
            f'<td align="right" style="padding:7px 9px">'
            f'{r["utilised"]} / {r["on_list"]}</td>'
            f'<td align="right" style="padding:7px 9px">'
            f'{money(r["on_list"] * TARGET)}</td></tr>')

    return f"""<div style="background:#f0f1f4;padding:20px 0">
<table width="640" cellpadding="0" cellspacing="0" align="center"
       style="background:#fff;border-radius:10px;overflow:hidden;
              font-family:Arial,sans-serif">
 <tr><td style="background:{NAVY};border-bottom:4px solid {ORANGE};padding:20px 28px 8px">
   <table cellpadding="0" cellspacing="0" style="border-collapse:collapse"><tr>
     <td style="font:900 34px 'Arial Black',Arial,sans-serif;color:#fff;
                letter-spacing:2px;padding-right:8px">FOX</td>
     <td style="background:{ORANGE};border-radius:6px;padding:4px 10px;
                font:900 15px Arial,sans-serif;color:#fff;letter-spacing:1px">GROUP</td>
     <td style="font:900 20px Arial,sans-serif;color:{GREEN};padding-left:4px">&#187;</td>
   </tr></table></td></tr>
 <tr><td style="background:{NAVY};padding:0 28px 16px">
   <table width="100%"><tr>
     <td style="font:700 11px Arial,sans-serif;color:{GREEN};letter-spacing:2px">
       COMBINED DAILY WAGON EARNINGS &middot; LEYLAND + LANCASHIRE</td>
     <td align="right"><span style="background:{ORANGE};border-radius:12px;
       padding:4px 12px;font:700 12px Arial,sans-serif;color:#fff">
       {d:%d %b %Y}</span></td>
   </tr></table></td></tr>

 {note}
 <tr><td style="padding:18px 22px 6px">
   <table width="100%" cellspacing="6"><tr>
     {tile("Leyland today", money(ley["earnings"]) if ley else "-", BLUE)}
     {tile("Lancashire today", money(lan["earnings"]) if lan else "-", ORANGE)}
     {tile("Combined today", money(t["earnings"]), NAVY)}
   </tr></table></td></tr>

 <tr><td style="padding:0 22px 6px">
   <table width="100%" cellspacing="6"><tr>
     {tile("Wagons utilised", util, NAVY)}
     {tile(f'Best day &middot; {s["best"]["date"]:%d %b}', money(s["best"]["earnings"]), NAVY)}
     {tile(f'Month to date &middot; {s["days"]} days', money(s["mtd"]), NAVY)}
   </tr></table></td></tr>

 <tr><td style="padding:8px 22px 4px">
   <table width="100%" style="background:{GREY};border-radius:8px">
     <tr><td style="padding:14px 16px">
       <div style="font:700 10px Arial,sans-serif;color:#666;letter-spacing:1px;
                   text-transform:uppercase">Against full utilisation
            &middot; every wagon at {money(TARGET)}</div>
       <table width="100%" style="padding-top:8px"><tr>
         <td style="font:700 20px Arial,sans-serif;color:{NAVY}">{money(s["mtd"])}</td>
         <td align="center" style="font:400 13px Arial,sans-serif;color:#888">
           of {money(s["potential"])}</td>
         <td align="right" style="font:700 20px Arial,sans-serif;color:{RED}">
           {money(s["gap"])} short</td>
       </tr></table>
       <div style="background:#ddd;border-radius:6px;height:12px;margin-top:10px">
         <div style="background:{GREEN};width:{min(s['pct'] or 0, 100):.1f}%;
                     height:12px;border-radius:6px"></div></div>
       <div style="font:700 11px Arial,sans-serif;color:{NAVY};padding-top:6px">
         {pct} of full utilisation</div>
     </td></tr></table></td></tr>

 <tr><td style="padding:14px 22px 4px;font:700 12px Arial,sans-serif;
                color:{NAVY};letter-spacing:1px">THIS MONTH, DAY BY DAY</td></tr>
 <tr><td style="padding:0 22px 18px">
   <table width="100%" cellspacing="0" style="border-collapse:collapse;
          font:400 12px Arial,sans-serif;color:#333">
     <tr style="background:{NAVY};color:#fff;font-weight:700">
       <td style="padding:8px 9px">Date</td>
       <td align="right" style="padding:8px 9px">Leyland</td>
       <td align="right" style="padding:8px 9px">Lancashire</td>
       <td align="right" style="padding:8px 9px">Combined</td>
       <td align="right" style="padding:8px 9px">Utilised</td>
       <td align="right" style="padding:8px 9px">At {money(TARGET)}</td></tr>
     {"".join(rows)}
     <tr style="background:#fdebd0;font-weight:700">
       <td style="padding:9px">Month to date</td>
       <td align="right" style="padding:9px">{money(s["mtd_ley"])}</td>
       <td align="right" style="padding:9px">{money(s["mtd_lan"])}</td>
       <td align="right" style="padding:9px">{money(s["mtd"])}</td>
       <td align="right" style="padding:9px">-</td>
       <td align="right" style="padding:9px">{money(s["potential"])}</td></tr>
   </table></td></tr>

 <tr><td style="background:{NAVY};border-top:4px solid {ORANGE};padding:14px 22px;
                font:400 10px Arial,sans-serif;color:#9a97b5">
   {skip_note}Draft only. Built automatically by danielwalsh.ai</td></tr>
</table></div>"""


def leave_draft(html, subject, to=None, dry_run=False):
    """APPEND to [Gmail]/Drafts. Deliberately never SMTP — this must not send."""
    to = to or DRAFT_TO
    m = EmailMessage()
    m["From"] = f"Daniel Walsh <{ENV['GMAIL_USER']}>"
    m["To"] = ", ".join(to)
    m["Subject"] = subject
    m["Date"] = dt.datetime.now(dt.timezone.utc).strftime("%a, %d %b %Y %H:%M:%S +0000")
    m.set_content("This report is HTML. Open it in a client that renders HTML.")
    m.add_alternative(html, subtype="html")
    if dry_run:
        print(f"[DRY RUN] would leave a draft to {', '.join(to)}: {subject}")
        return
    M = imaplib.IMAP4_SSL("imap.gmail.com", timeout=90)
    M.login(ENV["GMAIL_USER"], ENV["GMAIL_APP_PASSWORD"])
    try:
        M.append('"[Gmail]/Drafts"', "\\Draft",
                 imaplib.Time2Internaldate(time.time()), m.as_bytes())
        print(f"  draft left for {', '.join(to)}: {subject}")
    finally:
        try:
            M.logout()
        except Exception:
            pass


# ── entry point ─────────────────────────────────────────────────────────

def run(leyland=None, lancashire=None, html_out=None, dry_run=False, to=None):
    import tempfile
    with tempfile.TemporaryDirectory() as tmp:
        if leyland:
            ley_path = Path(leyland)
        else:
            found, _subj, when = wagon_auto.fetch_master_from_gmail(tmp)
            if not found:
                print("No Leyland master in Gmail — nothing to build.")
                return 1
            bare = Path(tmp) / "ley.xlsx"
            lancs_inject.strip(str(found), str(bare))
            ley_path = bare
            print(f"Leyland master: {found.name} (sent {when:%Y-%m-%d %H:%M})")

        if lancashire:
            lan_path = Path(lancashire)
        else:
            new, _heads = lancs_auto.fetch_mel(tmp, 21, wanted=lambda h: True)
            if not new:
                print("No Lancashire master from Mel in the last 21 days.")
                return 1
            lan_path = Path(new[-1]["files"][-1])
            print(f"Lancashire master: {new[-1]['subject']} "
                  f"({new[-1]['date']:%Y-%m-%d %H:%M})")

        ley, lan = leyland_series(ley_path), lancs_series(lan_path)
        print(f"  Leyland {len(ley)} days to {ley[-1]['date'] if ley else '-'}, "
              f"Lancashire {len(lan)} days to {lan[-1]['date'] if lan else '-'}")
        s = summarise(combine(ley, lan))
        if not s:
            print("No overlapping earnings data.")
            return 1
        d = s["today"]["date"]
        print(f"  {d}  combined {money(s['today']['earnings'])}  "
              f"MTD {money(s['mtd'])} of {money(s['potential'])} "
              f"({s['pct']:.1f}%)")
        html = build_html(s)
        if html_out:
            Path(html_out).write_text(html, encoding="utf8")
            print(f"  wrote {html_out}")
            return 0
        leave_draft(html, f"Fox Group - Combined Wagon Earnings - {d:%d %b %Y}",
                    to=to, dry_run=dry_run)
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--once", action="store_true")
    ap.add_argument("--scheduled", action="store_true")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--html", help="write the HTML here instead of leaving a draft")
    ap.add_argument("--leyland", help="local master, instead of fetching from Gmail")
    ap.add_argument("--lancashire", help="local master, instead of fetching from Gmail")
    a = ap.parse_args()
    if not (a.once or a.scheduled or a.html):
        ap.error("give --once, --scheduled or --html")
    return run(a.leyland, a.lancashire, a.html, a.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
