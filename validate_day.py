"""
Validate the classification brain against a known-good report column.

Usage: python validate_day.py <csv> <report_date> <base_xlsx> <base_col>
e.g.   python validate_day.py SupplierTransaction.csv 2026-07-01 base_report.xlsx 2
"""
import sys
import datetime as dt
from openpyxl import load_workbook
from classify import process_csv, sheet_to_row, COVER_TOP_ROW, COVER_BOTTOM_ROW, TOP_SHEETS

csv_path, report_date_s, base_path, col = sys.argv[1], sys.argv[2], sys.argv[3], int(sys.argv[4])
report_date = dt.date.fromisoformat(report_date_s)

df = process_csv(csv_path, report_date)
print(f"CSV lines for {report_date}: {len(df)}   total = £{df['Cost'].sum():,.2f}\n")

wb = load_workbook(base_path, data_only=True)
cover = wb['Cover']

def basev(r, c):
    v = cover.cell(r, c).value
    try:
        return float(v) if v not in (None, "") and not str(v).startswith("=") else 0.0
    except (TypeError, ValueError):
        return 0.0

# ── division comparison ──
div_tot = df.groupby('Sheet')['Cost'].sum().to_dict()
print(f"DIVISIONS         {'computed':>12} {'known-good':>12} {'diff':>10}")
tc = tb = 0.0
for name, row in sheet_to_row.items():
    c = round(div_tot.get(name, 0.0), 2)
    b = round(basev(row, col), 2)
    tc += c; tb += b
    if c or b:
        flag = "" if abs(c - b) < 0.01 else "  <-- DIFF"
        print(f"{name:<17} {c:>12,.2f} {b:>12,.2f} {c-b:>10,.2f}{flag}")
print(f"{'TOTAL':<17} {tc:>12,.2f} {tb:>12,.2f} {tc-tb:>10,.2f}")

# ── area comparison (vehicle-list dependent — differences expected) ──
top_df = df[df['Sheet'].isin(TOP_SHEETS)]
area_tot = top_df.groupby('Area')['Cost'].sum().to_dict()
print(f"\nAREAS             {'computed':>12} {'known-good':>12} {'diff':>10}")
ac = ab = 0.0
for name, row in COVER_BOTTOM_ROW.items():
    c = round(area_tot.get(name, 0.0), 2)
    b = round(basev(row, col), 2)
    ac += c; ab += b
    if c or b:
        flag = "" if abs(c - b) < 0.01 else "  <-- DIFF"
        print(f"{name:<17} {c:>12,.2f} {b:>12,.2f} {c-b:>10,.2f}{flag}")
print(f"{'AREA TOTAL':<17} {ac:>12,.2f} {ab:>12,.2f} {ac-ab:>10,.2f}")

print(f"\nBalance check (computed): top £{tc - round(div_tot.get('Capital',0.0),2):,.2f} vs area £{ac:,.2f}")
