"""
Like-for-like period comparison for both wagon earnings reports.

Paul's ask (20/08/2026): "comparing all time the current period to previous
month period". He had done it by hand — "1st 13 days of July vs 1st 13 days of
august" — and listed each section up or down in thousands.

The window is the first N TRADING days of each month, not calendar days 1..N.
Calendar days do not reproduce his figures: August 2026 opens on a Saturday, so
a calendar window compares nine trading days against thirteen and every section
looks worse than it is. On first-N-trading-days his seven sections come back
within £1k each, which is his own rounding:

    Hooks -7k (he said -6), 8x4 -12k (-12), Alloys -12k (-11), Artics +19k
    (+19), Grabs -2k (-1), Sweepers +21k (+21), Sleepers -13k (-12)

A trading day is a weekday that actually carries earnings, so bank holidays and
days not yet filled drop out of both sides on their own.
"""
import datetime as dt

import openpyxl

UP, DOWN, FLAT = "#0b7a3b", "#b42318", "#666666"
UP_BG, DOWN_BG = "#e6f4ea", "#fdeceb"
NAVY = "#24214a"


def date_columns(ws, date_row, first_col):
    """{date: column} for every dated column on the sheet."""
    out = {}
    for c in range(first_col, ws.max_column + 2):
        v = ws.cell(date_row, c).value
        if isinstance(v, dt.datetime):
            out[v.date()] = c
    return out


def _block_sum(ws, first_row, last_row, col):
    total, seen = 0.0, False
    for r in range(first_row, last_row + 1):
        v = ws.cell(r, col).value
        if isinstance(v, (int, float)):
            total += float(v)
            seen = True
    return total, seen


def compare(path, blocks, upto=None, sheet="DAILY", date_row=2, first_col=3):
    """Current month's trading days so far against the same count last month.

    `blocks` is [(display name, first row, last row)]. Returns None when there
    is nothing to compare — a first-of-the-month run, or a master with no
    previous month on it.
    """
    ws = openpyxl.load_workbook(str(path), data_only=True)[sheet]
    cols = date_columns(ws, date_row, first_col)
    if not cols:
        return None

    trading = {}
    for d, c in cols.items():
        if d.weekday() >= 5:                 # Sat/Sun earnings sit in the weekday cost
            continue
        if any(_block_sum(ws, a, b, c)[1] for _n, a, b in blocks):
            trading[d] = c
    if not trading:
        return None

    upto = upto or max(trading)
    this_month = [d for d in sorted(trading)
                  if (d.year, d.month) == (upto.year, upto.month) and d <= upto]
    if not this_month:
        return None
    first = dt.date(upto.year, upto.month, 1)
    prev_end = first - dt.timedelta(days=1)
    # Same number of trading days, taken from the START of last month, so the
    # comparison is like for like however far into the month we are.
    last_month = [d for d in sorted(trading)
                  if (d.year, d.month) == (prev_end.year, prev_end.month)][:len(this_month)]
    if not last_month:
        return None

    rows = []
    for name, a, b in blocks:
        cur = sum(_block_sum(ws, a, b, trading[d])[0] for d in this_month)
        prv = sum(_block_sum(ws, a, b, trading[d])[0] for d in last_month)
        rows.append((name, cur, prv, cur - prv))
    tot_cur = sum(r[1] for r in rows)
    tot_prv = sum(r[2] for r in rows)
    return {
        "days": len(this_month),
        "prev_days": len(last_month),
        "this_label": f"{upto:%B}",
        "prev_label": f"{prev_end:%B}",
        "rows": rows,
        "total": ("Total", tot_cur, tot_prv, tot_cur - tot_prv),
        "upto": upto,
    }


def _delta(v):
    if abs(v) < 500:
        return FLAT, None, ("+" if v >= 0 else "-") + f"£{abs(v):,.0f}"
    if v > 0:
        return UP, UP_BG, f"+£{v/1000:,.1f}k"
    return DOWN, DOWN_BG, f"-£{abs(v)/1000:,.1f}k"


def to_html(cmp_):
    """The comparison as an email table."""
    if not cmp_:
        return ""
    def row(name, cur, prv, d, bold=False):
        colour, bg, text = _delta(d)
        weight = "font-weight:bold;" if bold else ""
        return (f'<tr><td style="{weight}">{name}</td>'
                f'<td align="right" style="{weight}">£{cur:,.0f}</td>'
                f'<td align="right" style="{weight}">£{prv:,.0f}</td>'
                f'<td align="right" style="{weight}background:{bg or "transparent"};'
                f'color:{colour};font-weight:bold">{text}</td></tr>')
    body = "\n".join(row(*r) for r in cmp_["rows"])
    tname, pname, n = cmp_["this_label"], cmp_["prev_label"], cmp_["days"]
    return f"""
      <p style="margin-bottom:4px"><b>{tname} against {pname} — first {n}
         trading days of each</b></p>
      <table cellpadding="6" style="border-collapse:collapse;font-size:14px">
        <tr style="background:{NAVY};color:#fff">
          <th align="left">Section</th><th align="right">{tname}</th>
          <th align="right">{pname}</th><th align="right">Change</th>
        </tr>
        {body}
        {row(*cmp_["total"], bold=True)}
      </table>"""


def to_text(cmp_):
    if not cmp_:
        return ""
    lines = [f"{cmp_['this_label']} against {cmp_['prev_label']} "
             f"— first {cmp_['days']} trading days of each"]
    for name, cur, prv, d in list(cmp_["rows"]) + [cmp_["total"]]:
        lines.append(f"{name}: £{cur:,.0f} vs £{prv:,.0f}  {_delta(d)[2]}")
    return "\n".join(lines)


def build_dashboard_workbook(cmp_, out_path, company):
    """A standalone workbook holding just the DASHBOARD sheet, for injecting.

    Same approach as the cover and averages tabs: built clean here, transplanted
    into the master's zip, because re-saving the master through openpyxl would
    destroy its charts.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DASHBOARD"
    navy = "FF" + NAVY.lstrip("#")
    ws.sheet_view.showGridLines = False

    ws["A1"] = company
    ws["A1"].font = Font(name="Calibri", size=20, bold=True, color=navy)
    ws["A2"] = (f"{cmp_['this_label']} against {cmp_['prev_label']}"
                f" — first {cmp_['days']} trading days of each")
    ws["A2"].font = Font(name="Calibri", size=13, color=navy)
    ws["A3"] = ("Trading days only, so bank holidays and days not yet filled drop "
                "out of both sides. Updated automatically each run.")
    ws["A3"].font = Font(name="Calibri", size=9, italic=True, color="FF7F7F7F")
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 19

    hdr = 5
    for j, title in enumerate(["Section", cmp_["this_label"], cmp_["prev_label"],
                               "Change", "Change %"]):
        c = ws.cell(hdr, 1 + j, title)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=navy)
        c.alignment = Alignment(horizontal="left" if j == 0 else "right",
                                vertical="center")
    ws.row_dimensions[hdr].height = 20

    thin = Side(style="thin", color="FFD9D9D9")
    edge = Border(bottom=thin)
    green, red = "FF0B7A3B", "FFB42318"
    band = PatternFill("solid", fgColor="FFF4F6F9")
    for i, (name, cur, prv, d) in enumerate(list(cmp_["rows"]) + [cmp_["total"]]):
        rr = hdr + 1 + i
        last = i == len(cmp_["rows"])
        ws.cell(rr, 1, name).font = Font(name="Calibri", size=9.5, bold=True,
                                         color="FF262626")
        for j, v in enumerate((cur, prv, d)):
            c = ws.cell(rr, 2 + j, round(v, 2))
            c.number_format = '£#,##0;-£#,##0;"–"'
            c.alignment = Alignment(horizontal="right")
            c.font = Font(name="Calibri", size=9.5, bold=last or j == 2,
                          color=(green if d >= 0 else red) if j == 2 else "FF262626")
        pct = ws.cell(rr, 5, (d / prv) if prv else None)
        pct.number_format = '+0.0%;-0.0%;"–"'
        pct.alignment = Alignment(horizontal="right")
        pct.font = Font(name="Calibri", size=9.5, bold=last,
                        color=green if d >= 0 else red)
        for j in range(5):
            ws.cell(rr, 1 + j).border = edge
            if last:
                ws.cell(rr, 1 + j).fill = PatternFill("solid", fgColor="FFEAEDF3")
            elif i % 2:
                ws.cell(rr, 1 + j).fill = band

    ws.column_dimensions["A"].width = 24
    for j in range(1, 5):
        ws.column_dimensions[get_column_letter(1 + j)].width = 13
    wb.save(out_path)
    return out_path


def daily_series(path, blocks, upto=None, sheet="DAILY", date_row=2, first_col=3):
    """[(date, total earnings)] for the current month's trading days so far.

    Resets on the 1st by construction: only the month `upto` falls in is read.
    """
    ws = openpyxl.load_workbook(str(path), data_only=True)[sheet]
    cols = date_columns(ws, date_row, first_col)
    days = []
    for d in sorted(cols):
        if d.weekday() >= 5:
            continue
        total = sum(_block_sum(ws, a, b, cols[d])[0] for _n, a, b in blocks)
        if total:
            days.append((d, total))
    if not days:
        return []
    upto = upto or max(d for d, _t in days)
    return [(d, t) for d, t in days
            if (d.year, d.month) == (upto.year, upto.month) and d <= upto]


def daily_chart_png(series, prev_avg=None, month_label="", width=7.6, height=2.9):
    """Bar per trading day so far this month, as PNG bytes for the email.

    A new bar appears each day the report runs and the whole thing starts again
    on the 1st, which is what Paul asked for. The dashed line is last month's
    average over its equivalent days, so the shape is judged against something.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from io import BytesIO

    if not series:
        return None
    days = [d for d, _t in series]
    vals = [t for _d, t in series]
    fig, ax = plt.subplots(figsize=(width, height), dpi=160)
    colours = ["#1A2646" if v >= (prev_avg or 0) else "#8894b0" for v in vals]
    ax.bar(range(len(vals)), vals, color=colours, width=0.68, zorder=3)
    if prev_avg:
        ax.axhline(prev_avg, color="#C00000", linestyle="--", linewidth=1.2, zorder=4)
        ax.annotate(f"last month avg £{prev_avg/1000:,.0f}k",
                    xy=(len(vals) - 0.4, prev_avg), xytext=(0, 4),
                    textcoords="offset points", ha="right", va="bottom",
                    fontsize=7.5, color="#C00000")
    ax.set_xticks(range(len(vals)))
    ax.set_xticklabels([f"{d.day}" for d in days], fontsize=7.5, color="#595959")
    ax.set_title(f"Daily earnings, {month_label} to date", fontsize=9.5,
                 color="#1A2646", loc="left", pad=8)
    ax.yaxis.set_major_formatter(
        matplotlib.ticker.FuncFormatter(lambda v, _p: f"{v/1000:,.0f}k"))
    ax.tick_params(axis="y", labelsize=7.5, colors="#595959", length=0)
    ax.tick_params(axis="x", length=0)
    ax.grid(axis="y", color="#D9D9D9", linewidth=0.6, zorder=0)
    ax.set_axisbelow(True)
    for side in ("top", "right", "left"):
        ax.spines[side].set_visible(False)
    ax.spines["bottom"].set_color("#BFBFBF")
    fig.tight_layout(pad=0.6)
    buf = BytesIO()
    fig.savefig(buf, format="png", transparent=False, facecolor="white")
    plt.close(fig)
    return buf.getvalue()
