"""
Fox Brothers (Lancashire) — data layer for the Daily Wagon Earnings Pack.

Everything the pack needs comes out of Mel Vose's "Daily wagon earning master
Mel.xlsx", except driver names, which are joined in from the load sheets the
portal already collects.

Sheet notes (Mel's is laid out differently to the Leyland master):
  * dates are on row 1, not row 2, and start at column D
  * column B carries the vehicle make (VOLVO / SCANIA) — this is what the
    Volvo-vs-Scania page is built on
  * each category block ends with AVERAGE / TOTAL EARNINGS / TOTAL WAGES /
    TOTAL FUEL / MARGIN rows
"""
import re
import datetime as dt
from pathlib import Path

import openpyxl

# Category blocks: label row -> (first wagon row, last wagon row, display name).
# Targets are the conditional-formatting bands from the master, as printed on
# Simon's pack.
CATEGORIES = [
    ("EV's",          3,   7,  "EV's",          600),
    ("SLEEPERS",      14,  28, "Sleepers",      700),
    ("8 W DAY CABS",  35,  73, "8W Day Cabs",   700),
    ("WHITE VOLVOS",  80,  88, "White Volvos",  633),
    ("ARTICS",        95, 113, "Artics",        750),
    ("MIDLAND",      120, 129, "Midland",       700),
    ("TIPWORX",      136, 147, "Tipworx",       600),
]

ROW = {
    "maintenance": 156, "breakdowns": 157, "no_driver": 158, "vors": 159,
    "total_off_road": 160, "self_drive": 161,
    "total_earnings": 166, "no_wagons": 167, "avg_per_wagon": 168,
    "est_vor_cost": 169, "overheads": 170, "ev_lease": 171, "fuel": 172,
    "wages": 173, "breakdown_cost": 174, "hurt_fitters": 175, "leyland_parts": 176,
    "rm_toll_misc": 177, "tyres": 178, "misc_office": 179, "vehicle_tax": 180,
    "profit": 181, "ebitda_costs": 182, "ebitda": 183,
    "rolling_5day": 186, "wagons_under_target": 187,
}
COST_ROWS = [  # page 6, in the order they stack
    ("Wages", 173), ("Fuel", 172), ("Overheads", 170), ("EV lease", 171),
    ("Breakdown", 174), ("Hurt fitters", 175), ("Leyland parts", 176),
    ("R&M / toll / misc", 177), ("Tyres", 178), ("Misc office", 179),
    ("Vehicle tax", 180),
]
DATE_ROW = 1
FIRST_DATA_COL = 4


def _num(v):
    return float(v) if isinstance(v, (int, float)) else None


def reg_key(reg):
    """Mel writes compound regs (NV66ZKC/AFY, PN71HFX-XRV, PN70 EFF-XRD) while the
    load sheets carry the current plate (PN25 AFY, PO75 XRV). They agree on the
    last token, so that's the join key."""
    if not reg:
        return ""
    t = re.split(r"[/\-]", str(reg).upper().replace(" ", ""))
    return t[-1].strip() if t else ""


class LancsMaster:
    def __init__(self, path):
        self.path = Path(path)
        wb = openpyxl.load_workbook(str(path), data_only=True)
        self.ws = wb["DAILY"]
        self.vor = wb["VOR "] if "VOR " in wb.sheetnames else None
        self.cols = {}
        for c in range(FIRST_DATA_COL, self.ws.max_column + 1):
            v = self.ws.cell(DATE_ROW, c).value
            if isinstance(v, dt.datetime):
                self.cols[v.date()] = c

    # ── days ────────────────────────────────────────────────────────
    def trading_days(self, upto, n=7):
        """The last n days that actually carry earnings, oldest first."""
        out = []
        for d in sorted((x for x in self.cols if x <= upto), reverse=True):
            if _num(self.ws.cell(ROW["total_earnings"], self.cols[d]).value):
                out.append(d)
            if len(out) == n:
                break
        return sorted(out)

    def val(self, key, day):
        c = self.cols.get(day)
        return _num(self.ws.cell(ROW[key], c).value) if c else None

    def cost_breakdown(self, day):
        c = self.cols.get(day)
        if not c:
            return {}
        return {name: (_num(self.ws.cell(r, c).value) or 0.0) for name, r in COST_ROWS}

    # ── wagons ──────────────────────────────────────────────────────
    def wagons(self, day):
        """[{reg, make, category, target, earned}] for every wagon with a row."""
        c = self.cols.get(day)
        out = []
        for _label, first, last, display, target in CATEGORIES:
            for r in range(first, last + 1):
                reg = self.ws.cell(r, 1).value
                if not (isinstance(reg, str) and reg.strip()):
                    continue
                out.append({
                    "reg": reg.strip(),
                    "key": reg_key(reg),
                    "make": (self.ws.cell(r, 2).value or "").strip().title() or None,
                    "category": display,
                    "target": target,
                    "earned": _num(self.ws.cell(r, c).value) if c else None,
                })
        return out

    def category_average(self, day, first, last):
        c = self.cols.get(day)
        if not c:
            return None
        vals = [_num(self.ws.cell(r, c).value) for r in range(first, last + 1)
                if isinstance(self.ws.cell(r, 1).value, str)
                and str(self.ws.cell(r, 1).value).strip()]
        vals = [v for v in vals if v]          # in-service only, as the sheet does
        return sum(vals) / len(vals) if vals else None

    def category_averages(self, day):
        return [{"name": disp, "target": tgt,
                 "avg": self.category_average(day, first, last)}
                for _l, first, last, disp, tgt in CATEGORIES]

    # ── off road ────────────────────────────────────────────────────
    def off_road(self, day):
        return {k: (self.val(k, day) or 0) for k in
                ("maintenance", "breakdowns", "no_driver", "vors", "total_off_road")}

    def off_road_window(self, days):
        tot = {"vors": 0, "maintenance": 0, "no_driver": 0, "breakdowns": 0}
        for d in days:
            o = self.off_road(d)
            for k in tot:
                tot[k] += o.get(k, 0) or 0
        tot["wagon_days_lost"] = sum(
            (self.off_road(d).get("total_off_road") or 0) for d in days)
        return tot

    # ── headline ────────────────────────────────────────────────────
    def kpis(self, day, window):
        wag = [w for w in self.wagons(day)]
        in_service = [w for w in wag if w["earned"]]
        under = [w for w in in_service if w["earned"] < w["target"]]
        return {
            "total_earnings": self.val("total_earnings", day),
            "ebitda": self.val("ebitda", day),
            "profit": self.val("profit", day),
            "rolling_5day": self.val("rolling_5day", day),
            "under_target": len(under),
            "in_service": len(in_service),
            "off_road": self.off_road(day).get("total_off_road"),
            "wagon_days_lost": self.off_road_window(window)["wagon_days_lost"],
        }


def load_drivers(reports_dir):
    """{reg_key: driver} from the portal's load sheets — the only source of driver
    names; Mel's master doesn't carry them."""
    import json
    out = {}
    for p in sorted(Path(reports_dir).glob("*.json")):
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        for s in d.get("sheets", []):
            drv, reg = (s.get("driver") or "").strip(), s.get("reg")
            if drv and reg:
                out[reg_key(reg)] = drv.upper()
    return out
