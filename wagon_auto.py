"""
Wagon master automation — Katie's run sheets in, updated master out to Paul.

Watches Daniel's Gmail for "Wagon earnings ..." emails — normally Katie Ward's, but
Emmy Duckworth covers when she is off and Paul forwards them on — fills every
run sheet attached (she batches days and re-sends corrections), pulls parts/tyres
from the workshop transaction report, tidies the formatting, then emails the
master to Paul Fox with Daniel copied in.

Design notes:
  * A run sheet's days come from C2/D2, not P2/Q2 — a weekend sheet carries
    Saturday in col C and Sunday in col D, and its P2 is empty.
  * Katie re-sends corrections ("Please use this one"). Within a batch the newest
    email wins; a day already in the master is replaced and the change reported.
  * Parts/tyres come from the workshop report's Cover tab, which is month-to-date,
    so one file covers any day of that month. No database needed.
  * Nothing reaches Paul unless every day passes validation. If anything looks
    wrong the run emails Daniel alone and leaves the stored master untouched.

Usage:
    python wagon_auto.py --seed "Daily wagon earnings 23rd July.xlsx"   # first time only
    python wagon_auto.py --once --dry-run    # check + build, send nothing
    python wagon_auto.py --once              # run now
    python wagon_auto.py --scheduled         # cron: quiet when there's nothing new
"""
import os
import re
import ssl
import sys
import json
import email
import shutil
import smtplib
import imaplib
import argparse
import zipfile
import tempfile
import datetime as dt
import urllib.request
from pathlib import Path
from email.message import EmailMessage

import openpyxl

from openpyxl.utils import get_column_letter

from wagon_master_fill import (fill_master, master_state, read_transaction_report,
                               SheetXmlEditor, detect_layout)
import tidy_master
import lancs_cover                  # Paul wants the same cover tab on both masters
import lancs_inject

HERE = Path(__file__).parent
STATE_DIR = HERE / "state"
STATE_FILE = STATE_DIR / "wagon_auto.json"
MASTER_FILE = STATE_DIR / "wagon_master.xlsx"

# Katie normally sends, but Emmy Duckworth covers when she's off and Paul forwards
# them on. Matching is by subject + a valid run sheet, so this list only has to say
# whose mail is worth opening.
SENDERS = {
    "katie.ward@hurtplant.co.uk",     # Katie Ward
    "emmy@hurtplant.co.uk",           # Emmy Duckworth — covers for Katie
    "paulfox@foxbrothers.co.uk",      # Paul forwards them on
}
# Deliberately NOT simon@foxbrothers.co.uk: he sends the *Lancashire* pack, subject
# "Daily Wagon Earnings Pack — ...", which substring-matches our subject search.
# Lancashire is a separate business with a separate fleet — zero registrations in
# common with Leyland — so a sheet is also checked against the master's own wagons.
FLEET_MATCH_MIN = 0.70            # a real Leyland sheet scores 100%; Lancashire 0%
# Our own outgoing subject also contains "Wagon earnings"; never re-ingest it.
OWN_SUBJECT = "master updated to"
SEND_TO = ["paulfox@foxbrothers.co.uk"]
SEND_CC = ["daniel.walsh@kfltd.uk",
           "reports@foxgroup.co"]      # Paul's shared reports box, set up 05/08/2026
MODEL = "claude-sonnet-5"

# Rows are never addressed by number — wagon_master_fill.detect_layout finds every
# row by its column-A label, so finance can insert rows without breaking the fill.
MAX_PLAUSIBLE_DAY = 250_000      # a decimal slip would sail past this
MIN_PLAUSIBLE_WEEKDAY = 20_000   # weekdays have run £75k–£97k all month
DATE_WINDOW_DAYS = 90            # a mistyped year must not land in last year's column
IMAP_TIMEOUT = 90                # never let a stalled socket wedge a scheduled run
MAX_SCAN = 12                    # newest N candidates only; don't trawl the mailbox
ALERT_QUIET_HOURS = 12           # don't re-send the same alert on every 15-min run
TOPUP_DAYS = 7                   # only covers a day filled before that evening's
                                 # workshop report existed — costs land next run


def load_env():
    env = dict(os.environ)
    p = HERE / ".env"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env


ENV = load_env()


def ordinal(n):
    suffix = "th" if 11 <= n % 100 <= 13 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def master_filename(last_date):
    return f"Daily wagon earnings {ordinal(last_date.day)} {last_date:%B}.xlsx"


# ── state ───────────────────────────────────────────────────────────
def load_state():
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {"processed_message_ids": [], "last_run": None}


def save_state(state):
    STATE_DIR.mkdir(exist_ok=True)
    state["last_run"] = dt.datetime.now().isoformat(timespec="seconds")
    STATE_FILE.write_text(json.dumps(state, indent=2), encoding="utf-8")


# ── Gmail ───────────────────────────────────────────────────────────
def _attachments(msg, tmpdir, want=".xlsx"):
    out = []
    for part in msg.walk():
        name = part.get_filename()
        if not name or not name.lower().endswith(want):
            continue
        if name.startswith("~$"):          # Excel lock file
            continue
        p = Path(tmpdir) / name
        p.write_bytes(part.get_payload(decode=True))
        out.append(p)
    return out


def _connect():
    """IMAP with a timeout. Without one a stalled socket blocks the run forever —
    which is exactly how the first scheduled run hung."""
    M = imaplib.IMAP4_SSL("imap.gmail.com", timeout=IMAP_TIMEOUT)
    M.login(ENV["GMAIL_USER"], ENV["GMAIL_APP_PASSWORD"])
    return M


def fetch_katie_emails(tmpdir, since_days=21, wanted=None):
    """Newest-last list of {id, subject, date, files} from Katie's run-sheet emails.

    Headers are fetched first and bodies only for mail `wanted` still needs. Run
    sheets are ~350KB each, so pulling three weeks of them every 15 minutes moves
    tens of MB for nothing — that is what hung the first scheduled run.
    """
    M = _connect()
    try:
        M.select("INBOX")
        since = (dt.date.today() - dt.timedelta(days=since_days)).strftime("%d-%b-%Y")
        # Search on subject and filter senders here — an IMAP OR chain over four
        # addresses is unreadable, and "FW: Wagon earnings - 24/07/2026" still matches.
        typ, data = M.search(None, f'(SINCE "{since}" SUBJECT "Wagon earnings")')
        heads = []
        for uid in data[0].split():
            typ, md = M.fetch(uid, "(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID DATE SUBJECT FROM)])")
            if not md or not md[0]:
                continue
            h = email.message_from_bytes(md[0][1])
            sender = email.utils.parseaddr(h.get("From", ""))[1].lower()
            subject = str(email.header.make_header(
                email.header.decode_header(h.get("Subject", ""))))
            if sender not in SENDERS or OWN_SUBJECT in subject.lower():
                continue
            raw_date = h.get("Date")
            heads.append({
                "uid": uid,
                "id": (h.get("Message-ID") or uid.decode()).strip(),
                "date": email.utils.parsedate_to_datetime(raw_date) if raw_date else None,
                "subject": subject,
                "from": sender,
            })
        heads = [h for h in heads if h["date"]]
        heads.sort(key=lambda m: m["date"])       # oldest first — newest correction wins
        todo = [h for h in heads if wanted is None or wanted(h)]

        found = []
        for h in todo:                            # only now pay for the attachments
            typ, md = M.fetch(h["uid"], "(RFC822)")
            msg = email.message_from_bytes(md[0][1])
            box = Path(tmpdir) / h["id"].strip("<>").replace("/", "_")[:60]
            box.mkdir(parents=True, exist_ok=True)
            files = _attachments(msg, box)
            if files:
                found.append({"id": h["id"], "subject": h["subject"],
                              "date": h["date"], "files": files,
                              "from": h.get("from", "")})
        return found, heads
    finally:
        try:
            M.logout()
        except Exception:
            pass


def report_date_of(name):
    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", name)
    if not m:
        return None
    try:
        return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def _covers(report_day, target):
    """The Cover tab is month-to-date as of the report's own day, so a report only
    holds `target` if it is from the same month and dated on or after it."""
    return (report_day is not None and report_day.year == target.year
            and report_day.month == target.month and report_day >= target)


def fetch_transaction_report(tmpdir, target):
    """Newest workshop report that actually covers `target`, from Gmail. None if absent.

    Matched on subject, not filename: Gmail's filename: operator does not reliably
    match fox_transaction_report_DD-MM-YYYY.xlsx, and switching to it silently lost
    the parts/tyres figures."""
    M = _connect()
    try:
        M.select('"[Gmail]/All Mail"')
        since = target.strftime("%d-%b-%Y")
        typ, data = M.search(None,
                             f'(SINCE "{since}" SUBJECT "Daily Workshop Spend Report")')
        for uid in reversed(data[0].split()[-MAX_SCAN:]):     # newest first
            typ, md = M.fetch(uid, "(RFC822)")
            msg = email.message_from_bytes(md[0][1])
            box = Path(tmpdir) / f"tx_{uid.decode()}"
            box.mkdir(parents=True, exist_ok=True)
            for f in _attachments(msg, box):
                if _covers(report_date_of(f.name), target):
                    return f
        return None
    finally:
        try:
            M.logout()
        except Exception:
            pass


def fetch_master_from_gmail(tmpdir, since_days=30):
    """Newest 'Daily wagon earnings *.xlsx' attachment in the mailbox.

    Saves hauling a 1.6MB workbook onto the server by hand — email it to yourself
    (or let a previous run's email to Paul serve) and seed straight from Gmail."""
    M = _connect()
    try:
        M.select('"[Gmail]/All Mail"')
        since = (dt.date.today() - dt.timedelta(days=since_days)).strftime("%Y/%m/%d")
        # Match the filename directly rather than every xlsx in the mailbox.
        typ, data = M.search(None, "X-GM-RAW",
                             f'"filename:\\"Daily wagon earnings\\" after:{since}"')
        for uid in reversed(data[0].split()[-MAX_SCAN:]):     # newest first
            typ, md = M.fetch(uid, "(RFC822)")
            msg = email.message_from_bytes(md[0][1])
            box = Path(tmpdir) / f"seed_{uid.decode()}"
            box.mkdir(parents=True, exist_ok=True)
            for f in _attachments(msg, box):
                if f.name.lower().startswith("daily wagon earnings"):
                    return f, msg.get("Subject", ""), email.utils.parsedate_to_datetime(
                        msg.get("Date"))
        return None, None, None
    finally:
        try:
            M.logout()
        except Exception:
            pass


def ensure_master(tmpdir, state):
    """Make sure there's a master to work from, rebuilding from Gmail if not.

    The container has no persistent volume, so a redeploy wipes state/. That's fine:
    every run emails the master to Paul, so Gmail already holds the latest copy.
    Recover it, and treat everything Katie sent before that email as already done —
    which reconstructs exactly the state the wipe destroyed."""
    if MASTER_FILE.exists():
        return True
    found, subj, when = fetch_master_from_gmail(tmpdir)
    if not found:
        return False
    STATE_DIR.mkdir(exist_ok=True)
    # What comes back from Gmail is the copy Paul got, so it already carries our
    # tabs. Work from a clean master and add them back fresh at send time.
    bare = Path(tmpdir) / "bare.xlsx"
    if lancs_inject.strip(str(found), str(bare))["removed"]:
        print("  removed the previous cover tab before filling")
        found = bare
    bare2 = Path(tmpdir) / "bare2.xlsx"
    if lancs_inject.strip(str(found), str(bare2),
                          sheet_name=lancs_cover.AVG_SHEET)["removed"]:
        print("  removed the previous averages tab before filling")
        found = bare2
    shutil.copy(found, MASTER_FILE)
    state["watermark"] = when.isoformat()
    # Persist immediately. If this run finds nothing new it returns early, and a later
    # run would otherwise see the master present, skip recovery, and — with no
    # watermark saved — reprocess every email Katie has sent.
    save_state(state)
    print(f"No stored master — recovered {found.name} from Gmail "
          f"(sent {when:%Y-%m-%d %H:%M}).")
    print("  Katie's earlier emails treated as already done.")
    return True


def transaction_report_for(date, tmpdir, _cache={}):
    """Prefer a locally-built report that covers the day; otherwise fetch from Gmail."""
    if date in _cache:
        return _cache[date]
    local = [(report_date_of(f.name), f) for f in HERE.glob("fox_transaction_report_*.xlsx")]
    local = [(d, f) for d, f in local if _covers(d, date)]
    if local:
        _cache[date] = max(local)[1]
        return _cache[date]
    try:
        _cache[date] = fetch_transaction_report(tmpdir, date)
    except Exception as e:
        print(f"  ! could not fetch a transaction report for {date}: {e}")
        _cache[date] = None
    return _cache[date]


# ── run sheets ──────────────────────────────────────────────────────
def days_in_sheet(path):
    """[(date, value_col)] — C2/D2 carry the day(s); a weekend sheet has both."""
    ws = openpyxl.load_workbook(path, data_only=True)["Wagons"]
    days = []
    for col, ref in ((3, "C2"), (4, "D2")):
        v = ws[ref].value
        if isinstance(v, dt.datetime):
            days.append((v.date(), col))
    if not days:                                    # older sheets: fall back to P2/Q2
        for ref in ("P2", "Q2"):
            v = ws[ref].value
            if isinstance(v, dt.datetime):
                days.append((v.date(), 3))
                break
    return days


def filename_date(name):
    """Katie names sheets DD.MM.YYYY — a useful cross-check on the in-sheet date."""
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", name)
    if not m:
        return None
    try:
        return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def master_regs(master):
    ws = openpyxl.load_workbook(str(master))['DAILY']
    lay = detect_layout(ws)
    regs = {str(ws.cell(r, 1).value or '').strip().upper() for r in lay.vehicle_rows}
    regs.discard('')
    return regs


def fleet_match(regs, path, value_col):
    """Fraction of a sheet's wagons that are ours. Leyland sheets score 1.0;
    a Lancashire sheet scores 0.0 — the two fleets share no registrations."""
    from wagon_master_fill import read_daily_file
    try:
        _, earnings, _ = read_daily_file(str(path), dt.date(2000, 1, 1), value_col)
    except Exception:
        return None
    if not earnings:
        return None
    return sum(1 for r in earnings if r in regs) / len(earnings)


def collect_days(messages, regs=None):
    """One entry per date across every attachment; a later email supersedes an earlier.

    Dates outside a sane window are set aside rather than filled — a mistyped year
    would otherwise land in a column from a previous year. Sheets whose wagons
    aren't ours (Lancashire) are refused outright."""
    floor = dt.date.today() - dt.timedelta(days=DATE_WINDOW_DAYS)
    today = dt.date.today()
    by_date, rejected = {}, []
    for msg in messages:                            # already oldest-first
        for f in msg["files"]:
            try:
                found = days_in_sheet(f)
            except Exception as e:
                rejected.append(f"{f.name}: could not be read ({e})")
                continue
            if not found:
                rejected.append(f"{f.name}: no date found in C2/D2 or P2/Q2")
            if found and regs:
                score = fleet_match(regs, f, found[0][1])
                if score is not None and score < FLEET_MATCH_MIN:
                    rejected.append(
                        f"{f.name}: only {score:.0%} of its wagons are on the Leyland "
                        f"master — this looks like another fleet (Lancashire is a "
                        f"separate business). Not filled.")
                    continue
            for d, col in found:
                if d < floor or d > today:
                    fn = filename_date(f.name)
                    hint = (f" — the file is named {fn:%d.%m.%Y}, so this looks like a "
                            f"typo in the sheet's date cell") if fn and fn != d else ""
                    rejected.append(
                        f"{f.name}: sheet date reads {d:%d %b %Y}, which is outside the "
                        f"last {DATE_WINDOW_DAYS} days{hint}. Day skipped.")
                    continue
                by_date[d] = {"date": d, "file": f, "value_col": col,
                              "subject": msg["subject"]}
    return [by_date[d] for d in sorted(by_date)], rejected


# ── costs catch-up ──────────────────────────────────────────────────
def topup_costs(master, tmpdir):
    """Fill parts/tyres for any recent weekday that has earnings but no costs.

    Katie's run sheets often arrive before that evening's workshop report, so a day
    can land with earnings only. This picks them up once the report exists, which
    keeps the master converging without anyone having to notice."""
    sheet_path, date_cols, _, _, _, ws, wsv, lay = master_state(str(master))
    # Deliberately narrow: chasing reports for months-old days costs a 90s IMAP
    # timeout each and never finds anything.
    floor = dt.date.today() - dt.timedelta(days=TOPUP_DAYS)
    todo = []
    for d, c in sorted(date_cols.items()):
        if d < floor or d > dt.date.today() or d.weekday() >= 5:
            continue                                   # weekend cost is averaged into Mon–Fri
        f = ws.cell(lay.total_earnings, c).value
        if not (isinstance(f, str) and f.startswith("=")):
            continue                                   # no earnings there yet
        v = ws.cell(lay.parts, c).value
        if isinstance(v, (int, float)) and v:
            continue                                   # already carries a real figure
        todo.append((d, c))
    donor = max((c for c in date_cols.values()
                 if isinstance(ws.cell(lay.parts, c).value, (int, float))
                 and ws.cell(lay.parts, c).value), default=None)
    if not todo or donor is None:
        return None, []

    z = zipfile.ZipFile(str(master))
    ed = SheetXmlEditor(z.read(sheet_path))
    donor_L = get_column_letter(donor)
    filled = []
    for d, c in todo:
        tx = transaction_report_for(d, tmpdir)
        if not tx:
            continue
        try:
            parts, workshop, tyres = read_transaction_report(str(tx), d)
        except Exception as e:
            print(f"  ! costs for {d}: {e}")
            continue
        if not parts and not tyres:
            continue
        L = get_column_letter(c)
        for row, val in ((lay.parts, parts), (lay.workshop, workshop), (lay.tyres, tyres)):
            ed.write(row, L, value=round(float(val), 2), style=ed.style_of(row, donor_L))
        filled.append(d)
    if not filled:
        z.close()
        return None, []
    out = Path(tmpdir) / "topped_up.xlsx"
    data = ed.tobytes()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zo:
        for it in z.infolist():
            zo.writestr(it, data if it.filename == sheet_path else z.read(it.filename))
    z.close()
    return out, filled


# ── build ───────────────────────────────────────────────────────────
def check_structure(master):
    """Confirm the master still reads as the Leyland layout.

    Rows are found by label, not number, so inserted rows are fine — this only
    trips if a load-bearing label has been renamed, deleted or duplicated."""
    ws = openpyxl.load_workbook(str(master))['DAILY']
    try:
        detect_layout(ws)
    except ValueError as e:
        return [str(e)]
    return []


def suggest_placement(master, daily_path, value_col, new_regs):
    """Say where a new wagon belongs, by looking at who it sits between on Katie's
    sheet and finding those neighbours in the master."""
    ws_m = openpyxl.load_workbook(str(master))['DAILY']
    where = {}
    for name, a, b in detect_layout(ws_m).blocks:
        for r in range(a, b + 1):
            v = str(ws_m.cell(r, 1).value or '').strip().upper()
            if v:
                where[v] = (r, name)
    ws_d = openpyxl.load_workbook(str(daily_path), data_only=True)['Wagons']
    order = []
    for r in range(3, ws_d.max_row + 1):
        v = ws_d.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            if v.strip() == 'VEHICLE':
                break
            order.append(v.strip().upper())
    out = []
    for reg in new_regs:
        if reg not in order:
            out.append(f"{reg}: not found on the sheet"); continue
        i = order.index(reg)
        above = next((where[order[j]] for j in range(i - 1, -1, -1)
                      if order[j] in where), None)
        below = next((where[order[j]] for j in range(i + 1, len(order))
                      if order[j] in where), None)
        if above and below and above[1] == below[1]:
            out.append(f"{reg}: belongs in {above[1]} — insert a row at {above[0] + 1} "
                       f"(it sits between {order[i-1]} and {order[i+1]} on Katie's sheet)")
        elif above:
            out.append(f"{reg}: sits after {order[i-1]} in {above[1]} — "
                       f"insert a row at {above[0] + 1}")
        else:
            out.append(f"{reg}: could not place automatically — check which block it "
                       f"belongs to on Katie's sheet")
    return out


def explain_fill_error(day, err, master=None, daily=None, value_col=3):
    """Turn a fill failure into something Daniel can act on."""
    m = re.search(r"would be lost\): \{(.*)\}", str(err))
    if not m:
        return f"{day:%a %d %b}: {err}"
    regs = dict(re.findall(r"'([A-Z0-9 ]+)':\s*([0-9.]+)", m.group(1)))
    lines = [f"{day:%a %d %b}: NEW WAGON on Katie's sheet, not yet in the master."]
    for reg, amt in regs.items():
        lines.append(f"    {reg} earned £{float(amt):,.2f}")
    if master and daily:
        try:
            for hint in suggest_placement(master, daily, value_col, list(regs)):
                lines.append(f"    {hint}")
        except Exception:
            pass
    lines.append("    The day is skipped for now so the earnings aren't lost. Insert a "
                 "row in Excel (Excel fixes the formulas itself), then send me the "
                 "master — the day fills automatically once the rows line up.")
    return "\n".join(lines)


def build(master, days, tmpdir):
    """Fill each day. A day that fails is skipped, not fatal — one new wagon must
    not hold up the days that are perfectly fine."""
    results, skipped, current = [], [], master
    for i, day in enumerate(days):
        tx = transaction_report_for(day["date"], tmpdir)
        out = Path(tmpdir) / f"step_{i:02d}.xlsx"
        try:
            r = fill_master(str(current), str(day["file"]), str(tx) if tx else None,
                            str(out), date_override=day["date"],
                            value_col=day["value_col"], replace=True)
        except Exception as e:
            skipped.append(explain_fill_error(day["date"], e, master=current,
                                              daily=day["file"],
                                              value_col=day["value_col"]))
            print(f"  {day['date']} ({day['date']:%a}) SKIPPED: {e}")
            continue
        r["subject"] = day["subject"]
        r["had_costs"] = tx is not None
        results.append(r)
        current = out
        flag = "revised" if r["replaced"] else "new"
        print(f"  {day['date']} ({day['date']:%a}) col {r['column']}  "
              f"£{r['expected_total_earnings']:,.2f}  {flag}")
    if not results:
        return None, [], skipped
    final = Path(tmpdir) / "final.xlsx"
    tidy_master.tidy(str(current), str(final))
    return final, results, skipped


def validate(results):
    problems = []
    today = dt.date.today()
    for r in results:
        d = dt.date.fromisoformat(r["date"])
        t = r["expected_total_earnings"]
        if d > today:
            problems.append(f"{d} is in the future")
        if not r["no_wagons"] or r["no_wagons"] < 50:
            problems.append(f"{d}: wagon count reads {r['no_wagons']!r}, expected ~109")
        if t > MAX_PLAUSIBLE_DAY:
            problems.append(f"{d}: £{t:,.0f} is implausibly high — check for a stray decimal")
        if d.weekday() < 5 and 0 < t < MIN_PLAUSIBLE_WEEKDAY:
            problems.append(f"{d}: weekday total is only £{t:,.0f}, expected £75k–£95k")
    return problems


def month_context(path, upto):
    """Facts for the commentary, computed here so the model never invents a number."""
    ws = openpyxl.load_workbook(path)["DAILY"]
    wv = openpyxl.load_workbook(path, data_only=True)["DAILY"]
    lay = detect_layout(ws)
    cols = {wv.cell(2, c).value.date(): c for c in range(3, wv.max_column + 2)
            if isinstance(wv.cell(2, c).value, dt.datetime)}
    days = []
    for d, c in sorted(cols.items()):
        if d.month != upto.month or d.year != upto.year or d > upto:
            continue
        f = ws.cell(lay.total_earnings, c).value
        if not (isinstance(f, str) and f.startswith("=")):
            continue
        total = sum(float(wv.cell(r, c).value) for r in lay.vehicle_rows
                    if isinstance(wv.cell(r, c).value, (int, float)))
        days.append((d, round(total, 2)))
    weekdays = [(d, t) for d, t in days if d.weekday() < 5]
    if not weekdays:
        return {}
    first5 = weekdays[:5]
    last5 = weekdays[-5:]
    return {
        "weekday_count": len(weekdays),
        "weekday_avg": round(sum(t for _, t in weekdays) / len(weekdays)),
        "best": max(weekdays, key=lambda x: x[1]),
        "worst": min(weekdays, key=lambda x: x[1]),
        "avg_first5": round(sum(t for _, t in first5) / len(first5)),
        "avg_last5": round(sum(t for _, t in last5) / len(last5)),
        "month_total": round(sum(t for _, t in days)),
    }


# ── commentary ──────────────────────────────────────────────────────
VOICE = """You write as Daniel Walsh, who consults for Fox Group and maintains the
wagon earnings master. You are writing 2-4 short sentences to Paul Fox, the CEO.

How Daniel writes to Paul (real examples):
  "All in the sheet up to Wednesday 8th. Tuesday was a monster. £97k total, best day
   on the sheet, and the 7 day average is now running at £84k against £76k at the
   start of the month. PN72EFE has been VOR all week."
  "Updated master."
  "Tidy this and added older than 2021 spend and hook spend by registration at the bottom."

Rules:
  - Blunt, Northern, factual. No greeting, no sign-off, no "please find attached".
  - Never invent a number. Use only the figures given to you.
  - Round to the nearest £k in prose (£85k, not £85,446.25).
  - Lead with what changed, then anything genuinely worth flagging.
  - If a day was revised, say so plainly.
  - No corporate filler. Never say "I hope this finds you well" or "Let me know if".
"""


def commentary(results, ctx):
    if not ENV.get("ANTHROPIC_API_KEY"):
        return ""
    lines = []
    for r in results:
        d = dt.date.fromisoformat(r["date"])
        bit = f"{d:%A} {d.day} {d:%B}: £{r['expected_total_earnings']:,.0f}"
        if r["replaced"] and r["previous_total"] is not None \
                and abs(r["previous_total"] - r["expected_total_earnings"]) >= 1:
            bit += f" (revised, was £{r['previous_total']:,.0f})"
        elif r["replaced"]:
            bit += " (re-sent, unchanged)"
        if not r["had_costs"]:
            bit += " [no parts/tyres available yet]"
        lines.append(bit)
    facts = "Days just added:\n" + "\n".join(f"  {l}" for l in lines)
    if ctx:
        facts += (f"\n\nMonth to date: {ctx['weekday_count']} weekdays, average "
                  f"£{ctx['weekday_avg']:,}. Best £{ctx['best'][1]:,.0f} on "
                  f"{ctx['best'][0]:%A} {ctx['best'][0].day}. Weakest "
                  f"£{ctx['worst'][1]:,.0f} on {ctx['worst'][0]:%A} {ctx['worst'][0].day}. "
                  f"Average of the first five weekdays £{ctx['avg_first5']:,}, "
                  f"of the last five £{ctx['avg_last5']:,}.")
    body = {"model": MODEL, "max_tokens": 400, "system": VOICE,
            "messages": [{"role": "user", "content": facts}]}
    try:
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=json.dumps(body).encode(),
            headers={"x-api-key": ENV["ANTHROPIC_API_KEY"],
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            method="POST")
        with urllib.request.urlopen(req, timeout=60) as r:
            data = json.loads(r.read().decode())
        return "".join(b.get("text", "") for b in data.get("content", [])).strip()
    except Exception as e:
        print(f"  ! commentary unavailable ({e}) — sending the table alone")
        return ""


# ── email ───────────────────────────────────────────────────────────
def rows_html(results):
    out = []
    for r in results:
        d = dt.date.fromisoformat(r["date"])
        note = ""
        if r["replaced"] and r["previous_total"] is not None \
                and abs(r["previous_total"] - r["expected_total_earnings"]) >= 1:
            note = f' <span style="color:#b45309">revised from £{r["previous_total"]:,.2f}</span>'
        cost = (f'<td align="right">£{r["parts"]:,.2f}</td>'
                f'<td align="right">£{r["tyres"]:,.2f}</td>') if r["had_costs"] \
            else '<td align="right">—</td><td align="right">—</td>'
        out.append(
            f'<tr><td>{d:%a} {d.day} {d:%b}{note}</td>'
            f'<td align="right">£{r["expected_total_earnings"]:,.2f}</td>{cost}</tr>')
    return "\n".join(out)


# Paul's names for the blocks, in the order he listed them (08/08/2026). Nights
# is a total, not an average — four wagons and a category row, an average there
# means nothing. Hooks on hire is one wagon on hire, deliberately not shown.
SECTION_DISPLAY = [("HOOKS", "Hooks", "avg"), ("8W", "8x4", "avg"),
                   ("ALLY BODY", "Alloys", "avg"), ("ARTICS", "Artics", "avg"),
                   ("GRABS", "Grabs", "avg"), ("SWEEPER", "Sweepers", "avg"),
                   ("8W SLEEPERS", "Sleepers", "avg"),
                   ("ARTIC - NIGHT WORK - BREAKDOWN", "Nights total", "sum")]


def section_averages(master, day):
    """[(display name, £value or None)] for one day — the block averages Paul
    forwards on. Read straight off the wagon rows, which are literal values."""
    wb = openpyxl.load_workbook(str(master), data_only=True)
    ws = wb["DAILY"]
    lay = detect_layout(ws)
    col = next((c for c in range(3, ws.max_column + 2)
                if isinstance(ws.cell(2, c).value, dt.datetime)
                and ws.cell(2, c).value.date() == day), None)
    if col is None:
        return []
    blocks = {name: (a, b) for name, a, b in lay.blocks}
    out = []
    for block, shown, how in SECTION_DISPLAY:
        a, b = blocks[block]
        vals = [float(v) for r in range(a, b + 1)
                if isinstance((v := ws.cell(r, col).value), (int, float))]
        if not vals:
            out.append((shown, None))
        elif how == "sum":
            out.append((shown, sum(vals)))
        else:
            out.append((shown, sum(vals) / len(vals)))
    return out


def sections_html(secs, day):
    rows = "\n".join(
        f'<tr><td>{shown}</td><td align="right">'
        + (f"£{v:,.0f}" if v is not None else "—") + "</td></tr>"
        for shown, v in secs)
    return f"""
      <p style="margin-bottom:4px"><b>Section averages — {day:%a} {day.day} {day:%b}</b></p>
      <table cellpadding="6" style="border-collapse:collapse;font-size:14px">
        <tr style="background:#24214a;color:#fff">
          <th align="left">Section</th><th align="right">Per wagon</th>
        </tr>
        {rows}
      </table>"""


def send(final_path, results, note, dry_run=False, to_me=False, workbook=None):
    last = max(dt.date.fromisoformat(r["date"]) for r in results)
    try:
        secs = section_averages(final_path, last)
    except Exception as e:
        print(f"  ! section averages failed ({e}) — sending without them")
        secs = []
    fname = master_filename(last)
    # The covered copy is what Paul gets; the plain one stays as our stored master.
    attach = workbook if workbook and Path(workbook).exists() else final_path
    cover = " The cover tab is the first thing you'll see." if attach != final_path else ""
    subject = f"Wagon earnings — master updated to {ordinal(last.day)} {last:%B}"
    html = f"""
    <div style="font-family:Arial,Helvetica,sans-serif;color:#24214a;font-size:15px">
      {'<p>' + note.replace(chr(10) + chr(10), '</p><p>') + '</p>' if note else ''}
      <table cellpadding="6" style="border-collapse:collapse;font-size:14px">
        <tr style="background:#24214a;color:#fff">
          <th align="left">Day</th><th align="right">Earnings</th>
          <th align="right">Parts</th><th align="right">Tyres</th>
        </tr>
        {rows_html(results)}
      </table>
      {sections_html(secs, last) if secs else ''}
      <p style="font-size:13px;color:#666">Master attached, updated to
         {ordinal(last.day)} {last:%B}.{cover}</p>
      <hr style="border:none;border-top:1px solid #ccc">
      <p style="font-size:12px;color:#888">Updated automatically from Katie's run sheets
         by danielwalsh.ai</p>
    </div>"""
    plain = (note + "\n\n" if note else "") + "\n".join(
        f"{dt.date.fromisoformat(r['date']):%a} {dt.date.fromisoformat(r['date']).day} "
        f"{dt.date.fromisoformat(r['date']):%b}: "
        f"£{r['expected_total_earnings']:,.2f}" for r in results)
    if secs:
        plain += (f"\n\nSection averages — {last:%a} {last.day} {last:%b}\n"
                  + "\n".join(f"{shown}: " + (f"£{v:,.0f}" if v is not None else "—")
                              for shown, v in secs))

    # --to-me proves the whole path (fill, tidy, commentary, SMTP, attachment)
    # without anything reaching Paul.
    to = [ENV["GMAIL_USER"]] if to_me else SEND_TO
    cc = [] if to_me else SEND_CC
    m = EmailMessage()
    m["From"] = f"Daniel Walsh <{ENV['GMAIL_USER']}>"
    m["To"] = ", ".join(to)
    if cc:
        m["Cc"] = ", ".join(cc)
    m["Subject"] = ("[TEST — would go to Paul] " if to_me else "") + subject
    m.set_content(plain)
    m.add_alternative(html, subtype="html")
    m.add_attachment(Path(attach).read_bytes(), maintype="application",
                     subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     filename=fname)
    if dry_run:
        print(f"\n[DRY RUN] would send to {', '.join(SEND_TO)} cc {', '.join(SEND_CC)}")
        print(f"[DRY RUN] subject: {subject}")
        print(f"[DRY RUN] attachment: {fname} "
              f"({Path(attach).stat().st_size/1e6:.1f} MB"
              + (", with cover)" if cover else ")"))
        print(f"\n{plain}\n")
        return fname
    ctx = ssl.create_default_context()
    with smtplib.SMTP("smtp.gmail.com", 587) as s:
        s.starttls(context=ctx)
        s.login(ENV["GMAIL_USER"], ENV["GMAIL_APP_PASSWORD"])
        s.send_message(m)
    print(f"  sent to {', '.join(to)}"
          + (f" (cc {', '.join(cc)})" if cc else "") + f" — {fname}")
    return fname


def notify_daniel(subject, body, dry_run=False, state=None, key=None):
    """Anything Daniel needs to see but Paul doesn't.

    A day held back for a new registration is retried every run so it heals the
    moment the master is fixed — `key` stops that turning into an alert every 15
    minutes. The same alert goes out at most once per ALERT_QUIET_HOURS."""
    if key is not None and state is not None:
        seen = state.setdefault("alerts", {})
        last = seen.get(key)
        if last:
            age = dt.datetime.now() - dt.datetime.fromisoformat(last)
            if age < dt.timedelta(hours=ALERT_QUIET_HOURS):
                print(f"  (already flagged {age.seconds // 3600}h ago, not re-sending: {key})")
                return
        seen[key] = dt.datetime.now().isoformat(timespec="seconds")
        save_state(state)
    if dry_run:
        print(f"\n[DRY RUN] would email Daniel — {subject}\n{body}\n")
        return
    m = EmailMessage()
    m["From"] = f"Daniel Walsh <{ENV['GMAIL_USER']}>"
    m["To"] = ENV["GMAIL_USER"]
    m["Subject"] = subject
    m.set_content(body)
    ctx = ssl.create_default_context()
    with smtplib.SMTP("smtp.gmail.com", 587) as s:
        s.starttls(context=ctx)
        s.login(ENV["GMAIL_USER"], ENV["GMAIL_APP_PASSWORD"])
        s.send_message(m)
    print(f"  emailed Daniel — {subject}")


def alert_daniel(problems, results, dry_run=False):
    """Something looks wrong — tell Daniel only, send Paul nothing."""
    lines = "\n".join(f"  - {p}" for p in problems)
    body = ("The wagon master was built but NOT sent to Paul — these checks failed:\n\n"
            f"{lines}\n\nThe stored master is unchanged and these emails will be "
            "picked up again on the next run once the run sheets are corrected.\n\n"
            "Days in this batch:\n"
            + "\n".join(f"  {r['date']}: £{r['expected_total_earnings']:,.2f} "
                        f"({r['no_wagons']} wagons)" for r in results))
    notify_daniel("Wagon master — held back, needs a look", body, dry_run)


# ── main ────────────────────────────────────────────────────────────
def run(dry_run=False, since_days=21, out_dir=None, to_me=False):
    state = load_state()
    with tempfile.TemporaryDirectory() as tmp:
        if not ensure_master(tmp, state):
            print("No stored master and none found in Gmail. Seed it first:\n"
                  '  python wagon_auto.py --seed "<path to current master>"')
            return 2

        moved = check_structure(MASTER_FILE)
        if moved:
            print("MASTER LAYOUT CHANGED — refusing to fill:")
            for m in moved:
                print(f"  - {m}")
            notify_daniel(
                "Wagon master — layout has shifted, filling paused",
                "The master's rows have moved, most likely a row inserted for a new "
                "wagon. Filling is paused because the row numbers are hard-coded and "
                "earnings would land in the wrong rows.\n\n"
                + "\n".join(f"  - {m}" for m in moved)
                + "\n\nSend me the master and I'll re-point the code at the new rows; "
                  "it's a small change and everything resumes.",
                dry_run, state=state, key="layout:" + "|".join(moved)[:150])
            return 1
        # Costs first: days filled before that evening's workshop report catch up here.
        master_in = MASTER_FILE
        topped, topped_days = topup_costs(MASTER_FILE, tmp)
        if topped:
            print("costs caught up for: "
                  + ", ".join(f"{d:%d %b}" for d in topped_days))
            master_in = topped
            if not dry_run:
                shutil.copy(topped, MASTER_FILE)
                master_in = MASTER_FILE

        mark = state.get("watermark")
        mark = dt.datetime.fromisoformat(mark) if mark else None
        done = set(state["processed_message_ids"])

        def is_new(h):
            return h["id"] not in done and (mark is None or h["date"] > mark)

        new, heads = fetch_katie_emails(tmp, since_days, wanted=is_new)
        if not new:
            print(f"Nothing new from Katie ({len(heads)} email(s) checked).")
            return 0
        print(f"{len(new)} new run-sheet email(s):")
        for m in new:
            print(f"  {m['date']:%Y-%m-%d %H:%M}  {m['subject']}")
            print(f"      from {m.get('from', '?')}  "
                  f"({len(m['files'])} attachment(s))")

        days, rejected = collect_days(new, master_regs(MASTER_FILE))
        if rejected:
            print("\nset aside:")
            for r in rejected:
                print(f"  - {r}")
        if not days:
            print("No usable run sheets in those emails.")
            if rejected:
                notify_daniel("Wagon run sheets — nothing usable",
                              "Katie's latest email(s) had no run sheet I could use:\n\n"
                              + "\n".join(f"  - {r}" for r in rejected), dry_run)
            if not dry_run:
                state["processed_message_ids"] += [m["id"] for m in new]
                save_state(state)
            return 0

        print(f"\nfilling {len(days)} day(s):")
        try:
            final, results, skipped = build(master_in, days, tmp)
        except Exception as e:
            print(f"BUILD FAILED: {e}")
            alert_daniel([f"build failed: {e}"], [], dry_run)
            return 1
        rejected += skipped
        if not results:
            # Every day was held back. Leave the emails unprocessed so this heals
            # by itself once the master is corrected, but only say so once.
            print("No day could be filled.")
            notify_daniel("Wagon master — a day is waiting on the master",
                          "Nothing could be filled from Katie's latest:\n\n"
                          + "\n".join(f"  - {r}" for r in rejected),
                          dry_run, state=state, key="|".join(sorted(rejected))[:200])
            return 1

        problems = validate(results)
        if problems:
            print("\nvalidation failed:")
            for p in problems:
                print(f"  - {p}")
            alert_daniel(problems, results, dry_run)
            return 1

        last = max(dt.date.fromisoformat(r["date"]) for r in results)
        note = commentary(results, month_context(final, last))
        # Its own directory — the covered copy keeps the master's filename, so
        # writing it alongside `final` would mean reading and writing one file.
        cover_dir = Path(tmp) / "covered"
        cover_dir.mkdir(exist_ok=True)
        covered = lancs_cover.build_onto(final, cover_dir,
                                         company="Fox Brothers (Leyland)",
                                         averages=True)
        fname = send(final, results, note, dry_run, to_me=to_me, workbook=covered)

        if rejected:      # good days went to Paul; the data issue is Daniel's to chase
            notify_daniel(
                "Wagon run sheets — one to check",
                "The master went to Paul with the days that were fine. These were set "
                "aside:\n\n" + "\n".join(f"  - {r}" for r in rejected)
                + "\n\nEach is retried every run, so it will go through by itself once "
                  "the master is updated or Katie re-sends a corrected sheet.",
                dry_run, state=state, key="|".join(sorted(rejected))[:200])

        if not dry_run:
            shutil.copy(final, MASTER_FILE)           # promote only after a clean send
            if not skipped:
                # A skipped day must stay unprocessed so it retries and self-heals.
                state["processed_message_ids"] += [m["id"] for m in new]
                state["processed_message_ids"] = state["processed_message_ids"][-500:]
            save_state(state)
        if out_dir:
            Path(out_dir).mkdir(parents=True, exist_ok=True)
            dest = Path(out_dir) / fname
            # what Paul actually receives — the covered copy when there is one
            shutil.copy(covered if covered and Path(covered).exists() else final, dest)
            print(f"  copied to {dest}")
    return 0


def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--seed", metavar="XLSX", help="store this file as the starting master")
    ap.add_argument("--seed-from-gmail", action="store_true",
                    help="seed from the newest 'Daily wagon earnings' attachment in Gmail")
    ap.add_argument("--once", action="store_true", help="run now")
    ap.add_argument("--scheduled", action="store_true", help="run from cron")
    ap.add_argument("--dry-run", action="store_true", help="build but send nothing")
    ap.add_argument("--since-days", type=int, default=21)
    ap.add_argument("--out-dir", help="also drop the finished master here")
    ap.add_argument("--to-me", action="store_true",
                    help="send the real email to Daniel only, not Paul")
    a = ap.parse_args()

    if a.seed or a.seed_from_gmail:
        STATE_DIR.mkdir(exist_ok=True)
        if a.seed_from_gmail:
            with tempfile.TemporaryDirectory() as tmp:
                found, subj, when = fetch_master_from_gmail(tmp)
                if not found:
                    print("No 'Daily wagon earnings *.xlsx' attachment found in Gmail.\n"
                          "Email the current master to yourself, then run this again.")
                    return 1
                print(f"Found {found.name}\n  from: {subj}\n  sent: {when:%Y-%m-%d %H:%M}")
                shutil.copy(found, MASTER_FILE)
                a.seed = found.name
        else:
            shutil.copy(a.seed, MASTER_FILE)
        st = load_state()
        # The seeded master is already up to date, so everything Katie has sent so far
        # counts as done — otherwise the first run would refill weeks of history.
        with tempfile.TemporaryDirectory() as tmp:
            # headers are enough to mark history as done — skip the attachments
            _, seen = fetch_katie_emails(tmp, a.since_days, wanted=lambda h: False)
        st["processed_message_ids"] = [m["id"] for m in seen]
        # Watermark as well as ids: if the sender list is ever widened, mail that was
        # previously invisible must not suddenly replay weeks of history.
        st["watermark"] = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")
        save_state(st)
        print(f"Seeded master from {a.seed} -> {MASTER_FILE}")
        print(f"Marked {len(seen)} existing email(s) from Katie as already done.")
        print("From here it only acts on new ones.")
        return 0
    if not (a.once or a.scheduled):
        ap.print_help()
        return 1
    return run(dry_run=a.dry_run, since_days=a.since_days, out_dir=a.out_dir,
               to_me=a.to_me)


if __name__ == "__main__":
    sys.exit(main())
