"""
Fox Group / Clive Hurt Plant Hire - Daily Wagon Earnings master sheet filler.

Fills one day column in the Daily Wagon Earnings master workbook from:
  1. The daily wagon earnings file (e.g. 02.07.2026.xlsx) - Wagons tab
  2. The fox transaction report (parts/costs)          - Cover tab

Confirmed mapping rules (02/07/2026 trial, signed off by DW):
  - Per-wagon earnings: Wagons tab, reg in col A, value in col C.
    Text statuses (VOR / MN / BD / ABSENCE) written verbatim so master COUNTIFs work.
  - Regs in master with no row in the daily file are left blank.
  - NO WAGONS: Wagons tab count cell (A121 in current template).
  - PARTS    = Cover 'Leyland Wagons' row, day column. Leyland ONLY, never Fox Wagons.
  - WORKSHOP = 0 unless explicitly provided.
  - TYRES    = Cover 'Tyres' row (row 21 area section), day column.
  - Overheads/Fuel/Wages/Plant hire/Tax/Other/EBITDA-costs: standing absolute formulas
    copied from the last populated column (point at the hidden monthly block IX:JC).
  - Missing days between last populated column and target date are left empty.

CRITICAL: the master has 200+ SharePoint external links and chart sheets.
A plain openpyxl load/save DESTROYS the cached link values and the charts.
All writes therefore happen at raw sheet-XML level inside the xlsx zip.
workbook.xml gets fullCalcOnLoad=1 and calcChain is dropped so Excel
recalculates everything cleanly on first open.

Usage:
    python wagon_master_fill.py MASTER.xlsx DAILY.xlsx [TRANSACTIONS.xlsx] -o OUT.xlsx
    (date is read from the daily file's Wagons!P2; override with --date DD/MM/YYYY)

Exit code non-zero and no output file if any verification check fails.
"""

import argparse
import json
import re
import shutil
import sys
import zipfile
from datetime import date, datetime, timedelta

from lxml import etree
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.formula.translate import Translator

NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
def q(t): return f'{{{NS}}}{t}'

EPOCH = date(1899, 12, 30)

# ---- master layout ----------------------------------------------------------
# Nothing is addressed by row number. Finance insert rows into the DAILY tab (37
# went in during August 2026), so every row is found by its column-A label when
# the workbook is opened. If a label vanishes or turns up twice, the run stops
# rather than write figures into the wrong rows.
DATE_ROW, DAYNAME_ROW = 2, 1

# Each block is a heading row, then one wagon per row, then its AVERAGE and TOTAL
# rows. Wagons therefore live between the heading and the first AVERAGE after it.
# The night-work block has no AVERAGE row; it ends at its final NIGHT WORK total.
BLOCK_HEADINGS = ['HOOKS', 'HOOKS ON HIRE', '8W', 'ALLY BODY', 'ARTICS', 'GRABS',
                  'SWEEPER', '8W SLEEPERS', 'ARTIC - NIGHT WORK - BREAKDOWN']
NIGHT_BLOCK = 'ARTIC - NIGHT WORK - BREAKDOWN'
UNIQUE_ANCHORS = {'total_earnings': 'TOTAL EARNINGS',
                  'no_wagons': 'NO WAGONS',
                  'parts': 'DAILY COSTINGS PARTS',
                  'workshop': 'DAILY COSTINGS WORKSHOP',
                  'tyres': 'DAILY COSTINGS TYRES'}

from collections import namedtuple
Layout = namedtuple('Layout', 'blocks vehicle_rows total_earnings no_wagons '
                              'parts workshop tyres')


def detect_layout(ws):
    """Read the DAILY sheet's layout off its column-A labels.

    `ws` is any openpyxl DAILY worksheet (values or formulas — only column A is
    read). Verified to reproduce the pre-insert row constants exactly on the
    29th July master. Raises ValueError if the sheet no longer looks like the
    Leyland master."""
    labels = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip():
            labels[r] = v.strip().upper()

    def find(lbl):
        hits = [r for r, v in labels.items() if v == lbl]
        if len(hits) != 1:
            raise ValueError(f"master layout: expected exactly one {lbl!r} row "
                             f"in column A of DAILY, found {len(hits)}")
        return hits[0]

    anchors = {k: find(v) for k, v in UNIQUE_ANCHORS.items()}
    te = anchors['total_earnings']
    blocks, vehicle_rows = [], []
    for name in BLOCK_HEADINGS:
        h = find(name)
        if name == NIGHT_BLOCK:
            night_totals = [r for r, v in labels.items()
                            if v == 'NIGHT WORK' and h < r < te]
            if not night_totals:
                raise ValueError("master layout: no NIGHT WORK total row between "
                                 "the night-work heading and TOTAL EARNINGS")
            end = max(night_totals) - 1
        else:
            avgs = [r for r, v in labels.items() if r > h and v.endswith(' AVERAGE')]
            if not avgs:
                raise ValueError(f"master layout: no AVERAGE row after the "
                                 f"{name!r} heading")
            end = min(avgs) - 1
        blocks.append((name, h + 1, end))
        vehicle_rows.extend(range(h + 1, end + 1))
    n = len(vehicle_rows)
    if not 100 <= n <= 200:
        raise ValueError(f"master layout: found {n} wagon rows, expected 100-200 "
                         f"— the DAILY tab does not look right")
    return Layout(blocks, vehicle_rows, te, anchors['no_wagons'],
                  anchors['parts'], anchors['workshop'], anchors['tyres'])

# ---- source extraction -----------------------------------------------------

def read_daily_file(path, date_override=None, value_col=3):
    """Return (report_date, {REG: value}, wagon_count) from the daily wagon file.

    value_col/date_override let a weekend sheet (Sat in col C, Sun in col D,
    date in Q2 not P2) be filled one day at a time."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Wagons']
    report_date = date_override or ws['P2'].value
    if isinstance(report_date, datetime):
        report_date = report_date.date()
    mapping, count = {}, None
    for r in range(3, ws.max_row + 1):
        a, c = ws.cell(row=r, column=1).value, ws.cell(row=r, column=value_col).value
        if isinstance(a, str) and a.strip() == 'VEHICLE':
            break  # reached the TARGET table at the bottom - stop
        if isinstance(a, str) and a.strip():
            mapping[a.strip().upper()] = c
        elif isinstance(a, (int, float)) and ws.cell(row=r + 2, column=1).value is None \
                and ws.cell(row=r, column=value_col).value is None:
            count = int(a)  # standalone grand-count row (A121 style)
    if count is None:  # fallback: cell two below last category subtotal
        for r in range(ws.max_row, 3, -1):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, (int, float)) and v > 50:
                count = int(v); break
    return report_date, mapping, count


def read_transaction_report(path, report_date):
    """Return (parts, workshop, tyres) for the given date from the Cover tab.
    PARTS = Leyland Wagons row ONLY. TYRES = 'Tyres' area row. WORKSHOP = 0."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Cover']
    day_col = 1 + report_date.day  # col B = 1st of the month
    hdr = str(ws.cell(row=2, column=day_col).value or '')
    if not hdr.startswith(str(report_date.day)):
        raise ValueError(f"Cover day header mismatch: expected day {report_date.day}, "
                         f"found {hdr!r} in col {get_column_letter(day_col)}")
    leyland = tyres = None
    for r in range(3, 40):
        label = str(ws.cell(row=r, column=1).value or '').strip().lower()
        if label == 'leyland wagons':
            leyland = ws.cell(row=r, column=day_col).value
        elif label == 'tyres':
            tyres = ws.cell(row=r, column=day_col).value
    if leyland is None or tyres is None:
        raise ValueError("Could not locate 'Leyland Wagons' and/or 'Tyres' rows on Cover tab")
    return float(leyland), 0.0, float(tyres)

# ---- master helpers --------------------------------------------------------

def master_state(path):
    """Return (sheet_xml_path, date_cols, last_date_col, last_populated_col,
    row->reg map, ws, wsv, layout)."""
    wb = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    ws, wsv = wb['DAILY'], wbv['DAILY']
    layout = detect_layout(ws)
    date_cols, last_date_col, last_pop_col = {}, None, None
    for c in range(3, ws.max_column + 2):
        d = wsv.cell(row=DATE_ROW, column=c).value
        if isinstance(d, datetime):
            date_cols[d.date()] = c
            last_date_col = c
        f = ws.cell(row=layout.total_earnings, column=c).value
        if f not in (None, ''):
            last_pop_col = c
    regs = {}
    for r in layout.vehicle_rows:
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip():
            regs[r] = a.strip().upper()
    z = zipfile.ZipFile(path)
    wbxml = z.read('xl/workbook.xml').decode()
    rels = z.read('xl/_rels/workbook.xml.rels').decode()
    rid = re.search(r'<sheet[^>]*name="DAILY"[^>]*r:id="(rId\d+)"', wbxml).group(1)
    target = re.search(rf'<Relationship[^>]*Id="{rid}"[^>]*Target="([^"]+)"', rels).group(1)
    z.close()
    return 'xl/' + target, date_cols, last_date_col, last_pop_col, regs, ws, wsv, layout


def col_for_date(date_cols, last_date_col, target):
    """Exact date match in row 2 wins; otherwise extend past the last date column.
    The master's date row has occasional skipped days, so never assume consecutiveness."""
    if target in date_cols:
        return date_cols[target], False
    last_date = max(date_cols)
    if target <= last_date:
        raise ValueError(f"{target} falls inside the existing date range but has no "
                         f"column in row 2 - master dates skip this day; check manually")
    return last_date_col + (target - last_date).days, True

# ---- XML writer ------------------------------------------------------------

class SheetXmlEditor:
    def __init__(self, xml_bytes):
        self.tree = etree.fromstring(xml_bytes)
        self.sheetData = self.tree.find(q('sheetData'))
        self.rowmap = {int(r.get('r')): r for r in self.sheetData.findall(q('row'))}

    def _cell(self, rowel, ref):
        for c in rowel.findall(q('c')):
            if c.get('r') == ref:
                return c
        return None

    def style_of(self, rownum, colletter):
        rel = self.rowmap.get(rownum)
        if rel is None: return None
        c = self._cell(rel, f'{colletter}{rownum}')
        return c.get('s') if c is not None else None

    def write(self, rownum, colletter, *, value=None, formula=None, text=None, style=None):
        rel = self.rowmap.get(rownum)
        if rel is None:
            rel = etree.SubElement(self.sheetData, q('row')); rel.set('r', str(rownum))
            self.rowmap[rownum] = rel
        ref = f'{colletter}{rownum}'
        old = self._cell(rel, ref)
        if old is not None: rel.remove(old)
        c = etree.Element(q('c')); c.set('r', ref)
        if style: c.set('s', style)
        if formula is not None:
            etree.SubElement(c, q('f')).text = formula.lstrip('=')
        elif text is not None:
            c.set('t', 'inlineStr')
            etree.SubElement(etree.SubElement(c, q('is')), q('t')).text = text
        elif value is not None:
            etree.SubElement(c, q('v')).text = repr(value) if isinstance(value, float) else str(value)
        col_idx = column_index_from_string(colletter)
        for existing in rel.findall(q('c')):
            ecol = column_index_from_string(re.match(r'([A-Z]+)', existing.get('r')).group(1))
            if ecol > col_idx:
                existing.addprevious(c); return
        rel.append(c)

    def fix_dimension(self, colletter):
        dim = self.tree.find(q('dimension'))
        if dim is not None:
            dim.set('ref', re.sub(r':[A-Z]+(\d+)$', rf':{colletter}\1', dim.get('ref')))

    def extend_conditional_formatting(self, target_col):
        """Pull the sheet's conditional-formatting frontier out to a new column.

        The green over-target shading is conditional formatting whose ranges end
        wherever finance last dragged them (column OP, 31 July, when Paul spotted
        3rd/4th August uncoloured). Any range that reaches that shared frontier is
        treated as live and extended to the new column; ranges that stop short are
        historical patches and are left alone."""
        cfs = self.tree.findall(q('conditionalFormatting'))
        ends = []
        for cf in cfs:
            for part in (cf.get('sqref') or '').split():
                ends.append(range_boundaries(part)[2])
        if not ends:
            return 0
        frontier = max(ends)
        if target_col <= frontier:
            return 0
        n = 0
        for cf in cfs:
            parts, changed = [], False
            for part in (cf.get('sqref') or '').split():
                mn_c, mn_r, mx_c, mx_r = range_boundaries(part)
                if mx_c == frontier:
                    part = (f"{get_column_letter(mn_c)}{mn_r}:"
                            f"{get_column_letter(target_col)}{mx_r}")
                    changed = True
                    n += 1
                parts.append(part)
            if changed:
                cf.set('sqref', ' '.join(parts))
        return n

    def mirror_format(self, src_letter, tgt_letter):
        """Give the target column the source column's cell styles on any row the
        fill left empty — so a newly-extended column keeps the sheet's borders."""
        for rownum, rowel in list(self.rowmap.items()):
            src = self._cell(rowel, f'{src_letter}{rownum}')
            if src is None:
                continue
            s = src.get('s')
            if s is None:
                continue
            if self._cell(rowel, f'{tgt_letter}{rownum}') is None:
                self.write(rownum, tgt_letter, style=s)   # empty, styled (borders)

    def tobytes(self):
        return etree.tostring(self.tree, xml_declaration=True, encoding='UTF-8', standalone=True)

# ---- main fill -------------------------------------------------------------

def _column_earnings(wsv, col, vehicle_rows):
    """Sum the wagon rows of one date column (ignores 'VOR' and other text)."""
    total = 0.0
    for r in vehicle_rows:
        v = wsv.cell(row=r, column=col).value
        if isinstance(v, (int, float)):
            total += float(v)
    return round(total, 2)


def fill_master(master, daily, transactions, out, date_override=None, value_col=3,
                replace=False):
    """Fill one day column.  replace=True overwrites a day that is already
    populated (Katie re-sends corrections) and reports the before/after total."""
    report_date, earnings, wagon_count = read_daily_file(daily, date_override, value_col)
    parts = workshop = tyres = None
    if transactions:
        parts, workshop, tyres = read_transaction_report(transactions, report_date)

    (sheet_path, date_cols, last_date_col, last_pop_col, regs, ws, wsv,
     layout) = master_state(master)
    target_col, needs_extension = col_for_date(date_cols, last_date_col, report_date)
    # Per-column guard: refuse only if the TARGET column itself already holds data.
    # (The old `target_col <= last_pop_col` test blocked backfilling an empty gap that
    #  sits before a later populated column — e.g. filling 23-30 Jun when 2 Jul is done.)
    was_populated = ws.cell(row=layout.total_earnings, column=target_col).value not in (None, '')
    previous_total = _column_earnings(wsv, target_col, layout.vehicle_rows) if was_populated else None
    if was_populated and not replace:
        raise ValueError(f"{report_date} maps to column {get_column_letter(target_col)} "
                         f"which is already populated - refusing to overwrite")
    # Source column for styles + standing formulas = nearest populated column strictly
    # BEFORE the target (the adjacent real day), not the global last populated column.
    source_col = next((c for c in range(target_col - 1, 2, -1)
                       if ws.cell(row=layout.total_earnings, column=c).value not in (None, '')),
                      last_pop_col)
    TL = get_column_letter(target_col)
    src_col_letter = get_column_letter(source_col)

    fills, unmatched_master, matched = {}, [], set()
    for r, reg in regs.items():
        if reg in earnings:
            fills[r] = earnings[reg]; matched.add(reg)
        else:
            unmatched_master.append(reg)
    lost = {k: v for k, v in earnings.items() if k not in matched
            and isinstance(v, (int, float)) and v != 0}
    if lost:
        raise ValueError(f"Daily file regs with earnings not present in master (would be lost): {lost}")

    expected_total = round(sum(v for v in fills.values() if isinstance(v, (int, float))), 2)

    zin = zipfile.ZipFile(master)
    ed = SheetXmlEditor(zin.read(sheet_path))

    # extend dates + day names past the last existing date column if needed
    if needs_extension:
        s2 = ed.style_of(DATE_ROW, get_column_letter(last_date_col))
        s1 = ed.style_of(DAYNAME_ROW, get_column_letter(last_date_col))
        d, c = max(date_cols), last_date_col
        while d < report_date:
            d += timedelta(days=1); c += 1
            cl = get_column_letter(c)
            ed.write(DATE_ROW, cl, value=(d - EPOCH).days, style=s2)
            ed.write(DAYNAME_ROW, cl, formula=f'TEXT({cl}2, "dddd")', style=s1)

    # On a replace, clear any wagon row the corrected sheet no longer carries —
    # otherwise a reg dropped from the revision would keep yesterday's figure.
    if was_populated:
        for r in layout.vehicle_rows:
            if r not in fills:
                ed.write(r, TL, style=ed.style_of(r, src_col_letter))

    for r, v in fills.items():
        st = ed.style_of(r, src_col_letter)
        if isinstance(v, str):
            ed.write(r, TL, text=v, style=st)
        else:
            ed.write(r, TL, value=v, style=st)

    # Standing formulas are whatever the adjacent populated day carries, so rows
    # finance add later (they inserted 37 in Aug 2026) start copying across the
    # day after they first put a formula in. Value rows are excluded: a wagon or
    # costs cell must never be turned into a formula.
    skip = set(layout.vehicle_rows) | {layout.no_wagons, layout.parts,
                                       layout.workshop, layout.tyres,
                                       DATE_ROW, DAYNAME_ROW}
    formula_rows = [r for r in range(3, ws.max_row + 1)
                    if r not in skip
                    and isinstance(ws.cell(row=r, column=source_col).value, str)
                    and ws.cell(row=r, column=source_col).value.startswith('=')]
    if len(formula_rows) < 35:
        raise ValueError(f"only {len(formula_rows)} formula rows found in source "
                         f"column {src_col_letter} - expected 40+; refusing to fill "
                         f"from a column that looks half-built")
    for r in formula_rows:
        f = ws.cell(row=r, column=source_col).value
        ed.write(r, TL, formula=Translator(f, origin=f'{src_col_letter}{r}').translate_formula(f'{TL}{r}'),
                 style=ed.style_of(r, src_col_letter))

    ed.write(layout.no_wagons, TL, value=wagon_count, style=ed.style_of(layout.no_wagons, src_col_letter))
    if parts is not None:
        ed.write(layout.parts, TL, value=round(parts, 2), style=ed.style_of(layout.parts, src_col_letter))
        ed.write(layout.workshop, TL, value=round(workshop, 2), style=ed.style_of(layout.workshop, src_col_letter))
        ed.write(layout.tyres, TL, value=round(tyres, 2), style=ed.style_of(layout.tyres, src_col_letter))
    ed.mirror_format(src_col_letter, TL)   # borders on any rows the fill left empty
    ed.fix_dimension(TL)
    cf_extended = ed.extend_conditional_formatting(target_col)
    if cf_extended:
        print(f"  conditional formatting: {cf_extended} range(s) extended to {TL}")

    # repackage preserving everything else; force recalc on open, drop calcChain
    wb_tree = etree.fromstring(zin.read('xl/workbook.xml'))
    calcPr = wb_tree.find(q('calcPr'))
    if calcPr is None:
        calcPr = etree.SubElement(wb_tree, q('calcPr'))
    calcPr.set('fullCalcOnLoad', '1')
    ct = zin.read('[Content_Types].xml').decode().replace(
        '<Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-'
        'officedocument.spreadsheetml.calcChain+xml"/>', '')
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            if item.filename == 'xl/calcChain.xml':
                continue
            if item.filename == sheet_path:
                data = ed.tobytes()
            elif item.filename == 'xl/workbook.xml':
                data = etree.tostring(wb_tree, xml_declaration=True, encoding='UTF-8', standalone=True)
            elif item.filename == '[Content_Types].xml':
                data = ct.encode()
            else:
                data = zin.read(item.filename)
            zout.writestr(item, data)
    zin.close()

    # verify what we wrote reads back correctly
    wchk = openpyxl.load_workbook(out, data_only=False)['DAILY']
    tcol = target_col
    back = round(sum(wchk.cell(row=r, column=tcol).value for r, v in fills.items()
                     if isinstance(v, (int, float))), 2)
    assert back == expected_total, f"Read-back total {back} != expected {expected_total}"
    vors = sum(1 for r, v in fills.items() if v == 'VOR')

    return {
        'date': str(report_date), 'column': TL,
        'wagons_filled': len(fills), 'expected_total_earnings': expected_total,
        'vor_count': vors, 'no_wagons': wagon_count,
        'parts': parts, 'workshop': workshop, 'tyres': tyres,
        'master_regs_not_in_daily_file': sorted(unmatched_master),
        'replaced': was_populated, 'previous_total': previous_total,
    }


if __name__ == '__main__':
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument('master'); ap.add_argument('daily')
    ap.add_argument('transactions', nargs='?', default=None)
    ap.add_argument('-o', '--out', required=True)
    a = ap.parse_args()
    try:
        result = fill_master(a.master, a.daily, a.transactions, a.out)
    except Exception as e:
        print(json.dumps({'status': 'failed', 'error': str(e)}), file=sys.stderr)
        sys.exit(1)
    print(json.dumps({'status': 'ok', **result}, indent=2))
