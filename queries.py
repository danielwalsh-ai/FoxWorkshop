"""Read queries for the Workshop & Maintenance dashboard + API."""
import datetime as dt
from db import get_conn

# Divisions that carry a monthly budget (others show spend only)
BUDGET_DIVISIONS = ['Fox Wagons', 'Leyland Wagons', 'J FISHER', 'NMS CIVIL',
                    'Tyres', 'J Fisher Plant', 'NMS Plant']


def _bounds(y, m):
    first = dt.date(y, m, 1)
    nm, ny = (1, y + 1) if m == 12 else (m + 1, y)
    return first, dt.date(ny, nm, 1)


def available_months():
    with get_conn() as c, c.cursor() as cur:
        cur.execute("SELECT DISTINCT to_char(report_date,'YYYY-MM') FROM transactions ORDER BY 1 DESC")
        return [r[0] for r in cur.fetchall()]


def overview(y, m):
    first, nxt = _bounds(y, m)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT division, ROUND(SUM(cost),2) FROM transactions
                       WHERE report_date >= %s AND report_date < %s
                       GROUP BY division ORDER BY 2 DESC""", (first, nxt))
        divs = cur.fetchall()
        cur.execute("""SELECT area, ROUND(SUM(cost),2) FROM transactions
                       WHERE report_date >= %s AND report_date < %s
                       GROUP BY area ORDER BY 2 DESC""", (first, nxt))
        areas = [{"area": a or 'REFERENCE MISSING', "total": float(t)} for a, t in cur.fetchall()]
        cur.execute("""SELECT report_date, ROUND(SUM(cost),2) FROM transactions
                       WHERE report_date >= %s AND report_date < %s
                       GROUP BY report_date ORDER BY report_date""", (first, nxt))
        daily = [{"date": d, "total": float(t)} for d, t in cur.fetchall()]
        cur.execute("SELECT division, budget FROM budgets WHERE year=%s AND month=%s", (y, m))
        budgets = {d: float(b) for d, b in cur.fetchall()}

    div_map = {d: float(t) for d, t in divs}
    total = round(sum(div_map.values()), 2)

    # budget tracker rows
    budget_rows = []
    for name in BUDGET_DIVISIONS:
        spent = div_map.get(name, 0.0)
        bud = budgets.get(name, 0.0)
        rem = bud - spent
        pct = (spent / bud * 100) if bud else 0
        budget_rows.append({"division": name, "budget": bud, "spent": round(spent, 2),
                            "remaining": round(rem, 2), "pct": round(pct, 1)})

    div_rows = [{"division": d, "total": float(t)} for d, t in divs]
    biggest_day = max(daily, key=lambda r: r["total"]) if daily else None

    return {
        "total": total,
        "damage": div_map.get("Damage", 0.0),
        "tyres": div_map.get("Tyres", 0.0),
        "capital": div_map.get("Capital", 0.0),
        "windscreen": div_map.get("Windscreen & Glass", 0.0),
        "div_rows": div_rows,
        "area_rows": areas,
        "budget_rows": budget_rows,
        "daily": daily,
        "days": len(daily),
        "biggest_day": biggest_day,
    }


def monthly_totals(n=18):
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT to_char(report_date,'YYYY-MM') ym, ROUND(SUM(cost),2)
                       FROM transactions GROUP BY ym ORDER BY ym DESC LIMIT %s""", (n,))
        rows = cur.fetchall()[::-1]
    return [{"ym": ym, "total": float(t)} for ym, t in rows]


def month_rows(y, m):
    """Every transaction line of the month — for rebuilding the workbook cover."""
    first, nxt = _bounds(y, m)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT report_date, division, area, plate, supplier, cost, vehicle_reg
                       FROM transactions WHERE report_date >= %s AND report_date < %s""",
                    (first, nxt))
        return cur.fetchall()


def day_tab_rows(d):
    """A single day's transactions, for the division tabs in the workbook."""
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT division, supplier, supplier_source_depot, system_no, supplier_pn,
                              part_name, cost, po_no, attached_order_no, attached_customer,
                              po_created_date, supply_type, item_count, goods_received,
                              target_depot, assigned_depot, supplier_ref, custom_ref, area
                       FROM transactions WHERE report_date = %s ORDER BY division""", (d,))
        cols = [dd[0] for dd in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def reg_year_split(report_date):
    """Vehicle spend by registration year — today + month-to-date + year-to-date
    (YTD runs from 1 Jan of the report year; PF request 20/07/2026).
    Returns (today_by_year, mtd_by_year, ytd_by_year) dicts keyed by year int."""
    import datetime as _dt
    from collections import defaultdict
    from classify import reg_year, TOP_SHEETS
    TOP = {s.strip() for s in TOP_SHEETS}
    first, _ = _bounds(report_date.year, report_date.month)
    jan1 = _dt.date(report_date.year, 1, 1)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT report_date, division, vehicle_reg, cost FROM transactions
                       WHERE report_date >= %s AND report_date <= %s""", (jan1, report_date))
        rows = cur.fetchall()
    today, mtd, ytd = defaultdict(float), defaultdict(float), defaultdict(float)
    for rd, division, reg, cost in rows:
        if (division or '').strip() not in TOP:
            continue
        reg = (reg or '').strip()
        if not reg:
            continue
        key = reg_year(reg) or 'other'   # 2021-2026 int, else 'other' (older/private)
        cost = float(cost or 0)
        ytd[key] += cost
        if rd >= first:
            mtd[key] += cost
        if rd == report_date:
            today[key] += cost
    return dict(today), dict(mtd), dict(ytd)


def spend_per_day(report_date):
    """Average spend per day by AREA and by AGE range (Paul Fox request).
    'Per day' = period total / number of active days in the period. YTD runs from
    1 Jan; MTD from the 1st of the report month. Area uses all-division spend
    (matches the Spend-by-Area breakdown); age uses the workshop top-sheets."""
    import datetime as _dt
    from collections import defaultdict
    from classify import reg_year, TOP_SHEETS
    TOP = {s.strip() for s in TOP_SHEETS}
    jan1 = _dt.date(report_date.year, 1, 1)
    first, _ = _bounds(report_date.year, report_date.month)
    # Cars and vans are excluded from the averages entirely (Daniel 20/08/2026):
    # they are not wagons, so they distort the per-wagon figures. Their spend
    # still appears in full on the cover/division/area-total pages for balance.
    EXCLUDE = {'VAN', 'CAR'}
    # The fleet lookup: reg -> area for every wagon we report on.
    reg_to_area = {}
    try:
        import vrm_lookup
        reg_to_area, _lk2, _lk3 = vrm_lookup.load_lookup()
    except Exception as _e:
        print(f"fleet lookup unavailable: {_e}")

    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT report_date, division, area, vehicle_reg, cost FROM transactions
                       WHERE report_date >= %s AND report_date <= %s""", (jan1, report_date))
        rows = cur.fetchall()
    ytd_days = len({r[0] for r in rows}) or 1
    mtd_days = len({r[0] for r in rows if r[0] >= first}) or 1
    area_y, area_m = defaultdict(float), defaultdict(float)
    age_y, age_m = defaultdict(float), defaultdict(float)
    # distinct wagons the average is drawn from THIS MONTH (MTD) — a reg counts
    # once per bucket. Non-vehicle areas (Consumables, Reference Missing, ...)
    # carry no reg, so their wagon count is 0.
    area_wag, age_wag = defaultdict(set), defaultdict(set)          # this month (MTD)
    for rd, div, area, reg, cost in rows:
        cst = float(cost or 0)
        a = area or 'REFERENCE MISSING'
        if a in EXCLUDE:                       # no cars/vans in the area averages
            continue
        area_y[a] += cst
        r = (reg or '').strip().upper().replace(' ', '')
        if rd >= first:
            area_m[a] += cst
            if r:
                area_wag[a].add(r)
        # Age range = genuine fleet wagons only: reg must be a real fleet vehicle
        # (in the lookup) and not a car/van. This also drops internal reference
        # codes (BT32DAM, TR13MOT ...) that only look like plates.
        if (div or '').strip() in TOP and r in reg_to_area and reg_to_area[r] not in EXCLUDE:
            k = reg_year(r) or 'other'
            age_y[k] += cst
            if rd >= first:
                age_m[k] += cst
                age_wag[k].add(r)
    per = lambda d, days: {k: v / days for k, v in d.items()}
    # Fleet denominators: EVERY wagon we report on in each section (from the
    # master lookup), not just the ones with spend — so the per-wagon figure is
    # a true average for the section, not just that section's spend. Cars and
    # vans are excluded; non-vehicle sections (Workshop, Consumables, ...) have
    # no fleet.
    fleet_area, fleet_age = {}, {}
    try:
        from collections import Counter
        fa_area, fa_age = Counter(), Counter()
        for rg, ar in reg_to_area.items():
            if ar in EXCLUDE:
                continue
            fa_area[ar] += 1
            fa_age[reg_year(rg) or 'other'] += 1
        fleet_area = dict(fa_area)
        fleet_age = dict(fa_age)
    except Exception as _e:
        print(f"fleet counts unavailable: {_e}")
    return {
        'ytd_days': ytd_days, 'mtd_days': mtd_days,
        'area_ytd': per(area_y, ytd_days), 'area_mtd': per(area_m, mtd_days),
        'age_ytd': per(age_y, ytd_days), 'age_mtd': per(age_m, mtd_days),
        'area_wagons': {k: len(v) for k, v in area_wag.items()},
        'age_wagons': {k: len(v) for k, v in age_wag.items()},
        'area_fleet': fleet_area, 'age_fleet': fleet_age,
    }


def parts_category_split(report_date):
    """Parts-category breakdown for the 2025 & 2026-plate trucks.

    Answers Paul's ask: "categorise the spend on the trucks into part
    categories, interested in the 25 plate spend."  Unlike reg_year_split this
    is ALL divisions (incl. Capital) — the biggest chunk of new-truck fit-out
    (safety/camera/radar systems, livery) is booked to Capital, so restricting
    to the workshop top-sheets would hide exactly what he wants to see.

    Returns a dict:
      ltd / mtd  -> {2025: {category: £}, 2026: {category: £}}
      total_ltd / total_mtd -> {2025: £, 2026: £}
      trucks_ltd / trucks_mtd -> {2025: n, 2026: n}
    """
    from collections import defaultdict
    from classify import reg_year
    from parts_category import categorise
    first, _ = _bounds(report_date.year, report_date.month)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT report_date, vehicle_reg, part_name, cost
                       FROM transactions WHERE report_date <= %s""", (report_date,))
        rows = cur.fetchall()
    import datetime as _dt
    jan1 = _dt.date(report_date.year, 1, 1)
    ltd = {2025: defaultdict(float), 2026: defaultdict(float)}
    mtd = {2025: defaultdict(float), 2026: defaultdict(float)}
    ytd = {2025: defaultdict(float), 2026: defaultdict(float)}
    tr_ltd = {2025: set(), 2026: set()}
    tr_mtd = {2025: set(), 2026: set()}
    for rd, reg, part, cost in rows:
        reg = (reg or '').strip()
        if not reg:
            continue
        y = reg_year(reg)
        if y not in (2025, 2026):
            continue
        cost = float(cost or 0)
        cat = categorise(part)
        ltd[y][cat] += cost
        tr_ltd[y].add(reg)
        if rd >= jan1:
            ytd[y][cat] += cost
        if rd >= first:
            mtd[y][cat] += cost
            tr_mtd[y].add(reg)
    return {
        'ytd': {y: dict(ytd[y]) for y in (2025, 2026)},
        'total_ytd': {y: round(sum(ytd[y].values()), 2) for y in (2025, 2026)},
        'ltd': {y: dict(ltd[y]) for y in (2025, 2026)},
        'mtd': {y: dict(mtd[y]) for y in (2025, 2026)},
        'total_ltd': {y: round(sum(ltd[y].values()), 2) for y in (2025, 2026)},
        'total_mtd': {y: round(sum(mtd[y].values()), 2) for y in (2025, 2026)},
        'trucks_ltd': {y: len(tr_ltd[y]) for y in (2025, 2026)},
        'trucks_mtd': {y: len(tr_mtd[y]) for y in (2025, 2026)},
    }


def recent_transactions(y, m, limit=60):
    first, nxt = _bounds(y, m)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT po_created_date, supplier, part_name, division, area,
                              vehicle_reg, cost, po_no
                       FROM transactions
                       WHERE report_date >= %s AND report_date < %s
                       ORDER BY cost DESC LIMIT %s""", (first, nxt, limit))
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, r)) for r in cur.fetchall()]


def hook_fleet():
    """Hook registrations from vehicle_master.xlsx (AREA == HOOKS), file order."""
    from openpyxl import load_workbook
    from pathlib import Path
    wb = load_workbook(Path(__file__).parent / 'vehicle_master.xlsx', read_only=True)
    ws = wb.active
    return [str(r[0]).strip().upper().replace(' ', '')
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r[0] and str(r[2] or '').strip().upper() == 'HOOKS']


def hook_split(report_date):
    """Per-registration hook spend for the month (area = HOOKS, top-sheet
    divisions only — mirrors exactly what feeds the HOOKS area row).
    Returns (fleet_regs, per_reg, unmatched, ytd) where
      per_reg   -> {reg: {report_date: £}}   (current month)
      unmatched -> {report_date: £}          (area HOOKS but no reg recorded)
      ytd       -> {reg: £} + {'_UNMATCHED': £}  (from 1 Jan)"""
    import datetime as _dt
    from collections import defaultdict
    from classify import TOP_SHEETS
    TOP = {s.strip() for s in TOP_SHEETS}
    fleet = hook_fleet()
    first, _ = _bounds(report_date.year, report_date.month)
    jan1 = _dt.date(report_date.year, 1, 1)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT report_date, division, vehicle_reg, cost FROM transactions
                       WHERE report_date >= %s AND report_date <= %s AND area = 'HOOKS'""",
                    (jan1, report_date))
        rows = cur.fetchall()
    per_reg = defaultdict(lambda: defaultdict(float))
    unmatched = defaultdict(float)
    ytd = defaultdict(float)
    for rd, division, reg, cost in rows:
        if (division or '').strip() not in TOP:
            continue
        cost = float(cost or 0)
        reg = (reg or '').strip().upper().replace(' ', '')
        if reg:
            ytd[reg] += cost
            if rd >= first:
                per_reg[reg][rd] += cost
        else:
            ytd['_UNMATCHED'] += cost
            if rd >= first:
                unmatched[rd] += cost
    return fleet, {k: dict(v) for k, v in per_reg.items()}, dict(unmatched), dict(ytd)


FN_GROUPS = {'J FISHER': ('J FISHER', 'J Fisher Plant'),
             'NMS': ('NMS CIVIL', 'NMS Plant')}
FN_BANDS = ['2021', '2022', '2023', '2024', '2025', '2026',
            'Older / private plates', 'Unregistered / Plant']


def fisher_nms_split(report_date):
    """Age-band x parts-category matrices for J Fisher (trucks + plant) and
    NMS (civil + plant) — PF request 16/07/2026, 'same way as pages 3-4'.
    Registered vehicles band by plate year (2021-2026, else Older/private);
    lines with no reg (plant kit, stock, consumables) -> 'Unregistered / Plant'
    per DW instruction 18/07 (no plant-age lookup).
    Returns {group: {'mtd': {cat: {band: £}}, 'today': {cat: {band: £}},
                     'mtd_total': £, 'today_total': £}}"""
    from collections import defaultdict
    from classify import reg_year, extract_reg
    from parts_category import categorise
    import datetime as _dt
    first, _ = _bounds(report_date.year, report_date.month)
    jan1 = _dt.date(report_date.year, 1, 1)
    divs = tuple(d for pair in FN_GROUPS.values() for d in pair)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT report_date, division, vehicle_reg, part_name, cost
                       FROM transactions
                       WHERE report_date >= %s AND report_date <= %s
                         AND division IN %s""", (jan1, report_date, divs))
        rows = cur.fetchall()
    out = {g: {'mtd': defaultdict(lambda: defaultdict(float)),
               'today': defaultdict(lambda: defaultdict(float)),
               'ytd': defaultdict(lambda: defaultdict(float)),
               'mtd_total': 0.0, 'today_total': 0.0, 'ytd_total': 0.0} for g in FN_GROUPS}
    div_to_group = {d.strip(): g for g, pair in FN_GROUPS.items() for d in pair}
    for rd, division, reg, part, cost in rows:
        g = div_to_group.get((division or '').strip())
        if not g:
            continue
        cost = float(cost or 0)
        reg = (reg or '').strip()
        if not reg:
            band = 'Unregistered / Plant'
        else:
            y = reg_year(reg)
            band = str(y) if y else 'Older / private plates'
        cat = categorise(part)
        out[g]['ytd'][cat][band] += cost
        out[g]['ytd_total'] += cost
        if rd >= first:
            out[g]['mtd'][cat][band] += cost
            out[g]['mtd_total'] += cost
        if rd == report_date:
            out[g]['today'][cat][band] += cost
            out[g]['today_total'] += cost
    for g in out:
        out[g]['mtd'] = {k: dict(v) for k, v in out[g]['mtd'].items()}
        out[g]['today'] = {k: dict(v) for k, v in out[g]['today'].items()}
        out[g]['ytd'] = {k: dict(v) for k, v in out[g]['ytd'].items()}
        out[g]['mtd_total'] = round(out[g]['mtd_total'], 2)
        out[g]['today_total'] = round(out[g]['today_total'], 2)
        out[g]['ytd_total'] = round(out[g]['ytd_total'], 2)
    return out


def ytd_rollups(report_date):
    """Year-to-date (from 1 Jan) spend by division and by area — top sheets.
    Returns {'division': {name: £}, 'area': {name: £}}."""
    import datetime as _dt
    from collections import defaultdict
    from classify import TOP_SHEETS
    TOP = {s.strip() for s in TOP_SHEETS}
    jan1 = _dt.date(report_date.year, 1, 1)
    with get_conn() as c, c.cursor() as cur:
        cur.execute("""SELECT division, area, cost FROM transactions
                       WHERE report_date >= %s AND report_date <= %s""",
                    (jan1, report_date))
        rows = cur.fetchall()
    div, area = defaultdict(float), defaultdict(float)
    for d, a, cost in rows:
        d = (d or '').strip()
        if d not in TOP:
            continue
        cost = float(cost or 0)
        div[d] += cost
        if a:
            area[a.strip()] += cost
    return {'division': dict(div), 'area': dict(area)}
